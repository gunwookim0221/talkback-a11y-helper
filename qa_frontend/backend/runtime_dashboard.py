from __future__ import annotations

import ast
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from .paths import OUTPUT_DIR


SCENARIO_RE = re.compile(r"scenario='([^']+)'|scenario=([A-Za-z0-9_]+)|scenario_id='([^']+)'|scenario_id=([A-Za-z0-9_]+)")
STEP_RE = re.compile(r"step=(\d+)")
TOTAL_STEPS_RE = re.compile(r"total_steps=(\d+)")
SUMMARY_SAVE_EXCEL_RE = re.compile(r"save_excel_count=(\d+)")
ENABLED_IDS_RE = re.compile(r"enabled_ids=(\[[^\]]*\])")
FOCUS_PACKAGE_RE = re.compile(r"packageName': '([^']+)'|packageName=\"([^\"]+)\"|packageName=([^\s,]+)")
FOCUS_LABEL_RE = re.compile(r"talkbackLabel': '([^']+)'|focus_label='([^']*)'|visible='([^']*)'")
FINAL_RESULT_RE = re.compile(r"final_result='([^']+)'")
TRAVERSAL_RESULT_RE = re.compile(r"traversal_result='([^']+)'")
STOP_REASON_RE = re.compile(r"reason='([^']+)'")
POPUP_RESULT_RE = re.compile(r"\[QA_FRONTEND\]\[preflight\]\[popup\].*result='([^']*)'")
PREFLIGHT_FINAL_RE = re.compile(r"\[QA_FRONTEND\]\[preflight\] final_result='([^']*)'")
PRE_STATUS_RE = re.compile(r"\[QA_FRONTEND\]\[preflight\]\[(adb|helper)\] status='([^']*)'")
RUN_START_RE = re.compile(r"\[QA_FRONTEND\] start mode='([^']*)'.*launch_mode='([^']*)'(?:.*language_mode='([^']*)')?")
LANGUAGE_RE = re.compile(r"\[QA_FRONTEND\]\[language\].*language_mode='([^']*)'.*device_locale='([^']*)'")
SAVED_EXCEL_RE = re.compile(r"saved excel:\s+output/(?P<filename>[^/\s]+\.xlsx)", re.IGNORECASE)
ACCESSIBILITY_PASS_MISMATCH_TYPES = {
    "EXACT_MATCH",
    "NORMALIZED_MATCH",
    "PARTIAL_MATCH",
    "REPRESENTATIVE_CONTEXT",
}
SCENARIO_HARD_MISMATCH_TYPES = {
    "EMPTY_VISIBLE",
    "EMPTY_SPEECH",
    "LABEL_MISMATCH",
}
TRANSIENT_PRESENTATION_FAILURE_REASONS = {
    "move_failed",
    "repeat_no_progress",
    "terminal_not_handled",
    "no_unvisited_local_tab",
    "viewport_exhausted",
}


def _availability_signal() -> dict[str, object]:
    return {
        "pre_nav_action": "",
        "pre_nav_target": "",
        "pre_nav_failed": False,
        "pre_nav_failure_detail": "",
        "anchor_insufficient": False,
        "device_inventory_labels": "",
        "device_target_not_visible": False,
        "inventory_signature_unchanged": False,
        "summary_save_excel_count": None,
    }


def build_runtime_dashboard(
    *,
    status: dict[str, object],
    log_path: Path | None,
    scenario_ids: list[str] | None = None,
    parsed_log: dict[str, object] | None = None,
) -> dict[str, object]:
    started_at = _string_or_none(status.get("started_at"))
    dashboard = _empty_dashboard(status=status, started_at=started_at, scenario_ids=scenario_ids or [])
    if not log_path or not log_path.exists():
        return dashboard

    try:
        if parsed_log is not None:
            parsed = parsed_log
        else:
            log_text = log_path.read_text(encoding="utf-8", errors="replace")
            validation_failed_scenarios, validation_warning_scenarios = extract_validation_scenario_evidence_from_log(log_text)
            parsed = parse_runtime_log(
                log_text,
                scenario_ids=scenario_ids or [],
                validation_failed_scenarios=validation_failed_scenarios,
                validation_warning_scenarios=validation_warning_scenarios,
            )
            parsed["log_size"] = log_path.stat().st_size
        dashboard.update(parsed)
    except Exception as exc:
        dashboard["parse_error"] = str(exc)

    dashboard["elapsed_seconds"] = _elapsed_seconds(started_at, _string_or_none(status.get("finished_at")))
    dashboard["run_id"] = status.get("run_id")
    dashboard["mode"] = status.get("mode")
    dashboard["launch_mode"] = status.get("launch_mode")
    dashboard["language_mode"] = status.get("language_mode")
    dashboard["device_locale"] = status.get("device_locale") or dashboard.get("device_locale")
    dashboard["state"] = status.get("state")
    dashboard["preflight_state"] = status.get("preflight_state") or dashboard.get("preflight_state")
    dashboard["popup_result"] = status.get("popup_result") or dashboard.get("popup_result")
    dashboard["helper_status"] = status.get("helper_state") or dashboard.get("helper_status")
    return dashboard


def parse_runtime_log(
    log_text: str,
    *,
    scenario_ids: list[str] | None = None,
    validation_failed_scenarios: set[str] | None = None,
    validation_warning_scenarios: set[str] | None = None,
) -> dict[str, object]:
    lines = log_text.splitlines()
    selected_ids = list(scenario_ids or []) or _extract_enabled_ids(log_text)
    selected_filter = set(selected_ids)
    progress = {scenario_id: {"id": scenario_id, "status": "queued", "steps": 0} for scenario_id in selected_ids}
    step_end_counts: dict[str, int] = {}
    summary_steps: dict[str, int] = {}
    summary_save_excel_counts: dict[str, int] = {}
    entry_success_scenarios: set[str] = set()
    special_state_detected_at: dict[str, int] = {}
    special_state_handled_at: dict[str, int] = {}
    soft_entry_evidence_scenarios: set[str] = set()
    summary_scenarios: set[str] = set()
    global_nav_start_gate_passed: set[str] = set()
    global_nav_menu_reached: set[str] = set()
    global_nav_terminal: set[str] = set()
    events: list[dict[str, object]] = []
    current_scenario = None
    current_availability_scenario = None
    current_step = None
    total_step_count = 0
    overlay_count = 0
    save_excel_count = 0
    hard_failed_scenarios: set[str] = set(validation_failed_scenarios or set())
    warning_scenarios: set[str] = set(validation_warning_scenarios or set())
    last_focus_label = None
    last_focus_package = None
    stop_reason = None
    traversal_result = None
    popup_result = None
    preflight_state = None
    adb_status = None
    helper_status = None
    mode = None
    launch_mode = None
    language_mode = None
    device_locale = None
    availability_signals: dict[str, dict[str, object]] = {}
    fatal_or_crash_like_scenarios: set[str] = set()

    for index, line in enumerate(lines):
        raw_scenario = _extract_scenario(line)
        scenario = raw_scenario if raw_scenario and (not selected_filter or raw_scenario in selected_filter) else None
        step = _extract_step(line)
        if scenario:
            current_scenario = scenario
            progress.setdefault(scenario, {"id": scenario, "status": "queued", "steps": 0})
            availability_signals.setdefault(scenario, _availability_signal())
        signal_scenario = scenario or current_availability_scenario or current_scenario
        if (
            not scenario
            and "[SCENARIO][pre_nav] step=" in line
            and (current_scenario is None or current_scenario in summary_scenarios)
        ):
            next_scenario = _next_unfinished_selected_scenario(selected_ids, summary_scenarios)
            if next_scenario:
                signal_scenario = next_scenario
                current_availability_scenario = next_scenario
        if not signal_scenario and len(selected_ids) == 1 and _line_has_availability_signal(line):
            signal_scenario = selected_ids[0]
            current_availability_scenario = signal_scenario
        if signal_scenario and (not selected_filter or signal_scenario in selected_filter):
            availability_signals.setdefault(signal_scenario, _availability_signal())

        if signal_scenario and "[SCENARIO][pre_nav] step=" in line:
            action_match = re.search(r"action=([^\s]+)", line)
            target_match = re.search(r"target='([^']*)'", line)
            signal = availability_signals.setdefault(signal_scenario, _availability_signal())
            signal["pre_nav_action"] = action_match.group(1) if action_match else ""
            signal["pre_nav_target"] = target_match.group(1) if target_match else ""
        if signal_scenario and "[SCENARIO][pre_nav] failed" in line:
            signal = availability_signals.setdefault(signal_scenario, _availability_signal())
            signal["pre_nav_failed"] = True
            if "detail=" in line:
                signal["pre_nav_failure_detail"] = line
        if signal_scenario and "[DEVICE_ENTRY][inventory]" in line:
            labels_match = re.search(r"labels='([^']*)'", line)
            signal = availability_signals.setdefault(signal_scenario, _availability_signal())
            signal["device_inventory_labels"] = labels_match.group(1) if labels_match else ""
        if signal_scenario and "target_not_visible" in line:
            availability_signals.setdefault(signal_scenario, _availability_signal())["device_target_not_visible"] = True
        if signal_scenario and "inventory_signature_changed=false" in line:
            availability_signals.setdefault(signal_scenario, _availability_signal())["inventory_signature_unchanged"] = True
        if signal_scenario and "insufficient_new_screen_evidence" in line:
            availability_signals.setdefault(signal_scenario, _availability_signal())["anchor_insufficient"] = True

        if "[GLOBAL_NAV][start_gate] passed" in line and scenario:
            global_nav_start_gate_passed.add(scenario)
        if scenario and _line_reaches_menu_tab(line):
            global_nav_menu_reached.add(scenario)
        if scenario and "[SCENARIO][entry_contract] success" in line:
            entry_success_scenarios.add(scenario)
        if scenario and _is_special_state_detected(line):
            special_state_detected_at.setdefault(scenario, index)
        if scenario and _is_special_state_handled_contract(line):
            special_state_handled_at.setdefault(scenario, index)
        if scenario and "[ENTRY][post_open_identity]" in line and _has_plugin_entry_identity_evidence(scenario, line):
            soft_entry_evidence_scenarios.add(scenario)
        if scenario and "[SCENARIO][entry_contract]" in line and " fail" in line.lower():
            if "post_open_verify_miss" in line and scenario in soft_entry_evidence_scenarios:
                warning_scenarios.add(scenario)
                progress[scenario]["status"] = "warning"
                _add_event(events, index, "scenario_warning", line, scenario=scenario)
            else:
                hard_failed_scenarios.add(scenario)
                progress[scenario]["status"] = "failed"
                _add_event(events, index, "scenario_failed", line, scenario=scenario)

        if "[QA_FRONTEND] start" in line:
            run_match = RUN_START_RE.search(line)
            if run_match:
                mode = run_match.group(1)
                launch_mode = run_match.group(2)
                language_mode = run_match.group(3)
                _add_event(events, index, "run_started", line, scenario=None)
        elif "[QA_FRONTEND][language]" in line:
            language_match = LANGUAGE_RE.search(line)
            if language_match:
                language_mode = language_match.group(1)
                device_locale = language_match.group(2) or None
                _add_event(events, index, "language_ready", line, scenario=None)
        elif "[QA_FRONTEND][preflight][popup]" in line:
            popup_match = POPUP_RESULT_RE.search(line)
            if popup_match:
                popup_result = popup_match.group(1)
                _add_event(events, index, f"popup_{popup_result or 'observed'}", line, scenario=None)
        elif "[QA_FRONTEND][preflight]" in line:
            final_match = PREFLIGHT_FINAL_RE.search(line)
            if final_match:
                preflight_state = final_match.group(1)
                _add_event(events, index, f"preflight_{preflight_state or 'unknown'}", line, scenario=None)
            status_match = PRE_STATUS_RE.search(line)
            if status_match:
                if status_match.group(1) == "adb":
                    adb_status = status_match.group(2)
                if status_match.group(1) == "helper":
                    helper_status = status_match.group(2)
        elif "[STEP] START" in line:
            current_step = step
            total_step_count += 1
            if scenario:
                progress[scenario]["status"] = "running"
                progress[scenario]["steps"] = max(int(progress[scenario].get("steps", 0)), step + 1 if step is not None else 0)
                _add_event(events, index, "scenario_running", line, scenario=scenario)
        elif "[STEP] END" in line:
            current_step = step
            if scenario:
                step_end_counts[scenario] = step_end_counts.get(scenario, 0) + 1
                progress[scenario]["steps"] = max(int(progress[scenario].get("steps", 0)), step + 1 if step is not None else 0)
            last_focus_label = _extract_focus_label(line) or last_focus_label
        elif "[OVERLAY]" in line:
            overlay_count += 1
        elif "[SAVE] saved excel:" in line:
            save_excel_count += 1
            _add_event(events, index, "save_excel", line, scenario=current_scenario)
        elif "[STOP][eval]" in line:
            final_result = _extract_first(FINAL_RESULT_RE, line)
            event_traversal_result = _extract_first(TRAVERSAL_RESULT_RE, line)
            event_stop_reason = _extract_first(STOP_REASON_RE, line)
            traversal_result = event_traversal_result or traversal_result
            stop_reason = event_stop_reason or stop_reason
            if scenario:
                # These are scenario-scoped runner facts.  Keeping them on the
                # progress item prevents the final stop of a full run from being
                # copied into every scenario summary.
                if event_traversal_result:
                    progress[scenario]["traversal_result"] = event_traversal_result
                if event_stop_reason:
                    progress[scenario]["stop_reason"] = event_stop_reason
            if scenario and _is_global_nav_terminal(line, scenario):
                global_nav_terminal.add(scenario)
                progress[scenario]["status"] = "passed"
                _add_event(events, index, "traversal_terminal", line, scenario=scenario)
            elif scenario and str(final_result).upper() == "FAIL":
                benign_reasons = {
                    "repeat_no_progress",
                    "viewport_exhausted",
                    "terminal_reached",
                    "end_of_content",
                    "no_unvisited_local_tab"
                }
                if stop_reason and stop_reason in benign_reasons:
                    if scenario in entry_success_scenarios:
                        warning_scenarios.add(scenario)
                        if progress[scenario].get("status") not in {"failed"}:
                            progress[scenario]["status"] = "warning"
                    _add_event(events, index, "traversal_terminal", line, scenario=scenario)
                else:
                    warning_scenarios.add(scenario)
                    if progress[scenario].get("status") not in {"failed"}:
                        progress[scenario]["status"] = "warning"
                    _add_event(events, index, "scenario_warning", line, scenario=scenario)
            elif scenario and "decision='stop'" in line:
                _add_event(events, index, "traversal_terminal", line, scenario=scenario)
        elif "[TAB][select] stabilization failed" in line or "no_bottom_nav_candidates" in line:
            if scenario:
                hard_failed_scenarios.add(scenario)
                progress[scenario]["status"] = "failed"
                _add_event(events, index, "scenario_failed", line, scenario=scenario)
        elif "[PERF][scenario_summary]" in line:
            if scenario:
                summary_scenarios.add(scenario)
                if current_availability_scenario == scenario:
                    current_availability_scenario = None
                total_steps = _extract_total_steps(line)
                if total_steps is not None:
                    summary_steps[scenario] = total_steps
                    progress[scenario]["steps"] = total_steps
                scenario_save_excel_count = _extract_summary_save_excel_count(line)
                if scenario_save_excel_count is not None:
                    summary_save_excel_counts[scenario] = scenario_save_excel_count
                    availability_signals.setdefault(scenario, _availability_signal())["summary_save_excel_count"] = scenario_save_excel_count
                if progress[scenario].get("status") not in {"failed", "warning"}:
                    progress[scenario]["status"] = "passed"
                _add_event(events, index, "scenario_completed", line, scenario=scenario)
        elif "[FATAL]" in line or "Traceback" in line or "Exception" in line:
            if scenario:
                fatal_or_crash_like_scenarios.add(scenario)
                hard_failed_scenarios.add(scenario)
                progress[scenario]["status"] = "failed"
                _add_event(events, index, "scenario_failed", line, scenario=scenario)
        elif "APP_TERMINATED" in line:
            if signal_scenario:
                fatal_or_crash_like_scenarios.add(signal_scenario)
        elif "[MAIN] skip disabled scenario_id=" in line:
            skipped = _extract_skip_scenario(line)
            if skipped in progress:
                progress[skipped]["status"] = "skipped"
        elif "[FOCUS_RESULT" in line or "packageName" in line:
            last_focus_package = _extract_focus_package(line) or last_focus_package
            last_focus_label = _extract_focus_label(line) or last_focus_label
        elif "[QA_FRONTEND][run] stop_requested=true" in line:
            _add_event(events, index, "stop_requested", line, scenario=current_scenario)

    special_state_handled_scenarios = {
        scenario_id
        for scenario_id, handled_index in special_state_handled_at.items()
        if (detected_index := special_state_detected_at.get(scenario_id)) is not None
        and detected_index < handled_index
    }
    entry_success_scenarios.update(special_state_handled_scenarios)

    for scenario_id, item in progress.items():
        if scenario_id in special_state_handled_scenarios:
            item["entry_contract_status"] = "handled"
            item["special_state_handled"] = True
        if scenario_id in summary_steps:
            item["steps"] = summary_steps[scenario_id]
        elif scenario_id in step_end_counts:
            item["steps"] = step_end_counts[scenario_id]
        elif item.get("steps") in {None, 0} and current_scenario == scenario_id and current_step is not None:
            item["steps"] = current_step + 1
        if _is_completed_global_nav_terminal(
            scenario_id,
            start_gate_passed=global_nav_start_gate_passed,
            menu_reached=global_nav_menu_reached,
            terminal=global_nav_terminal,
        ):
            hard_failed_scenarios.discard(scenario_id)
            warning_scenarios.discard(scenario_id)
            item["status"] = "passed"
            continue
        if scenario_id in hard_failed_scenarios:
            item["status"] = "failed"
            continue
        if scenario_id in warning_scenarios and scenario_id in soft_entry_evidence_scenarios:
            item["status"] = "warning"
            continue
        has_summary = scenario_id in summary_scenarios
        has_rows = step_end_counts.get(scenario_id, 0) > 0 or int(item.get("steps") or 0) > 0
        has_entry = scenario_id in entry_success_scenarios or scenario_id.startswith("global_nav")
        if item.get("status") == "skipped":
            continue
        if not has_summary or not has_rows:
            if has_entry or item.get("status") not in {"queued", "running"}:
                item["status"] = "failed"
            continue
        if scenario_id in warning_scenarios:
            item["status"] = "warning"
        elif _is_availability_candidate(
            scenario_id,
            item,
            has_entry=has_entry,
            has_summary=has_summary,
            summary_save_excel_count=summary_save_excel_counts.get(scenario_id),
            signal=availability_signals.get(scenario_id, {}),
            fatal_or_crash_like=scenario_id in fatal_or_crash_like_scenarios,
        ):
            availability = _classify_availability_signal(availability_signals.get(scenario_id, {}))
            item.update(availability)
            item["status"] = str(availability.get("status") or "not_available_candidate")
        elif item.get("status") not in {"failed", "warning"}:
            item["status"] = "passed"

    passed_count = len([item for item in progress.values() if item["status"] == "passed"])
    warning_count = len([item for item in progress.values() if item["status"] == "warning"])
    failed_count = len([item for item in progress.values() if item["status"] == "failed"])
    not_available_count = len([item for item in progress.values() if item["status"] == "not_available"])
    not_available_candidate_count = len([item for item in progress.values() if item["status"] == "not_available_candidate"])
    no_target_candidate_count = len([item for item in progress.values() if item["status"] == "no_target_candidate"])
    availability_candidate_count = not_available_count + not_available_candidate_count + no_target_candidate_count
    executed_count = passed_count + warning_count + failed_count
    completed_count = len([item for item in progress.values() if item["status"] in {"passed", "warning", "failed", "skipped", "not_available", "not_available_candidate", "no_target_candidate"}])
    return {
        "mode": mode,
        "launch_mode": launch_mode,
        "language_mode": language_mode,
        "device_locale": device_locale,
        "current_scenario": current_scenario,
        "passed_scenarios": passed_count,
        "special_state_handled_scenarios": len(special_state_handled_scenarios),
        "warning_scenarios": warning_count,
        "completed_scenarios": passed_count + warning_count,
        "executed_scenarios": executed_count,
        "not_available_scenarios": not_available_count,
        "not_available_candidate_scenarios": not_available_candidate_count,
        "no_target_candidate_scenarios": no_target_candidate_count,
        "availability_candidate_scenarios": availability_candidate_count,
        "remaining_scenarios": len([item for item in progress.values() if item["status"] in {"queued", "running"}]),
        "failed_scenarios": failed_count,
        "scenario_progress": list(progress.values()),
        "current_step": current_step,
        "total_step_count": total_step_count,
        "overlay_count": overlay_count,
        "save_excel_count": save_excel_count,
        "popup_result": popup_result,
        "preflight_state": preflight_state,
        "helper_status": helper_status,
        "adb_status": adb_status,
        "last_focus_label": last_focus_label,
        "last_focus_package": last_focus_package,
        "stop_reason": stop_reason,
        "traversal_result": traversal_result,
        "event_feed": events[-30:],
        "completed_or_terminal_scenarios": completed_count,
        "parse_error": None,
    }


def _empty_dashboard(*, status: dict[str, object], started_at: str | None, scenario_ids: list[str]) -> dict[str, object]:
    return {
        "run_id": status.get("run_id"),
        "mode": status.get("mode"),
        "launch_mode": status.get("launch_mode"),
        "language_mode": status.get("language_mode"),
        "device_locale": status.get("device_locale"),
        "state": status.get("state"),
        "started_at": started_at,
        "elapsed_seconds": _elapsed_seconds(started_at, _string_or_none(status.get("finished_at"))),
        "current_scenario": None,
        "completed_scenarios": 0,
        "executed_scenarios": 0,
        "not_available_scenarios": 0,
        "not_available_candidate_scenarios": 0,
        "no_target_candidate_scenarios": 0,
        "availability_candidate_scenarios": 0,
        "remaining_scenarios": len(scenario_ids),
        "failed_scenarios": 0,
        "scenario_progress": [{"id": item, "status": "queued", "steps": 0} for item in scenario_ids],
        "passed_scenarios": 0,
        "special_state_handled_scenarios": 0,
        "warning_scenarios": 0,
        "current_step": None,
        "total_step_count": 0,
        "overlay_count": 0,
        "save_excel_count": 0,
        "popup_result": status.get("popup_result"),
        "preflight_state": status.get("preflight_state"),
        "helper_status": status.get("helper_state"),
        "adb_status": None,
        "last_focus_label": None,
        "last_focus_package": None,
        "stop_reason": None,
        "traversal_result": None,
        "event_feed": [],
        "log_size": 0,
        "parse_error": None,
    }


def _extract_summary_save_excel_count(line: str) -> int | None:
    match = SUMMARY_SAVE_EXCEL_RE.search(line)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _line_has_availability_signal(line: str) -> bool:
    return any(
        marker in line
        for marker in (
            "[SCENARIO][pre_nav]",
            "[DEVICE_ENTRY][inventory]",
            "target_not_visible",
            "inventory_signature_changed=false",
            "insufficient_new_screen_evidence",
        )
    )


def _is_special_state_detected(line: str) -> bool:
    return (
        "[SCENARIO][special_state] detected" in line
        and "handling='back_after_read'" in line
    )


def _is_special_state_handled_contract(line: str) -> bool:
    return (
        "[SCENARIO][entry_contract] handled" in line
        and "reason='special_state_handled'" in line
        and "detail='onboarding_back_exit_recovered'" in line
    )


def _next_unfinished_selected_scenario(selected_ids: list[str], finished: set[str]) -> str | None:
    for scenario_id in selected_ids:
        if scenario_id not in finished:
            return scenario_id
    return None


def _is_availability_candidate(
    scenario_id: str,
    item: dict[str, object],
    *,
    has_entry: bool,
    has_summary: bool,
    summary_save_excel_count: int | None,
    signal: dict[str, object],
    fatal_or_crash_like: bool,
) -> bool:
    if fatal_or_crash_like or has_entry or not has_summary:
        return False
    if int(item.get("steps") or 0) > 1:
        return False
    if summary_save_excel_count not in {0, None}:
        return False
    pre_nav_action = str(signal.get("pre_nav_action") or "")
    if pre_nav_action in {"enter_device_card_plugin", "enter_safe_favorite_card", "xml_scroll_search_tap", "scrolltouch"}:
        return True
    return bool(signal.get("anchor_insufficient") or signal.get("pre_nav_failed"))


def _classify_availability_signal(signal: dict[str, object]) -> dict[str, object]:
    action = str(signal.get("pre_nav_action") or "")
    target = str(signal.get("pre_nav_target") or "").strip()
    inventory = str(signal.get("device_inventory_labels") or "").strip()
    device_target_missing = bool(signal.get("device_target_not_visible") or signal.get("inventory_signature_unchanged"))
    if action in {"enter_device_card_plugin", "enter_safe_favorite_card"} and (inventory or device_target_missing or action == "enter_safe_favorite_card"):
        reason = "target device card not found"
        if action == "enter_safe_favorite_card":
            reason = "optional Safe card not found"
        if inventory:
            reason = f"{reason}; inventory only {inventory}"
        return {
            "status": "not_available",
            "availability_status": "NOT_AVAILABLE",
            "availability_confidence": "high",
            "availability_reason": reason,
            "availability_target": target,
        }
    if action in {"xml_scroll_search_tap", "scrolltouch"}:
        reason = "target plugin/card anchor not found"
        if bool(signal.get("anchor_insufficient")):
            reason = "anchor insufficient new-screen evidence"
        elif bool(signal.get("pre_nav_failed")):
            reason = "pre-navigation target search failed"
        return {
            "status": "no_target_candidate",
            "availability_status": "NO_TARGET_CANDIDATE",
            "availability_confidence": "medium",
            "availability_reason": reason,
            "availability_target": target,
        }
    return {
        "status": "not_available_candidate",
        "availability_status": "NOT_AVAILABLE_CANDIDATE",
        "availability_confidence": "low",
        "availability_reason": "scenario ended before entry contract success",
        "availability_target": target,
    }


def _extract_enabled_ids(log_text: str) -> list[str]:
    match = ENABLED_IDS_RE.search(log_text)
    if not match:
        return []
    try:
        value = ast.literal_eval(match.group(1))
    except (SyntaxError, ValueError):
        return []
    if not isinstance(value, list):
        return []
    return [str(item) for item in value]


def extract_validation_scenario_evidence_from_log(log_text: str) -> tuple[set[str], set[str]]:
    filename = _extract_saved_excel_filename(log_text)
    if not filename:
        return set(), set()
    return extract_validation_scenario_evidence_from_xlsx(OUTPUT_DIR / filename)


def extract_validation_scenario_evidence_from_xlsx(path: Path) -> tuple[set[str], set[str]]:
    if not path.exists():
        return set(), set()
    try:
        import openpyxl
    except ImportError:
        return set(), set()
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "result" not in workbook.sheetnames:
            return set(), set()
        sheet = workbook["result"]
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        scenario_col = headers.index("scenario_id") + 1
        result_col = headers.index("final_result") + 1
        failure_col = headers.index("failure_reason") + 1 if "failure_reason" in headers else None
        mismatch_col = headers.index("mismatch_type") + 1 if "mismatch_type" in headers else None
    except (OSError, ValueError, KeyError):
        return set(), set()

    failed: set[str] = set()
    warning: set[str] = set()
    try:
        for row in range(2, sheet.max_row + 1):
            result = str(sheet.cell(row, result_col).value or "").strip().upper()
            scenario = str(sheet.cell(row, scenario_col).value or "").strip()
            if not scenario:
                continue
            if result == "WARN":
                warning.add(scenario)
            elif result == "FAIL":
                failure_reason = str(sheet.cell(row, failure_col).value or "").strip() if failure_col else ""
                mismatch_type = str(sheet.cell(row, mismatch_col).value or "").strip().upper() if mismatch_col else ""
                normalized_failure = failure_reason.lower()
                if mismatch_type in SCENARIO_HARD_MISMATCH_TYPES:
                    failed.add(scenario)
                elif (
                    mismatch_type in ACCESSIBILITY_PASS_MISMATCH_TYPES
                    and normalized_failure in TRANSIENT_PRESENTATION_FAILURE_REASONS
                ):
                    warning.add(scenario)
                elif failure_reason:
                    failed.add(scenario)
                else:
                    warning.add(scenario)
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    return failed, warning


def _extract_saved_excel_filename(log_text: str) -> str | None:
    matches = SAVED_EXCEL_RE.findall(log_text)
    return matches[-1] if matches else None


def _extract_scenario(line: str) -> str | None:
    match = SCENARIO_RE.search(line)
    if not match:
        return None
    return match.group(1) or match.group(2) or match.group(3) or match.group(4)


def _extract_skip_scenario(line: str) -> str | None:
    match = re.search(r"scenario_id='([^']+)'", line)
    return match.group(1) if match else None


def _extract_step(line: str) -> int | None:
    match = STEP_RE.search(line)
    if not match:
        return None
    return int(match.group(1))


def _extract_total_steps(line: str) -> int | None:
    match = TOTAL_STEPS_RE.search(line)
    if not match:
        return None
    return int(match.group(1))


def _extract_focus_package(line: str) -> str | None:
    match = FOCUS_PACKAGE_RE.search(line)
    if not match:
        return None
    return match.group(1) or match.group(2) or match.group(3)


def _extract_focus_label(line: str) -> str | None:
    match = FOCUS_LABEL_RE.search(line)
    if not match:
        return None
    return match.group(1) or match.group(2) or match.group(3)


def _line_reaches_menu_tab(line: str) -> bool:
    return "Menu, Tab 5 of 5" in line or "menu tab 5 of 5" in line.lower()


def _is_global_nav_terminal(line: str, scenario: str) -> bool:
    return (
        scenario == "global_nav_main"
        or "scenario_type='global_nav'" in line
        or "is_global_nav=true" in line
    ) and "reason='smart_nav_terminal'" in line


def _is_completed_global_nav_terminal(
    scenario: str,
    *,
    start_gate_passed: set[str],
    menu_reached: set[str],
    terminal: set[str],
) -> bool:
    return (
        scenario in terminal
        and scenario in menu_reached
        and (scenario in start_gate_passed or scenario == "global_nav_main")
    )


def _has_plugin_entry_identity_evidence(scenario: str, line: str) -> bool:
    lowered = line.lower()
    evidence_tokens = {
        "life_find_plugin": (
            "current location",
            "my devices",
            "offline",
            "com.samsung.android.plugin.fme",
        ),
        "life_video_plugin": (
            "live view",
            "daily clips",
            "home camera",
            "홈카메라",
        ),
    }.get(scenario, ())
    if not evidence_tokens:
        return False
    negative_tokens = ("location qr code", "change location", "장소 qr 코드", "장소 변경")
    has_evidence = any(token in lowered for token in evidence_tokens)
    has_negative_only = any(token in lowered for token in negative_tokens) and not has_evidence
    return bool(has_evidence and not has_negative_only)


def _extract_first(pattern: re.Pattern[str], line: str) -> str | None:
    match = pattern.search(line)
    return match.group(1) if match else None


def _add_event(events: list[dict[str, object]], line_index: int, event_type: str, line: str, *, scenario: str | None) -> None:
    events.append(
        {
            "line": line_index + 1,
            "type": event_type,
            "scenario": scenario,
            "message": _trim(line),
        }
    )


def _trim(value: str, limit: int = 240) -> str:
    return value if len(value) <= limit else f"{value[: limit - 3]}..."


def _string_or_none(value: object) -> str | None:
    return str(value) if value else None


def _elapsed_seconds(started_at: str | None, finished_at: str | None) -> int:
    if not started_at:
        return 0
    try:
        started = datetime.fromisoformat(started_at)
        finished = datetime.fromisoformat(finished_at) if finished_at else datetime.now()
    except ValueError:
        return 0
    if started.tzinfo is not None:
        finished = finished.replace(tzinfo=timezone.utc) if finished.tzinfo is None else finished
    return max(0, int((finished - started).total_seconds()))
