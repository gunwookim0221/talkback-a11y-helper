from __future__ import annotations

import openpyxl
import json
from pathlib import Path
from .paths import OUTPUT_DIR
from .recent_runs import safe_recent_run_log_path
from .run_summary import read_summary_file, summary_path_for_log

SEVERE_MISMATCH_TYPES = {
    "EMPTY_VISIBLE",
    "EMPTY_SPEECH",
    "TEXT_MISMATCH",
    "LABEL_MISMATCH",
    "SPOKEN_MISMATCH",
    "MISMATCH",
    "FAIL_MISMATCH",
}

def get_run_mismatch_summary(run_id: str) -> dict[str, object]:
    try:
        log_path = safe_recent_run_log_path(run_id)
    except (FileNotFoundError, ValueError):
        return {"error": "run not found"}

    summary = read_summary_file(summary_path_for_log(log_path))
    xlsx_path = None
    if summary:
        if summary.get("xlsx_path"):
            xlsx_path = Path(str(summary["xlsx_path"]))
        elif summary.get("xlsx_filename"):
            xlsx_path = OUTPUT_DIR / str(summary["xlsx_filename"])

    if not xlsx_path:
        return {"error": "xlsx output not available"}

    return get_mismatch_summary_from_xlsx(xlsx_path)

def get_mismatch_summary_from_xlsx(xlsx_path: Path) -> dict[str, object]:
    if not xlsx_path.exists():
        return {"error": "xlsx file not found"}

    try:
        workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
        if "result" not in workbook.sheetnames:
            return {"error": "result sheet not found"}
        
        sheet = workbook["result"]
        headers = [str(sheet.cell(1, c).value or "") for c in range(1, sheet.max_column + 1)]
        
        def _get_col(name: str) -> int | None:
            return headers.index(name) + 1 if name in headers else None

        scenario_col = _get_col("scenario_id") or _get_col("scenario")
        plugin_name_col = _get_col("plugin_name")
        step_col = _get_col("step")
        visible_col = _get_col("visible_label")
        speech_col = _get_col("merged_announcement")
        mismatch_col = _get_col("mismatch_type")
        result_col = _get_col("final_result")
        failure_col = _get_col("failure_reason")
        focus_col = _get_col("focus_confidence")
        context_col = _get_col("context_type")
        note_col = _get_col("review_note")
        crop_col = _get_col("result_crop_thumbnail")
        repeat_count_col = _get_col("repeat_count")
        first_step_col = _get_col("first_step")
        last_step_col = _get_col("last_step")
        steps_col = _get_col("steps")
        repeated_group_col = _get_col("is_repeated_issue_group")
        shadow_verdict_col = _get_col("shadow_verdict")
        shadow_reason_col = _get_col("shadow_verdict_reason")
        shadow_source_col = _get_col("shadow_verdict_source")
        scenario_shadow_col = _get_col("scenario_shadow_verdict")

        def _int_cell(row: int, col: int | None, default: int = 0) -> int:
            if not col:
                return default
            try:
                return int(sheet.cell(row, col).value or default)
            except (TypeError, ValueError):
                return default

        summary_matched = 0
        summary_true_mismatch = 0
        summary_empty_speech = 0
        summary_empty_visible = 0
        summary_review = 0
        summary_runtime_warning = 0

        summary_fail_count = 0
        summary_issue_count = 0
        summary_review_count = 0
        summary_clean_count = 0
        
        scenario_stats = {}
        all_previews = []

        for row in range(2, sheet.max_row + 1):
            scenario = str(sheet.cell(row, scenario_col).value or "").strip() if scenario_col else ""
            if not scenario:
                scenario = "unknown"
            
            if scenario not in scenario_stats:
                scenario_stats[scenario] = {
                    "scenario_id": scenario,
                    "plugin_name": "",
                    "matched": 0,
                    "true_mismatch": 0,
                    "empty_speech": 0,
                    "empty_visible": 0,
                    "review": 0,
                    "runtime_warning": 0,
                    "fail_count": 0,
                    "issue_count": 0,
                    "review_count": 0,
                    "clean_count": 0,
                    "shadow_pass_count": 0,
                    "shadow_review_count": 0,
                    "shadow_warn_count": 0,
                    "shadow_fail_count": 0,
                    "scenario_shadow_verdict": "",
                    "focusable_required_missed": 0,
                    "focusable_review_unknown": 0,
                    "focusable_coverage_rate": None,
                    "status": "clean"
                }

            plugin_name = str(sheet.cell(row, plugin_name_col).value or "").strip() if plugin_name_col else ""
            step = str(sheet.cell(row, step_col).value or "").strip() if step_col else ""
            visible = str(sheet.cell(row, visible_col).value or "").strip() if visible_col else ""
            speech = str(sheet.cell(row, speech_col).value or "").strip() if speech_col else ""
            mismatch_type = str(sheet.cell(row, mismatch_col).value or "").strip().upper() if mismatch_col else ""
            final_result = str(sheet.cell(row, result_col).value or "").strip().upper() if result_col else ""
            failure_reason = str(sheet.cell(row, failure_col).value or "").strip() if failure_col else ""
            focus_confidence = str(sheet.cell(row, focus_col).value or "").strip() if focus_col else ""
            context_type = str(sheet.cell(row, context_col).value or "").strip() if context_col else ""
            review_note = str(sheet.cell(row, note_col).value or "").strip() if note_col else ""
            crop_thumbnail = str(sheet.cell(row, crop_col).value or "").strip() if crop_col else ""
            repeat_count = _int_cell(row, repeat_count_col, 1)
            first_step = str(sheet.cell(row, first_step_col).value or "").strip() if first_step_col else step
            last_step = str(sheet.cell(row, last_step_col).value or "").strip() if last_step_col else step
            repeated_steps = str(sheet.cell(row, steps_col).value or "").strip() if steps_col else step
            is_repeated_issue_group = False
            if repeated_group_col:
                raw_repeated = sheet.cell(row, repeated_group_col).value
                is_repeated_issue_group = str(raw_repeated).strip().lower() in {"true", "1", "yes"}
            shadow_verdict = str(sheet.cell(row, shadow_verdict_col).value or "").strip().upper() if shadow_verdict_col else ""
            shadow_reason = str(sheet.cell(row, shadow_reason_col).value or "").strip() if shadow_reason_col else ""
            shadow_source = str(sheet.cell(row, shadow_source_col).value or "").strip() if shadow_source_col else ""
            scenario_shadow_verdict = str(sheet.cell(row, scenario_shadow_col).value or "").strip().upper() if scenario_shadow_col else ""

            if not scenario_stats[scenario]["plugin_name"] and plugin_name:
                scenario_stats[scenario]["plugin_name"] = plugin_name
            if scenario_shadow_verdict and not scenario_stats[scenario]["scenario_shadow_verdict"]:
                scenario_stats[scenario]["scenario_shadow_verdict"] = scenario_shadow_verdict
            if shadow_verdict == "SHADOW_PASS":
                scenario_stats[scenario]["shadow_pass_count"] += 1
            elif shadow_verdict == "SHADOW_REVIEW":
                scenario_stats[scenario]["shadow_review_count"] += 1
            elif shadow_verdict == "SHADOW_WARN":
                scenario_stats[scenario]["shadow_warn_count"] += 1
            elif shadow_verdict == "SHADOW_FAIL":
                scenario_stats[scenario]["shadow_fail_count"] += 1

            category = ""
            is_fail = False
            is_issue = False
            is_review = False
            is_clean = False

            both_text_empty = not visible and not speech
            exact_visible_speech_match = bool(visible and speech and visible == speech)
            severe_fail_row = final_result == "FAIL" and mismatch_type in SEVERE_MISMATCH_TYPES

            if mismatch_type == "EMPTY_VISIBLE":
                summary_empty_visible += 1
                scenario_stats[scenario]["empty_visible"] += 1
                if severe_fail_row and both_text_empty:
                    is_fail = True
                    category = "EMPTY_VISIBLE_FAIL"
                elif both_text_empty:
                    summary_review += 1
                    scenario_stats[scenario]["review"] += 1
                    is_review = True
                    category = "REVIEW"
                else:
                    is_issue = True
                    category = "EMPTY_VISIBLE"
            elif mismatch_type == "EMPTY_SPEECH" or (visible and not speech):
                summary_empty_speech += 1
                scenario_stats[scenario]["empty_speech"] += 1
                is_fail = True
                category = "EMPTY_SPEECH"
            elif mismatch_type in {"PARTIAL_MATCH", "REPRESENTATIVE_CONTEXT"}:
                summary_review += 1
                scenario_stats[scenario]["review"] += 1
                is_review = True
                category = "REVIEW"
            elif mismatch_type in {"TEXT_MISMATCH", "LABEL_MISMATCH", "SPOKEN_MISMATCH", "MISMATCH", "FAIL_MISMATCH"} or (mismatch_type and "MATCH" not in mismatch_type and "EMPTY" not in mismatch_type):
                summary_true_mismatch += 1
                scenario_stats[scenario]["true_mismatch"] += 1
                is_fail = True
                category = "TRUE_MISMATCH"
            elif mismatch_type in {"EXACT_MATCH", "NORMALIZED_MATCH"} or (final_result == "PASS" and visible and speech):
                summary_matched += 1
                scenario_stats[scenario]["matched"] += 1
                is_clean = True
                category = "MATCHED"
            else:
                # Catch-all
                if final_result == "PASS":
                    summary_matched += 1
                    scenario_stats[scenario]["matched"] += 1
                    is_clean = True
                    category = "MATCHED"
                else:
                    summary_review += 1
                    scenario_stats[scenario]["review"] += 1
                    is_review = True
                    category = "REVIEW"

            # Runtime Warning Override
            if final_result == "WARN" and failure_reason:
                lower_reason = failure_reason.lower()
                if any(ignored in lower_reason for ignored in ["repeat_no_progress", "viewport_exhausted", "terminal_reached", "end_of_content", "no_unvisited_local_tab"]):
                    # Ignored warning, treated as clean
                    pass
                elif any(reason in lower_reason for reason in ["plugin_open_failed", "terminal_not_handled", "activation_fail", "parse_error", "fatal", "exception"]):
                    summary_runtime_warning += 1
                    scenario_stats[scenario]["runtime_warning"] += 1
                    is_issue = True
                    category = "RUNTIME_WARNING"

            # Top-level Categorization
            if is_fail:
                summary_fail_count += 1
                scenario_stats[scenario]["fail_count"] += 1
                top_category = "FAIL"
            elif is_issue:
                summary_issue_count += 1
                scenario_stats[scenario]["issue_count"] += 1
                top_category = "ISSUE"
            elif is_review:
                summary_review_count += 1
                scenario_stats[scenario]["review_count"] += 1
                top_category = "REVIEW"
            else:
                summary_clean_count += 1
                scenario_stats[scenario]["clean_count"] += 1
                top_category = "CLEAN"

            add_to_preview = top_category in {"FAIL", "ISSUE"}
            if exact_visible_speech_match:
                add_to_preview = False

            if add_to_preview:
                all_previews.append({
                    "scenario_id": scenario,
                    "plugin_name": plugin_name,
                    "step": step,
                    "context_type": context_type,
                    "visible_label": visible,
                    "merged_announcement": speech,
                    "mismatch_type": mismatch_type,
                    "final_result": final_result,
                    "failure_reason": failure_reason,
                    "focus_confidence": focus_confidence,
                    "review_note": review_note,
                    "crop_thumbnail": crop_thumbnail,
                    "repeat_count": repeat_count,
                    "first_step": first_step,
                    "last_step": last_step,
                    "steps": repeated_steps,
                    "is_repeated_issue_group": is_repeated_issue_group,
                    "shadow_verdict": shadow_verdict,
                    "shadow_verdict_reason": shadow_reason,
                    "shadow_verdict_source": shadow_source,
                    "scenario_shadow_verdict": scenario_shadow_verdict,
                    "category": category,
                    "top_category": top_category
                })

        workbook.close()

        # Sort previews by priority: FAIL > ISSUE > REVIEW
        priority_map = {
            "FAIL": 1,
            "ISSUE": 2,
            "REVIEW": 3
        }
        all_previews.sort(key=lambda x: priority_map.get(x["top_category"], 99))
        previews = all_previews[:20]

        focusable_coverage = _read_focusable_coverage_for_xlsx(xlsx_path)
        coverage_probe_summary = _read_coverage_probe_summary_for_xlsx(xlsx_path, sheet, headers)
        focusable_summary = focusable_coverage.get("summary", {})
        focusable_by_scenario = {
            str(item.get("scenario_id", "") or ""): item
            for item in focusable_coverage.get("scenarios", [])
            if isinstance(item, dict)
        }

        for scenario, coverage_stats in focusable_by_scenario.items():
            stats = scenario_stats.setdefault(
                scenario,
                {
                    "scenario_id": scenario,
                    "plugin_name": "",
                    "matched": 0,
                    "true_mismatch": 0,
                    "empty_speech": 0,
                    "empty_visible": 0,
                    "review": 0,
                    "runtime_warning": 0,
                    "fail_count": 0,
                    "issue_count": 0,
                    "review_count": 0,
                    "clean_count": 0,
                    "shadow_pass_count": 0,
                    "shadow_review_count": 0,
                    "shadow_warn_count": 0,
                    "shadow_fail_count": 0,
                    "scenario_shadow_verdict": "",
                    "focusable_required_missed": 0,
                    "focusable_review_unknown": 0,
                    "focusable_coverage_rate": None,
                    "status": "clean",
                },
            )
            stats["focusable_required_missed"] = int(coverage_stats.get("focusable_required_missed") or 0)
            stats["focusable_review_unknown"] = int(coverage_stats.get("focusable_review_unknown") or 0)
            stats["focusable_coverage_rate"] = coverage_stats.get("focusable_coverage_rate")

        # Calculate scenario status after all reporting-only scenario metrics are attached.
        scenario_summary = []
        for s_id, stats in scenario_stats.items():
            if stats["fail_count"] > 0:
                stats["status"] = "fail"
            elif stats["issue_count"] > 0:
                stats["status"] = "issue"
            elif stats["review_count"] > 0:
                stats["status"] = "review"
            else:
                stats["status"] = "clean"
            scenario_summary.append(stats)

        shadow_summary = _read_shadow_summary_from_xlsx(xlsx_path)
        if not shadow_summary:
            shadow_summary = {
                "shadow_pass_count": sum(stats["shadow_pass_count"] for stats in scenario_stats.values()),
                "shadow_review_count": sum(stats["shadow_review_count"] for stats in scenario_stats.values()),
                "shadow_warn_count": sum(stats["shadow_warn_count"] for stats in scenario_stats.values()),
                "shadow_fail_count": sum(stats["shadow_fail_count"] for stats in scenario_stats.values()),
            }

        return {
            "summary": {
                "fail_count": summary_fail_count,
                "issue_count": summary_issue_count,
                "review_count": summary_review_count,
                "clean_count": summary_clean_count,
                "matched": summary_matched,
                "true_mismatch": summary_true_mismatch,
                "empty_speech": summary_empty_speech,
                "empty_visible": summary_empty_visible,
                "review": summary_review,
                "runtime_warning": summary_runtime_warning,
                "shadow_pass_count": int(shadow_summary.get("shadow_pass_count") or 0),
                "shadow_review_count": int(shadow_summary.get("shadow_review_count") or 0),
                "shadow_warn_count": int(shadow_summary.get("shadow_warn_count") or 0),
                "shadow_fail_count": int(shadow_summary.get("shadow_fail_count") or 0),
                "focusable_required_expected_count": int(focusable_summary.get("focusable_required_expected_count") or 0),
                "focusable_required_covered_count": int(focusable_summary.get("focusable_required_covered_count") or 0),
                "focusable_required_missed_count": int(focusable_summary.get("focusable_required_missed_count") or 0),
                "focusable_review_expected_count": int(focusable_summary.get("focusable_review_expected_count") or 0),
                "focusable_review_unknown_count": int(focusable_summary.get("focusable_review_unknown_count") or 0),
                "focusable_optional_expected_count": int(focusable_summary.get("focusable_optional_expected_count") or 0),
                "focusable_coverage_rate": focusable_summary.get("focusable_coverage_rate"),
            },
            "scenario_summary": scenario_summary,
            "signals": previews,
            "focusable_coverage": focusable_coverage,
            "coverage_probe_summary": coverage_probe_summary,
            "coverage_probe": coverage_probe_summary,
            "debug": {
                "xlsx_path": xlsx_path.name,
                "result_rows": sheet.max_row - 1 if sheet.max_row > 1 else 0
            }
        }

    except Exception as exc:
        return {"error": f"Failed to parse xlsx: {str(exc)}"}


def _read_shadow_summary_from_xlsx(xlsx_path: Path) -> dict[str, int]:
    try:
        workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "summary" not in workbook.sheetnames:
            workbook.close()
            return {}
        sheet = workbook["summary"]
        headers = [str(sheet.cell(1, c).value or "") for c in range(1, sheet.max_column + 1)]
        if "metric" not in headers or "value" not in headers:
            workbook.close()
            return {}
        metric_col = headers.index("metric") + 1
        value_col = headers.index("value") + 1
        wanted = {
            "shadow_pass_count",
            "shadow_review_count",
            "shadow_warn_count",
            "shadow_fail_count",
        }
        values: dict[str, int] = {}
        for row in range(2, sheet.max_row + 1):
            metric = str(sheet.cell(row, metric_col).value or "").strip()
            if metric not in wanted:
                continue
            try:
                values[metric] = int(float(sheet.cell(row, value_col).value or 0))
            except (TypeError, ValueError):
                values[metric] = 0
        workbook.close()
        return values
    except Exception:
        return {}


def _read_json_object(path: Path) -> dict[str, object] | None:
    if not path.is_file():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return payload if isinstance(payload, dict) else None


def _artifact_metric(payload: dict[str, object] | None, *keys: str) -> int | None:
    if not payload:
        return None
    summary = payload.get("summary")
    sources = (payload, summary if isinstance(summary, dict) else {})
    for source in sources:
        for key in keys:
            if key not in source:
                continue
            try:
                return int(float(source[key] or 0))
            except (TypeError, ValueError):
                return 0
    return None


def _read_coverage_probe_summary_for_xlsx(
    xlsx_path: Path,
    result_sheet: openpyxl.worksheet.worksheet.Worksheet,
    result_headers: list[str],
) -> dict[str, object]:
    aggregate_path = xlsx_path.with_name(f"{xlsx_path.stem}.coverage_probe_validation.aggregate.json")
    fallback_path = xlsx_path.with_name(f"{xlsx_path.stem}.coverage_probe_validation.json")
    validation_path: Path | None = aggregate_path
    validation = _read_json_object(aggregate_path)
    source = "aggregate"
    if validation is None:
        validation_path = fallback_path
        validation = _read_json_object(validation_path)
        source = "scenario"
    if validation is None:
        return {
            "available": False,
            "source": "none",
            "results_artifact": None,
            "validation_artifact": None,
        }

    results_aggregate_path = xlsx_path.with_name(f"{xlsx_path.stem}.coverage_probe_results.aggregate.json")
    results_fallback_path = xlsx_path.with_name(f"{xlsx_path.stem}.coverage_probe_results.json")
    results = _read_json_object(results_aggregate_path)
    results_path: Path | None = results_aggregate_path if results is not None else None
    if results is None:
        results = _read_json_object(results_fallback_path)
        results_path = results_fallback_path if results is not None else None

    row_source_col = result_headers.index("row_source") + 1 if "row_source" in result_headers else None
    dedup_status_col = (
        result_headers.index("promotion_dedup_status") + 1
        if "promotion_dedup_status" in result_headers
        else None
    )
    promoted_row_count = _artifact_metric(validation, "promoted_row_count")
    dedup_skipped_count = _artifact_metric(validation, "promotion_dedup_skipped_count")
    if promoted_row_count is None or dedup_skipped_count is None:
        promoted_rows = 0
        dedup_skipped_rows = 0
        for row in range(2, result_sheet.max_row + 1):
            row_source = (
                str(result_sheet.cell(row, row_source_col).value or "").strip().upper()
                if row_source_col
                else ""
            )
            dedup_status = (
                str(result_sheet.cell(row, dedup_status_col).value or "").strip().upper()
                if dedup_status_col
                else ""
            )
            if row_source == "COVERAGE_PROBE_PROMOTED":
                promoted_rows += 1
            if row_source == "COVERAGE_PROBE_SHADOW" and dedup_status == "SKIPPED":
                dedup_skipped_rows += 1
        if promoted_row_count is None:
            promoted_row_count = promoted_rows
        if dedup_skipped_count is None:
            dedup_skipped_count = dedup_skipped_rows

    screen_skipped = _artifact_metric(validation, "total_screen_skipped_count", "screen_skipped_count")
    if screen_skipped is None:
        screen_skipped = _artifact_metric(results, "total_screen_skipped_count", "screen_skipped_count")
    scenario_filtered = _artifact_metric(
        validation,
        "total_scenario_filtered_count",
        "scenario_filtered_count",
    )
    if scenario_filtered is None:
        scenario_filtered = _artifact_metric(
            results,
            "total_scenario_filtered_count",
            "scenario_filtered_count",
        )

    return {
        "available": True,
        "source": source,
        "results_artifact": results_path.name if results_path else None,
        "validation_artifact": validation_path.name,
        "candidate_count": _artifact_metric(results, "total_candidate_count", "candidate_count") or 0,
        "attempted_count": _artifact_metric(results, "total_attempted_count", "attempted_count") or 0,
        "success_count": _artifact_metric(results, "total_success_count", "success_count") or 0,
        "failed_count": _artifact_metric(results, "total_failed_count", "failed_count") or 0,
        "match_count": _artifact_metric(validation, "total_match_count", "match_count") or 0,
        "promotable_count": _artifact_metric(validation, "promotable_count") or 0,
        "not_promotable_count": _artifact_metric(validation, "not_promotable_count") or 0,
        "promoted_row_count": promoted_row_count or 0,
        "dedup_skipped_count": dedup_skipped_count or 0,
        "screen_skipped_count": screen_skipped,
        "scenario_filtered_count": scenario_filtered,
    }


def _empty_focusable_coverage() -> dict[str, object]:
    return {
        "summary": {
            "focusable_required_expected_count": 0,
            "focusable_required_covered_count": 0,
            "focusable_required_missed_count": 0,
            "focusable_review_expected_count": 0,
            "focusable_review_unknown_count": 0,
            "focusable_optional_expected_count": 0,
            "focusable_coverage_rate": None,
        },
        "scenarios": [],
        "issues": [],
    }


def _read_focusable_coverage_for_xlsx(xlsx_path: Path) -> dict[str, object]:
    coverage_path = xlsx_path.with_name(f"{xlsx_path.stem}.focusable_coverage.json")
    if not coverage_path.is_file():
        return _empty_focusable_coverage()
    try:
        payload = json.loads(coverage_path.read_text(encoding="utf-8"))
    except Exception:
        return _empty_focusable_coverage()

    summaries = payload.get("summary")
    records = payload.get("records")
    if not isinstance(summaries, list):
        summaries = []
    if not isinstance(records, list):
        records = []

    total_required_expected = 0
    total_required_covered = 0
    total_required_missed = 0
    total_review_expected = 0
    total_review_unknown = 0
    total_optional_expected = 0
    expected_total = 0
    covered_total = 0
    scenarios: list[dict[str, object]] = []

    for item in summaries:
        if not isinstance(item, dict):
            continue
        required_expected = _safe_int(item.get("required_expected_count"))
        required_covered = _safe_int(item.get("required_covered_count"))
        required_missed = _safe_int(item.get("required_missed_count"))
        review_expected = _safe_int(item.get("review_expected_count"))
        review_unknown = _safe_int(item.get("review_unknown_count"))
        optional_expected = _safe_int(item.get("optional_expected_count"))
        expected = _safe_int(item.get("expected_count"))
        covered = _safe_int(item.get("covered_count"))
        total_required_expected += required_expected
        total_required_covered += required_covered
        total_required_missed += required_missed
        total_review_expected += review_expected
        total_review_unknown += review_unknown
        total_optional_expected += optional_expected
        expected_total += expected
        covered_total += covered
        scenarios.append(
            {
                "scenario_id": str(item.get("scenario_id", "") or ""),
                "focusable_required_missed": required_missed,
                "focusable_review_unknown": review_unknown,
                "focusable_coverage_rate": item.get("coverage_rate"),
            }
        )

    issues = []
    for record in records:
        if not isinstance(record, dict):
            continue
        taxonomy = str(record.get("taxonomy", "") or "").upper()
        status = str(record.get("coverage_status", "") or "").upper()
        if not ((taxonomy == "REQUIRED" and status == "MISSED") or (taxonomy == "REVIEW" and status == "UNKNOWN")):
            continue
        issues.append(
            {
                "scenario_id": str(record.get("scenario_id", "") or ""),
                "focusable_label": str(record.get("label", "") or ""),
                "focusable_view_id": str(record.get("view_id", "") or ""),
                "focusable_taxonomy": taxonomy,
                "focusable_coverage_status": status,
                "focusable_coverage_reason": str(record.get("coverage_reason", "") or ""),
                "focusable_taxonomy_reason": str(record.get("taxonomy_reason", "") or ""),
            }
        )

    return {
        "summary": {
            "focusable_required_expected_count": total_required_expected,
            "focusable_required_covered_count": total_required_covered,
            "focusable_required_missed_count": total_required_missed,
            "focusable_review_expected_count": total_review_expected,
            "focusable_review_unknown_count": total_review_unknown,
            "focusable_optional_expected_count": total_optional_expected,
            "focusable_coverage_rate": round((covered_total / expected_total) * 100.0, 1) if expected_total else None,
        },
        "scenarios": scenarios,
        "issues": issues,
    }


def _safe_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0
