from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl


COVERAGE_HEALTH_SCHEMA_VERSION = 1
UNAVAILABLE_STATES = {
    "NOT_AVAILABLE",
    "NOT_AVAILABLE_CANDIDATE",
    "NO_TARGET_CANDIDATE",
}

_STABILIZATION_RE = re.compile(
    r"\[SCENARIO\]\[stabilization\].*?scenario='(?P<scenario>[^']+)'"
    r".*?stabilization_mode='(?P<mode>[^']+)'"
)
_ANCHOR_START_RE = re.compile(
    r"\[ANCHOR\]\[scenario_start\]\s+(?P<state>\w+)\s+scenario='(?P<scenario>[^']+)'"
)
_ENTRY_CONTRACT_RE = re.compile(
    r"\[SCENARIO\]\[entry_contract\]\s+(?P<state>\w+)\s+scenario='(?P<scenario>[^']+)'"
)
_STOP_RE = re.compile(
    r"\[STOP\]\[eval\].*?scenario='(?P<scenario>[^']+)'.*?"
    r"traversal_result='(?P<traversal>[^']+)'"
)


def build_coverage_health_report(
    *,
    scenarios: object,
    log_path: Path | None = None,
    xlsx_path: Path | None = None,
    focusable_coverage: object = None,
) -> dict[str, object]:
    """Build the validator-facing coverage projection from existing evidence.

    This function is deliberately downstream-only. It interprets persisted
    scenario summaries, the normal runtime log, and the result/focusable
    artifacts; it does not participate in traversal or scenario decisions.
    """

    log_evidence = _read_log_evidence(log_path)
    content_rows, semantic_values = _read_result_evidence(xlsx_path)
    focusable = _focusable_projection(focusable_coverage, xlsx_path=xlsx_path)
    normalized_scenarios: list[dict[str, object]] = []

    raw_scenarios = scenarios if isinstance(scenarios, list) else []
    for raw in raw_scenarios:
        if not isinstance(raw, dict):
            continue
        scenario_id = str(raw.get("id") or raw.get("scenario_id") or "").strip()
        if not scenario_id:
            continue

        mode = log_evidence["stabilization_mode"].get(scenario_id)
        if mode is None:
            value = raw.get("stabilization_mode")
            mode = str(value) if value else None
        entry_state = log_evidence["entry_state"].get(scenario_id)
        special = bool(raw.get("special_state_handled")) or entry_state == "handled"
        row_count = (
            content_rows.get(scenario_id, 0)
            if xlsx_path and xlsx_path.is_file()
            else None
        )
        availability = _availability_state(raw)
        anchor_present = log_evidence["anchor_present"].get(scenario_id)
        entry_success = entry_state == "success"
        entry_event_failed = entry_state in {"failed", "abort"}

        if special:
            content_entered: bool | None = False
        elif availability in UNAVAILABLE_STATES:
            content_entered = False
        elif row_count is not None and row_count > 0:
            content_entered = True
        elif entry_success:
            content_entered = True
        else:
            content_entered = None

        if special or content_entered is True:
            entry_failed: bool | None = False
        elif availability in UNAVAILABLE_STATES or entry_event_failed:
            entry_failed = True
        elif content_entered is False:
            entry_failed = True
        else:
            entry_failed = None

        if special:
            derived = "HANDLED_SPECIAL_STATE"
            availability_value = "HANDLED_SPECIAL_STATE"
        elif availability in UNAVAILABLE_STATES:
            derived = "ENTRY_FAILED_OR_UNAVAILABLE"
            availability_value = availability
        elif content_entered is True:
            derived = "CONTENT_TRAVERSED"
            availability_value = "AVAILABLE_ENTERED"
        elif entry_failed is True:
            derived = "ENTRY_FAILED_OR_UNAVAILABLE"
            availability_value = "ENTRY_FAILED"
        else:
            derived = "INDETERMINATE"
            availability_value = availability or "UNKNOWN"

        terminal = raw.get("traversal_result")
        if not terminal:
            terminal = log_evidence["traversal_state"].get(scenario_id)

        normalized_scenarios.append(
            {
                "scenario_id": scenario_id,
                "stabilization_mode": mode,
                "anchor_present": anchor_present,
                "entry_failed": entry_failed,
                "content_entered": content_entered,
                "content_row_count": row_count,
                "traversal_terminal_state": str(terminal) if terminal else None,
                "availability_state": availability_value,
                "derived_classification": derived,
                "scenario_result": raw.get("status"),
                "stop_reason": raw.get("stop_reason"),
            }
        )

    anchor_mode = [
        item for item in normalized_scenarios if item.get("stabilization_mode") == "anchor_only"
    ]
    content_traversed = [item for item in anchor_mode if item.get("derived_classification") == "CONTENT_TRAVERSED"]
    unavailable = [
        item
        for item in anchor_mode
        if item.get("derived_classification") == "ENTRY_FAILED_OR_UNAVAILABLE"
        and item.get("availability_state") in UNAVAILABLE_STATES
    ]
    handled = [
        item for item in anchor_mode if item.get("derived_classification") == "HANDLED_SPECIAL_STATE"
    ]
    true_anchor_failures = [
        item
        for item in anchor_mode
        if item.get("derived_classification") == "ENTRY_FAILED_OR_UNAVAILABLE"
        and item.get("availability_state") == "ENTRY_FAILED"
    ]
    semantic_total = int(semantic_values["expected_count"])
    semantic_covered = int(semantic_values["covered_count"])

    return {
        "schema_version": COVERAGE_HEALTH_SCHEMA_VERSION,
        "summary": {
            "anchor_mode_count": len(anchor_mode),
            "true_anchor_traversal_failure_count": len(true_anchor_failures),
            "anchor_mode_content_traversed_count": len(content_traversed),
            "unavailable_or_no_target_count": len(unavailable),
            "handled_or_ambiguous_count": len(handled),
            "content_entered_count": sum(item.get("content_entered") is True for item in normalized_scenarios),
            "entry_failed_count": sum(item.get("entry_failed") is True for item in normalized_scenarios),
        },
        "scenarios": normalized_scenarios,
        "focusable_candidate_coverage": focusable,
        "semantic_value_coverage": {
            "label": "Semantic value coverage",
            "formula": "covered / expected",
            "covered_count": semantic_covered,
            "expected_count": semantic_total,
            "rate": round((semantic_covered / semantic_total) * 100.0, 1) if semantic_total else None,
        },
        "evidence": {
            "stabilization_mode_source": "normal_runtime_log" if log_evidence["stabilization_mode"] else "unavailable",
            "content_row_source": "result_workbook" if xlsx_path and xlsx_path.is_file() else "unavailable",
            "semantic_value_source": "result_workbook" if semantic_total else "unavailable",
        },
    }


def _read_log_evidence(log_path: Path | None) -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {
        "stabilization_mode": {},
        "anchor_present": {},
        "entry_state": {},
        "traversal_state": {},
    }
    if not log_path or not log_path.is_file():
        return result
    try:
        lines = log_path.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError:
        return result

    for line in lines:
        match = _STABILIZATION_RE.search(line)
        if match:
            result["stabilization_mode"][match.group("scenario")] = match.group("mode")
        match = _ANCHOR_START_RE.search(line)
        if match:
            state = match.group("state").lower()
            if state in {"success", "verified"}:
                result["anchor_present"][match.group("scenario")] = True
                result["entry_state"].setdefault(match.group("scenario"), None)
            elif state == "abort":
                result["entry_state"][match.group("scenario")] = "abort"
        match = _ENTRY_CONTRACT_RE.search(line)
        if match:
            result["entry_state"][match.group("scenario")] = match.group("state").lower()
        match = _STOP_RE.search(line)
        if match:
            result["traversal_state"][match.group("scenario")] = match.group("traversal")
    return result


def _read_result_evidence(xlsx_path: Path | None) -> tuple[Counter[str], dict[str, int]]:
    content_rows: Counter[str] = Counter()
    semantic = {"covered_count": 0, "expected_count": 0}
    if not xlsx_path or not xlsx_path.is_file():
        return content_rows, semantic
    try:
        workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "result" not in workbook.sheetnames:
            workbook.close()
            return content_rows, semantic
        sheet = workbook["result"]
        headers = [str(sheet.cell(1, column).value or "") for column in range(1, sheet.max_column + 1)]
        columns = {name: headers.index(name) for name in headers if name}
        scenario_column = columns.get("scenario_id", columns.get("scenario"))
        total_column = columns.get("semantic_value_total_count")
        matched_column = columns.get("semantic_value_matched_count")
        if scenario_column is None:
            workbook.close()
            return content_rows, semantic
        for row in sheet.iter_rows(min_row=2, values_only=True):
            scenario = str(row[scenario_column] or "").strip()
            if scenario:
                content_rows[scenario] += 1
            if total_column is not None:
                semantic["expected_count"] += _safe_int(row[total_column])
            if matched_column is not None:
                semantic["covered_count"] += _safe_int(row[matched_column])
        workbook.close()
    except Exception:
        return Counter(), {"covered_count": 0, "expected_count": 0}
    return content_rows, semantic


def _availability_state(source: dict[str, Any]) -> str | None:
    value = source.get("availability_status")
    if value:
        return str(value).upper()
    status = str(source.get("status") or "").lower()
    if status == "no_target_candidate":
        return "NO_TARGET_CANDIDATE"
    if status == "not_available_candidate":
        return "NOT_AVAILABLE_CANDIDATE"
    if status == "not_available":
        return "NOT_AVAILABLE"
    return None


def _focusable_projection(value: object, *, xlsx_path: Path | None = None) -> dict[str, object]:
    source = value if isinstance(value, dict) else {}
    summary = source.get("summary") if isinstance(source.get("summary"), dict) else {}
    expected = _safe_int(summary.get("focusable_expected_count"))
    covered = _safe_int(summary.get("focusable_covered_count"))
    missed = _safe_int(summary.get("focusable_missed_count"))
    unknown = _safe_int(summary.get("focusable_unknown_count"))
    ignored = _safe_int(summary.get("focusable_ignored_count"))
    if xlsx_path and (not expected or not covered or not missed or not unknown or not ignored):
        sidecar = xlsx_path.with_name(f"{xlsx_path.stem}.focusable_coverage.json")
        try:
            import json

            payload = json.loads(sidecar.read_text(encoding="utf-8"))
            sidecar_summaries = payload.get("summary") if isinstance(payload, dict) else None
            if isinstance(sidecar_summaries, list):
                expected = sum(_safe_int(item.get("expected_count")) for item in sidecar_summaries if isinstance(item, dict))
                covered = sum(_safe_int(item.get("covered_count")) for item in sidecar_summaries if isinstance(item, dict))
                missed = sum(_safe_int(item.get("missed_count")) for item in sidecar_summaries if isinstance(item, dict))
                unknown = sum(_safe_int(item.get("unknown_count")) for item in sidecar_summaries if isinstance(item, dict))
                ignored = sum(_safe_int(item.get("ignore_count")) for item in sidecar_summaries if isinstance(item, dict))
        except (OSError, ValueError, TypeError):
            pass
    if not expected:
        expected = _safe_int(summary.get("focusable_required_expected_count")) + _safe_int(summary.get("focusable_review_expected_count")) + _safe_int(summary.get("focusable_optional_expected_count"))
    if not covered:
        covered = _safe_int(summary.get("focusable_required_covered_count"))
    if not missed and not unknown:
        missed = _safe_int(summary.get("focusable_required_missed_count"))
        unknown = _safe_int(summary.get("focusable_review_unknown_count"))
    rate = summary.get("focusable_coverage_rate")
    if not isinstance(rate, (int, float)):
        rate = round((covered / expected) * 100.0, 1) if expected else None
    return {
        "label": "Focusable candidate coverage",
        "formula": "covered / expected",
        "covered_count": covered,
        "expected_count": expected,
        "missed_count": missed,
        "unknown_count": unknown,
        "ignored_count": ignored,
        "rate": rate,
    }


def _safe_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0
