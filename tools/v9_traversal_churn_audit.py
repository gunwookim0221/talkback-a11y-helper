"""Audit traversal churn from completed run artifacts without changing traversal behavior."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from tools.v8_fullrun_profile import resolve_artifacts


if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


AUDIT_JSON = "v9_traversal_churn_audit.json"
AUDIT_MARKDOWN = "v9_traversal_churn_audit.md"
AUDIT_CSV = "v9_traversal_churn_audit.csv"

PRIMARY_SCENARIOS = [
    "life_energy_plugin",
    "life_pet_care_plugin",
    "life_plant_care_plugin",
    "device_washer_plugin",
    "life_family_care_plugin",
]

TRAVERSAL_ROW_SOURCES = {
    "",
    "ACTUAL_FOCUS",
    "REPRESENTATIVE",
    "REPRESENTATIVE_FALLBACK",
}

BOTTOM_STRIP_LABELS = {
    "eventsbutton",
    "activitybutton",
    "locationbutton",
    "controls",
    "history",
    "routines",
    "events",
    "activity",
    "location",
}

CATEGORY_PRODUCTIVE = "PRODUCTIVE"
CATEGORY_DUPLICATE = "DUPLICATE"
CATEGORY_BOTTOM_STRIP = "BOTTOM_STRIP_CHURN"
CATEGORY_LOCAL_TAB = "LOCAL_TAB_CHURN"
CATEGORY_OVERLAY = "OVERLAY_CHURN"
CATEGORY_SCROLL_RETRY = "SCROLL_RETRY_CHURN"
CATEGORY_VIEWPORT_EXHAUSTED = "VIEWPORT_EXHAUSTED_CHURN"
CATEGORY_UNKNOWN = "UNKNOWN"

CATEGORY_ORDER = [
    CATEGORY_PRODUCTIVE,
    CATEGORY_DUPLICATE,
    CATEGORY_BOTTOM_STRIP,
    CATEGORY_LOCAL_TAB,
    CATEGORY_OVERLAY,
    CATEGORY_SCROLL_RETRY,
    CATEGORY_VIEWPORT_EXHAUSTED,
    CATEGORY_UNKNOWN,
]

SCENARIO_FIELDS = [
    "scenario_id",
    "runtime_sec",
    "attempted_step_count",
    "persisted_step_count",
    "suppressed_row_count",
    "nonstep_result_row_count",
    "productive_count",
    "unique_meaningful_count",
    "duplicate_count",
    "repeat_no_progress_count",
    "bottom_strip_churn_count",
    "local_tab_churn_count",
    "overlay_churn_count",
    "scroll_retry_churn_count",
    "viewport_exhausted_churn_count",
    "unknown_count",
    "churn_ratio",
    "productive_ratio",
    "stop_reason",
]

RUN_STEM_RE = re.compile(r"(talkback_compare_\d{8}_\d{6})")
SCENARIO_START_RE = re.compile(
    r"\[SCENARIO\]\[(?:stabilization|start|begin)\].*?\bscenario=(?:'([^']+)'|([^\s]+))"
)
SCENARIO_FIELD_RE = re.compile(r"\bscenario=(?:'([^']+)'|([^\s]+))")
PERF_SUMMARY_RE = re.compile(r"\[PERF\]\[scenario_summary\].*?\bscenario=(?:'([^']+)'|([^\s]+))")
STOP_SUMMARY_RE = re.compile(
    r"\[STOP\]\[summary\].*?\bscenario=(?:'([^']+)'|([^\s]+)).*?\breason=(?:'([^']*)'|(\S+))"
)
STEP_RE = re.compile(r"\bstep=(\d+)")
TIMESTAMP_RE = re.compile(r"^\[(\d{2}):(\d{2}):(\d{2})\]")
TOTAL_RUNTIME_RE = re.compile(r"\btotal_runtime=(\d+(?:\.\d+)?)")
TOTAL_STEPS_RE = re.compile(r"\btotal_steps=(\d+)")


def _match_value(match: re.Match[str] | None) -> str | None:
    if not match:
        return None
    return next((value for value in match.groups() if value), None)


def _clock_seconds(line: str) -> int | None:
    match = TIMESTAMP_RE.match(line)
    if not match:
        return None
    hour, minute, second = (int(value) for value in match.groups())
    return hour * 3600 + minute * 60 + second


def _artifact_stem(*paths: Path) -> str:
    for path in paths:
        match = RUN_STEM_RE.search(path.name)
        if match:
            return match.group(1)
    return "talkback_compare_unknown"


def _normalize_text(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _display(value: Any) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, float):
        return f"{value:.3f}".rstrip("0").rstrip(".")
    return str(value)


def _empty_step(scenario_id: str, step: int) -> dict[str, Any]:
    return {
        "scenario_id": scenario_id,
        "step": step,
        "start_clock": None,
        "end_clock": None,
        "lines": [],
        "rows": [],
        "classification": CATEGORY_UNKNOWN,
        "reason": "no_signals",
        "evidence": [],
        "signals": {
            "duplicate": False,
            "bottom_strip": False,
            "local_tab": False,
            "overlay": False,
            "scroll_retry": False,
            "viewport_exhausted": False,
            "repeat_no_progress": False,
            "productive_row": False,
        },
    }


def _empty_scenario(scenario_id: str) -> dict[str, Any]:
    row: dict[str, Any] = {
        "scenario_id": scenario_id,
        "runtime_sec": None,
        "attempted_step_count": 0,
        "persisted_step_count": 0,
        "suppressed_row_count": 0,
        "nonstep_result_row_count": 0,
        "result_row_count": 0,
        "productive_count": 0,
        "unique_meaningful_count": 0,
        "duplicate_count": 0,
        "repeat_no_progress_count": 0,
        "bottom_strip_churn_count": 0,
        "local_tab_churn_count": 0,
        "overlay_churn_count": 0,
        "scroll_retry_churn_count": 0,
        "viewport_exhausted_churn_count": 0,
        "unknown_count": 0,
        "churn_ratio": 0.0,
        "productive_ratio": 0.0,
        "stop_reason": "",
        "steps": [],
        "_step_map": {},
        "_repeat_keys": set(),
        "_meaningful_signatures": set(),
        "_persisted_step_ids": set(),
        "_started_step_ids": set(),
    }
    return row


def _scenario(
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    scenario_id: str | None,
) -> dict[str, Any] | None:
    normalized = str(scenario_id or "").strip().strip("'\"")
    if not normalized:
        return None
    if normalized not in scenarios:
        scenarios[normalized] = _empty_scenario(normalized)
        execution_order.append(normalized)
    return scenarios[normalized]


def _step_record(
    scenario: dict[str, Any],
    step: int,
) -> dict[str, Any]:
    step_map = scenario["_step_map"]
    if step not in step_map:
        step_map[step] = _empty_step(scenario["scenario_id"], step)
    return step_map[step]


def parse_xlsx(
    path: Path,
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    warnings: list[str],
) -> None:
    if not path.is_file():
        warnings.append(f"xlsx file not found: {path}")
        return
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        warnings.append(f"xlsx could not be read: {path} ({exc})")
        return

    try:
        if "result" not in workbook.sheetnames:
            warnings.append(f"xlsx result sheet not found: {path}")
            return
        sheet = workbook["result"]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, ())
        index = {
            str(value).strip(): position
            for position, value in enumerate(headers)
            if value is not None and str(value).strip()
        }
        if "scenario_id" not in index or "step" not in index:
            warnings.append(f"xlsx result sheet missing scenario_id/step columns: {path}")
            return

        for values in rows:
            scenario_id = _cell(values, index, "scenario_id")
            scenario = _scenario(scenarios, execution_order, scenario_id)
            if scenario is None:
                continue
            row = {
                "scenario_id": scenario_id,
                "step": _int_value(_cell(values, index, "step")),
                "context_type": _cell(values, index, "context_type"),
                "visible_label": _cell(values, index, "visible_label"),
                "merged_announcement": _cell(values, index, "merged_announcement"),
                "row_source": _cell(values, index, "row_source"),
                "final_result": _cell(values, index, "final_result"),
                "focus_view_id": _cell(values, index, "focus_view_id"),
            }
            step = row["step"]
            if step is None or step < 0:
                scenario["nonstep_result_row_count"] += 1
                continue
            record = _step_record(scenario, step)
            record["rows"].append(row)
            scenario["_persisted_step_ids"].add(step)
            scenario["result_row_count"] += 1
    finally:
        workbook.close()


def _cell(values: tuple[Any, ...], index: dict[str, int], key: str) -> str:
    position = index.get(key)
    if position is None or position >= len(values):
        return ""
    value = values[position]
    return str(value or "").strip()


def _int_value(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_log(
    path: Path,
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    warnings: list[str],
) -> None:
    if not path.is_file():
        warnings.append(f"normal log file not found: {path}")
        return
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError as exc:
        warnings.append(f"normal log could not be read: {path} ({exc})")
        return

    current_scenario: str | None = None
    previous_clock: int | None = None
    day_offset = 0
    active_step_by_scenario: dict[str, int] = {}
    scenario_start_seconds: dict[str, int] = {}
    scenario_end_seconds: dict[str, int] = {}

    for line in lines:
        clock = _clock_seconds(line)
        if clock is not None:
            if previous_clock is not None and clock < previous_clock:
                day_offset += 24 * 60 * 60
            previous_clock = clock
            clock += day_offset

        start_match = SCENARIO_START_RE.search(line)
        if start_match:
            current_scenario = _match_value(start_match)
            scenario = _scenario(scenarios, execution_order, current_scenario)
            if scenario is not None and clock is not None:
                scenario_start_seconds.setdefault(scenario["scenario_id"], clock)

        perf_match = PERF_SUMMARY_RE.search(line)
        explicit_scenario = _match_value(perf_match or SCENARIO_FIELD_RE.search(line))
        scenario_id = explicit_scenario or current_scenario
        scenario = _scenario(scenarios, execution_order, scenario_id)
        if scenario is None:
            continue

        if "[STEP] START" in line:
            step_match = STEP_RE.search(line)
            if step_match:
                step = int(step_match.group(1))
                record = _step_record(scenario, step)
                record["start_clock"] = clock if record["start_clock"] is None else record["start_clock"]
                record["lines"].append(line)
                active_step_by_scenario[scenario["scenario_id"]] = step
                scenario["_started_step_ids"].add(step)
                continue

        if "[STEP] END" in line:
            step_match = STEP_RE.search(line)
            if step_match:
                step = int(step_match.group(1))
                record = _step_record(scenario, step)
                record["end_clock"] = clock
                record["lines"].append(line)
                active_step_by_scenario[scenario["scenario_id"]] = step
                continue

        assigned_step = None
        step_match = STEP_RE.search(line)
        if scenario["scenario_id"] in active_step_by_scenario:
            assigned_step = active_step_by_scenario[scenario["scenario_id"]]
        elif step_match:
            assigned_step = int(step_match.group(1))

        if assigned_step is not None and assigned_step >= 0:
            record = _step_record(scenario, assigned_step)
            record["lines"].append(line)

        if "[STOP][eval]" in line and "repeat_no_progress" in line:
            repeat_key = (
                scenario["scenario_id"],
                step_match.group(1) if step_match else str(len(scenario["_repeat_keys"])),
            )
            if repeat_key not in scenario["_repeat_keys"]:
                scenario["_repeat_keys"].add(repeat_key)
                scenario["repeat_no_progress_count"] += 1
            if assigned_step is not None and assigned_step >= 0:
                record = _step_record(scenario, assigned_step)
                record["signals"]["repeat_no_progress"] = True

        stop_match = STOP_SUMMARY_RE.search(line)
        if stop_match:
            stop_scenario = next((value for value in stop_match.groups()[:2] if value), None)
            stop_reason = next((value for value in stop_match.groups()[2:] if value is not None), "")
            stop_target = _scenario(scenarios, execution_order, stop_scenario)
            if stop_target is not None:
                stop_target["stop_reason"] = str(stop_reason or "").strip().strip("'\"")
                if clock is not None:
                    scenario_end_seconds[stop_target["scenario_id"]] = clock

        if perf_match:
            runtime_match = TOTAL_RUNTIME_RE.search(line)
            steps_match = TOTAL_STEPS_RE.search(line)
            if runtime_match:
                scenario["runtime_sec"] = float(runtime_match.group(1))
            if steps_match:
                scenario["persisted_step_count"] = int(steps_match.group(1))
            if clock is not None:
                scenario_end_seconds[scenario["scenario_id"]] = clock

    for scenario_id, scenario in scenarios.items():
        if scenario["runtime_sec"] is None:
            start = scenario_start_seconds.get(scenario_id)
            end = scenario_end_seconds.get(scenario_id)
            if start is not None and end is not None and end >= start:
                scenario["runtime_sec"] = float(end - start)


def _row_signature(row: dict[str, Any]) -> str:
    visible = _normalize_text(row.get("visible_label", ""))
    merged = _normalize_text(row.get("merged_announcement", ""))
    focus = _normalize_text(row.get("focus_view_id", ""))
    parts = [part for part in (visible, merged, focus) if part]
    return " || ".join(parts)


def _primary_row(rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    traversal_rows = [
        row
        for row in rows
        if str(row.get("row_source", "") or "").strip().upper() in TRAVERSAL_ROW_SOURCES
    ]
    if traversal_rows:
        return traversal_rows[0]
    return rows[0] if rows else None


def _rows_have_meaningful_content(rows: list[dict[str, Any]]) -> bool:
    for row in rows:
        row_source = str(row.get("row_source", "") or "").strip().upper()
        if row_source not in TRAVERSAL_ROW_SOURCES:
            continue
        if _row_signature(row):
            return True
    return False


def _rows_look_like_bottom_strip(rows: list[dict[str, Any]]) -> bool:
    for row in rows:
        visible = _normalize_text(row.get("visible_label", ""))
        merged = _normalize_text(row.get("merged_announcement", ""))
        if visible in BOTTOM_STRIP_LABELS or merged in BOTTOM_STRIP_LABELS:
            return True
    return False


def _extract_step_signals(step: dict[str, Any]) -> tuple[dict[str, bool], list[str]]:
    signals = dict(step["signals"])
    evidence: list[str] = []
    rows = step["rows"]
    has_meaningful_row = _rows_have_meaningful_content(rows)
    signals["productive_row"] = has_meaningful_row

    for row in rows:
        if str(row.get("context_type", "") or "").strip().lower() == "overlay":
            signals["overlay"] = True
            evidence.append("xlsx:context_type=overlay")

    for line in step["lines"]:
        normalized = line.lower()
        if "[step][row_filter]" in normalized:
            signals["duplicate"] = True
            reason = _extract_field(line, "reason")
            evidence.append(f"log:row_filter:{reason or 'unknown'}")
        if "[overlay]" in normalized:
            signals["overlay"] = True
            if "blocked by scenario policy" not in normalized:
                evidence.append("log:overlay_activity")
        if "bottom_strip" in normalized or "strip_focus_context=true" in normalized:
            signals["bottom_strip"] = True
            evidence.append("log:bottom_strip_signal")
        if "[step][local_tab_" in normalized or "local_tab_gate" in normalized:
            signals["local_tab"] = True
            reason = _extract_field(line, "reason")
            evidence.append(f"log:local_tab:{reason or 'signal'}")
        if "forced_local_tab_navigation" in normalized:
            signals["local_tab"] = True
            evidence.append("log:forced_local_tab_navigation")
        if "scroll_fallback" in normalized or "last_scroll_fallback" in normalized:
            if "not_evaluated" not in normalized:
                signals["scroll_retry"] = True
                reason = _extract_field(line, "reason") or _extract_field(line, "scroll_fallback_block_reason")
                evidence.append(f"log:scroll_fallback:{reason or 'signal'}")
        if "[step][viewport_exhausted_eval]" in normalized and "result=true" in normalized:
            signals["viewport_exhausted"] = True
            reason = _extract_field(line, "reason")
            evidence.append(f"log:viewport_exhausted:{reason or 'true'}")

    deduped = []
    seen = set()
    for item in evidence:
        if item not in seen:
            seen.add(item)
            deduped.append(item)
    return signals, deduped


def _extract_field(line: str, key: str) -> str:
    match = re.search(rf"\b{re.escape(key)}=(?:'([^']*)'|(\S+))", line)
    if not match:
        return ""
    return next((value for value in match.groups() if value is not None), "")


def classify_step(step: dict[str, Any]) -> None:
    signals, evidence = _extract_step_signals(step)
    rows = step["rows"]
    has_meaningful_row = signals["productive_row"]
    rows_look_strip = _rows_look_like_bottom_strip(rows)

    if signals["duplicate"]:
        category = CATEGORY_DUPLICATE
        reason = next((item for item in evidence if item.startswith("log:row_filter:")), "log:row_filter")
    elif signals["overlay"] and not has_meaningful_row:
        category = CATEGORY_OVERLAY
        reason = "overlay_without_persisted_content"
    elif signals["bottom_strip"] and (rows_look_strip or not has_meaningful_row):
        category = CATEGORY_BOTTOM_STRIP
        reason = "bottom_strip_navigation_only"
    elif signals["local_tab"] and not has_meaningful_row:
        category = CATEGORY_LOCAL_TAB
        reason = "local_tab_navigation_only"
    elif signals["scroll_retry"] and not has_meaningful_row:
        category = CATEGORY_SCROLL_RETRY
        reason = "scroll_retry_without_content"
    elif signals["viewport_exhausted"] and not has_meaningful_row:
        category = CATEGORY_VIEWPORT_EXHAUSTED
        reason = "viewport_exhausted_without_content"
    elif has_meaningful_row:
        category = CATEGORY_PRODUCTIVE
        reason = "persisted_content_row"
    elif signals["overlay"]:
        category = CATEGORY_OVERLAY
        reason = "overlay_signal_only"
    elif signals["bottom_strip"]:
        category = CATEGORY_BOTTOM_STRIP
        reason = "bottom_strip_signal_only"
    elif signals["local_tab"]:
        category = CATEGORY_LOCAL_TAB
        reason = "local_tab_signal_only"
    elif signals["scroll_retry"]:
        category = CATEGORY_SCROLL_RETRY
        reason = "scroll_retry_signal_only"
    elif signals["viewport_exhausted"]:
        category = CATEGORY_VIEWPORT_EXHAUSTED
        reason = "viewport_exhausted_signal_only"
    else:
        category = CATEGORY_UNKNOWN
        reason = "no_classification_signal"

    primary = _primary_row(rows)
    step["classification"] = category
    step["reason"] = reason
    step["signals"] = signals
    step["evidence"] = evidence
    step["row_signature"] = _row_signature(primary or {})
    step["visible_label"] = str((primary or {}).get("visible_label", "") or "")
    step["merged_announcement"] = str((primary or {}).get("merged_announcement", "") or "")


def finalize_scenarios(
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for scenario_id in execution_order:
        scenario = scenarios[scenario_id]
        started_step_ids = set(scenario["_started_step_ids"])
        steps = [
            scenario["_step_map"][step]
            for step in sorted(scenario["_step_map"])
            if step > 0 and (step in started_step_ids or scenario["_step_map"][step]["rows"])
        ]
        scenario["steps"] = steps
        if started_step_ids:
            scenario["attempted_step_count"] = len(started_step_ids)
        elif scenario["attempted_step_count"] == 0:
            scenario["attempted_step_count"] = len(steps)
        scenario["productive_count"] = 0
        scenario["duplicate_count"] = 0
        scenario["bottom_strip_churn_count"] = 0
        scenario["local_tab_churn_count"] = 0
        scenario["overlay_churn_count"] = 0
        scenario["scroll_retry_churn_count"] = 0
        scenario["viewport_exhausted_churn_count"] = 0
        scenario["unknown_count"] = 0
        for step in steps:
            classify_step(step)
            category = step["classification"]
            if category == CATEGORY_PRODUCTIVE:
                scenario["productive_count"] += 1
                if step["row_signature"]:
                    scenario["_meaningful_signatures"].add(step["row_signature"])
            elif category == CATEGORY_DUPLICATE:
                scenario["duplicate_count"] += 1
            elif category == CATEGORY_BOTTOM_STRIP:
                scenario["bottom_strip_churn_count"] += 1
            elif category == CATEGORY_LOCAL_TAB:
                scenario["local_tab_churn_count"] += 1
            elif category == CATEGORY_OVERLAY:
                scenario["overlay_churn_count"] += 1
            elif category == CATEGORY_SCROLL_RETRY:
                scenario["scroll_retry_churn_count"] += 1
            elif category == CATEGORY_VIEWPORT_EXHAUSTED:
                scenario["viewport_exhausted_churn_count"] += 1
            else:
                scenario["unknown_count"] += 1

        scenario["unique_meaningful_count"] = len(scenario["_meaningful_signatures"])
        persisted_steps = len(scenario["_persisted_step_ids"])
        if persisted_steps > 0:
            scenario["persisted_step_count"] = persisted_steps
        elif scenario["persisted_step_count"] == 0:
            scenario["persisted_step_count"] = sum(1 for step in steps if step["rows"])
        if scenario["attempted_step_count"] == 0:
            scenario["attempted_step_count"] = len(steps)
        scenario["suppressed_row_count"] = max(
            0,
            int(scenario["attempted_step_count"]) - int(scenario["persisted_step_count"]),
        )
        churn_total = (
            scenario["duplicate_count"]
            + scenario["bottom_strip_churn_count"]
            + scenario["local_tab_churn_count"]
            + scenario["overlay_churn_count"]
            + scenario["scroll_retry_churn_count"]
            + scenario["viewport_exhausted_churn_count"]
        )
        attempted = int(scenario["attempted_step_count"])
        scenario["churn_ratio"] = round(churn_total / attempted, 4) if attempted else 0.0
        scenario["productive_ratio"] = round(scenario["productive_count"] / attempted, 4) if attempted else 0.0
        scenario["category_counts"] = {
            CATEGORY_PRODUCTIVE: scenario["productive_count"],
            CATEGORY_DUPLICATE: scenario["duplicate_count"],
            CATEGORY_BOTTOM_STRIP: scenario["bottom_strip_churn_count"],
            CATEGORY_LOCAL_TAB: scenario["local_tab_churn_count"],
            CATEGORY_OVERLAY: scenario["overlay_churn_count"],
            CATEGORY_SCROLL_RETRY: scenario["scroll_retry_churn_count"],
            CATEGORY_VIEWPORT_EXHAUSTED: scenario["viewport_exhausted_churn_count"],
            CATEGORY_UNKNOWN: scenario["unknown_count"],
        }
        scenario["is_priority_scenario"] = scenario_id in PRIMARY_SCENARIOS
        del scenario["_step_map"]
        del scenario["_repeat_keys"]
        del scenario["_meaningful_signatures"]
        del scenario["_persisted_step_ids"]
        del scenario["_started_step_ids"]
        rows.append(scenario)
    return rows


def build_summary(scenarios: list[dict[str, Any]]) -> dict[str, Any]:
    durations = [float(row["runtime_sec"]) for row in scenarios if row["runtime_sec"] is not None]
    total_churn_steps = sum(
        int(row["duplicate_count"])
        + int(row["bottom_strip_churn_count"])
        + int(row["local_tab_churn_count"])
        + int(row["overlay_churn_count"])
        + int(row["scroll_retry_churn_count"])
        + int(row["viewport_exhausted_churn_count"])
        for row in scenarios
    )
    return {
        "scenario_count": len(scenarios),
        "total_runtime_sec": round(sum(durations), 3) if durations else None,
        "total_attempted_steps": sum(int(row["attempted_step_count"]) for row in scenarios),
        "total_persisted_steps": sum(int(row["persisted_step_count"]) for row in scenarios),
        "total_suppressed_rows": sum(int(row["suppressed_row_count"]) for row in scenarios),
        "total_productive_steps": sum(int(row["productive_count"]) for row in scenarios),
        "total_duplicate_steps": sum(int(row["duplicate_count"]) for row in scenarios),
        "total_repeat_no_progress": sum(int(row["repeat_no_progress_count"]) for row in scenarios),
        "total_churn_steps": total_churn_steps,
        "total_unknown_steps": sum(int(row["unknown_count"]) for row in scenarios),
        "priority_scenarios": [scenario_id for scenario_id in PRIMARY_SCENARIOS if any(row["scenario_id"] == scenario_id for row in scenarios)],
        "slowest_scenarios": _rank(scenarios, "runtime_sec", include_zero=False),
        "highest_churn_ratio": _rank(scenarios, "churn_ratio", include_zero=False),
        "highest_duplicate_scenarios": _rank(scenarios, "duplicate_count", include_zero=False),
    }


def _rank(
    scenarios: list[dict[str, Any]],
    field: str,
    *,
    include_zero: bool,
    limit: int = 10,
) -> list[dict[str, Any]]:
    ranked = [
        {"scenario_id": row["scenario_id"], "value": row[field]}
        for row in scenarios
        if row.get(field) is not None and (include_zero or float(row[field]) > 0)
    ]
    ranked.sort(key=lambda item: (-float(item["value"]), item["scenario_id"]))
    return ranked[:limit]


def build_findings(scenarios: list[dict[str, Any]]) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    for row in scenarios:
        scenario_id = row["scenario_id"]
        if float(row["churn_ratio"]) >= 0.6:
            findings.append(
                _finding(
                    scenario_id,
                    "high_churn_ratio",
                    f"Churn ratio is {_display(row['churn_ratio'])} with stop_reason='{row['stop_reason'] or 'none'}'.",
                )
            )
        if int(row["duplicate_count"]) >= 3:
            findings.append(
                _finding(
                    scenario_id,
                    "duplicate_churn_cluster",
                    f"Duplicate churn occurred {row['duplicate_count']} times.",
                )
            )
        if int(row["viewport_exhausted_churn_count"]) >= 3:
            findings.append(
                _finding(
                    scenario_id,
                    "repeated_viewport_exhausted",
                    f"Viewport exhausted churn occurred {row['viewport_exhausted_churn_count']} times.",
                )
            )
    return findings


def _finding(scenario_id: str, code: str, message: str) -> dict[str, Any]:
    return {
        "severity": "WARN",
        "scenario_id": scenario_id,
        "code": code,
        "message": message,
    }


def build_audit(
    *,
    xlsx: Path,
    log: Path,
    profile_json: Path | None = None,
) -> dict[str, Any]:
    warnings: list[str] = []
    scenarios: dict[str, dict[str, Any]] = {}
    execution_order: list[str] = []

    parse_xlsx(xlsx, scenarios, execution_order, warnings)
    parse_log(log, scenarios, execution_order, warnings)
    scenario_rows = finalize_scenarios(scenarios, execution_order)
    summary = build_summary(scenario_rows)
    findings = build_findings(scenario_rows)

    profile_summary = None
    if profile_json and profile_json.is_file():
        try:
            payload = json.loads(profile_json.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            warnings.append(f"profile json could not be read: {profile_json} ({exc})")
        else:
            if isinstance(payload, dict):
                profile_summary = payload.get("summary")

    return {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "artifact_stem": _artifact_stem(xlsx, log),
        "inputs": {
            "xlsx": str(xlsx),
            "log": str(log),
            "profile_json": str(profile_json) if profile_json else "",
        },
        "warnings": warnings,
        "summary": summary,
        "profile_summary": profile_summary,
        "scenarios": scenario_rows,
        "findings": findings,
    }


def write_outputs(audit: dict[str, Any], output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / AUDIT_JSON
    markdown_path = output_dir / AUDIT_MARKDOWN
    csv_path = output_dir / AUDIT_CSV

    json_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(render_markdown(audit), encoding="utf-8")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=SCENARIO_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(audit["scenarios"])
    return {"json": json_path, "markdown": markdown_path, "csv": csv_path}


def render_markdown(audit: dict[str, Any]) -> str:
    summary = audit["summary"]
    scenarios = audit["scenarios"]
    lines = [
        "# V9 Traversal Churn Audit",
        "",
        f"Artifact: `{audit.get('artifact_stem', '')}`",
        "",
        "## Run Summary",
        "",
        "| Metric | Value |",
        "| --- | ---: |",
    ]
    for key in (
        "scenario_count",
        "total_runtime_sec",
        "total_attempted_steps",
        "total_persisted_steps",
        "total_suppressed_rows",
        "total_productive_steps",
        "total_duplicate_steps",
        "total_repeat_no_progress",
        "total_churn_steps",
        "total_unknown_steps",
    ):
        lines.append(f"| {key} | {_display(summary.get(key))} |")

    lines.extend(
        [
            "",
            "## Scenario Table",
            "",
            "| Scenario | Runtime (s) | Attempted | Persisted | Suppressed | Productive | Duplicate | Repeat | Churn ratio | Stop reason |",
            "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | --- |",
        ]
    )
    for row in scenarios:
        lines.append(
            f"| {row['scenario_id']} | {_display(row['runtime_sec'])} | {row['attempted_step_count']} "
            f"| {row['persisted_step_count']} | {row['suppressed_row_count']} | {row['productive_count']} "
            f"| {row['duplicate_count']} | {row['repeat_no_progress_count']} | {_display(row['churn_ratio'])} "
            f"| {row['stop_reason'] or 'none'} |"
        )

    priority = [row for row in scenarios if row["scenario_id"] in PRIMARY_SCENARIOS]
    lines.extend(
        [
            "",
            "## Priority Scenarios",
            "",
            "| Scenario | Bottom strip | Local tab | Overlay | Scroll retry | Viewport exhausted | Unique meaningful |",
            "| --- | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in priority:
        lines.append(
            f"| {row['scenario_id']} | {row['bottom_strip_churn_count']} | {row['local_tab_churn_count']} "
            f"| {row['overlay_churn_count']} | {row['scroll_retry_churn_count']} | {row['viewport_exhausted_churn_count']} "
            f"| {row['unique_meaningful_count']} |"
        )

    lines.extend(
        [
            "",
            "## Category Definitions",
            "",
            "- `PRODUCTIVE`: the step persisted traversal content and did not collapse into a churn-only signal.",
            "- `DUPLICATE`: the step was filtered as an already-seen logical row.",
            "- `BOTTOM_STRIP_CHURN`: the step revolved around bottom-strip controls or strip-only focus context.",
            "- `LOCAL_TAB_CHURN`: the step stayed in local-tab transition/recovery without new content.",
            "- `OVERLAY_CHURN`: the step was consumed by overlay work without new main content.",
            "- `SCROLL_RETRY_CHURN`: the step spent time in scroll fallback / retry without new content.",
            "- `VIEWPORT_EXHAUSTED_CHURN`: the step hit exhausted viewport logic without persisting new content.",
            "- `UNKNOWN`: no stable classification signal was found.",
            "",
        ]
    )

    _append_ranking(lines, "Top Slow Scenarios", summary["slowest_scenarios"], "s")
    _append_ranking(lines, "Top Churn Ratios", summary["highest_churn_ratio"], "")
    _append_ranking(lines, "Top Duplicate Churn", summary["highest_duplicate_scenarios"], "")

    lines.extend(["## Findings", ""])
    if not audit["findings"]:
        lines.append("- None.")
    else:
        for finding in audit["findings"]:
            lines.append(f"- `{finding['scenario_id']}` `{finding['code']}`: {finding['message']}")
    lines.append("")
    return "\n".join(lines)


def _append_ranking(
    lines: list[str],
    title: str,
    ranking: list[dict[str, Any]],
    suffix: str,
) -> None:
    lines.append(f"### {title}")
    lines.append("")
    if not ranking:
        lines.append("- No data.")
    else:
        for item in ranking:
            lines.append(f"- `{item['scenario_id']}`: {_display(item['value'])}{suffix}")
    lines.append("")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path)
    parser.add_argument("--log", type=Path)
    parser.add_argument("--profile-json", type=Path)
    parser.add_argument("--artifact-dir", type=Path)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args(argv)
    if not args.artifact_dir and not any((args.xlsx, args.log)):
        parser.error("provide --artifact-dir or both --xlsx and --log")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    detected: dict[str, Path] = {}
    if args.artifact_dir:
        try:
            detected = resolve_artifacts(args.artifact_dir)
        except FileNotFoundError as exc:
            print(f"error: {exc}", file=sys.stderr)
            return 2

    explicit_paths = [path for path in (args.xlsx, args.log) if path is not None]
    base_dir = explicit_paths[0].parent if explicit_paths else next(iter(detected.values())).parent
    xlsx = args.xlsx or detected.get("xlsx") or base_dir / "talkback_compare_unknown.xlsx"
    log = args.log or detected.get("log") or base_dir / "talkback_compare_unknown.normal.log"
    profile_json = args.profile_json
    if profile_json is None and args.artifact_dir:
        candidates = sorted(args.artifact_dir.rglob("v8_fullrun_profile.json"))
        if candidates:
            profile_json = candidates[-1]
    output_dir = args.output_dir or base_dir / f"v9_churn_{_artifact_stem(xlsx, log).removeprefix('talkback_compare_')}"

    audit = build_audit(xlsx=xlsx, log=log, profile_json=profile_json)
    outputs = write_outputs(audit, output_dir)
    print(f"audited {audit['summary']['scenario_count']} scenarios")
    print(
        f"attempted={audit['summary']['total_attempted_steps']} "
        f"persisted={audit['summary']['total_persisted_steps']} "
        f"churn={audit['summary']['total_churn_steps']}"
    )
    for label, path in outputs.items():
        print(f"{label}={path}")
    if audit["warnings"]:
        print(f"warnings={len(audit['warnings'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
