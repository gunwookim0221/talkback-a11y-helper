from collections import Counter
from math import ceil
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .models import RunMetadata, SourceRow
from .source import deterministic_pass_sample, read_pass_rows, read_review_rows

QA_REVIEW_COLUMNS = [
    "Review ID", "Scenario", "Focus Target", "Approximate Position", "Review Description",
    "Screenshot", "TalkBack Speech", "Visible Text", "Validator Checklist", "Validator Comment",
    "Step",
]
DECISIONS = ["미검토", "정상 발화", "실제 접근성 문제", "False Positive", "재현 불가", "추가 조사 필요"]
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")


def _style_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, widths: dict[int, int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column, width in widths.items():
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row[0].row].height = 60


def _setup_workbook() -> Workbook:
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    return workbook


def _source_counts(source: Path) -> Counter[str]:
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["result"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    headers = [str(value or "") for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    result_index = headers.index("final_result")
    counts = Counter(str(row[result_index] or "").upper() for row in rows)
    workbook.close()
    return counts


def _write_summary(sheet: openpyxl.worksheet.worksheet.Worksheet, metadata: RunMetadata, counts: Counter[str], qa_rows: list[SourceRow], diagnostics: list[SourceRow]) -> None:
    sheet.append(["Metric", "Value"])
    review_count = len(qa_rows)
    scenario_counts = Counter(row.scenario_id for row in qa_rows)
    position_counts = Counter(row.approximate_position for row in qa_rows)
    unknown_count = sum(row.focus_target == "Unknown" for row in qa_rows)
    resource_target_count = sum(row.focus_target_source == "resource id" for row in qa_rows)
    screenshot_missing = sum(row.screenshot_evidence_type == "No screenshot" for row in qa_rows)
    values: list[tuple[str, object]] = [
        ("Run ID", metadata.run_id), ("Batch ID", metadata.batch_id), ("Device", metadata.device_model),
        ("Android / One UI", f"{metadata.android_version} / {metadata.one_ui_version}"),
        ("TalkBack", metadata.talkback), ("App version", metadata.app), ("Locale", metadata.locale),
        ("Total raw rows", sum(counts.values())), ("PASS count", counts["PASS"]), ("WARN count", counts["WARN"]),
        ("FAIL count", counts["FAIL"]), ("QA Review Count", review_count),
        ("Automation Diagnostic Count", len(diagnostics)), ("Scenario Count", len(scenario_counts)),
        ("QA 예상 검토 시간", f"약 {ceil(review_count * 1.5)}분"), ("Unknown Target count", unknown_count),
        ("Screenshot 없는 항목 count", screenshot_missing),
        ("Resource-derived Target count", resource_target_count),
        ("정상 발화", '=COUNTIF(\'Review Checklist\'!$I:$I,"정상 발화")'),
        ("실제 접근성 문제", '=COUNTIF(\'Review Checklist\'!$I:$I,"실제 접근성 문제")'),
        ("False Positive", '=COUNTIF(\'Review Checklist\'!$I:$I,"False Positive")'),
        ("재현 불가", '=COUNTIF(\'Review Checklist\'!$I:$I,"재현 불가")'),
        ("추가 조사 필요", '=COUNTIF(\'Review Checklist\'!$I:$I,"추가 조사 필요")'),
        ("미검토", '=COUNTIF(\'Review Checklist\'!$I:$I,"미검토")'),
        ("Review completion %", '=IF(B13=0,1,(B18+B19+B20+B21+B22)/B13)'),
        ("Overall Human Review Status", '=IF(B13=0,"COMPLETED_NO_ISSUE",IF(B23=B13,"NOT_STARTED",IF(B23>0,"IN_PROGRESS",IF(B22>0,"RETEST_REQUIRED",IF(B19>0,"COMPLETED_WITH_ISSUES","COMPLETED_NO_ISSUE")))))'),
    ]
    for item in values:
        sheet.append(list(item))
    sheet["B25"].number_format = "0%"
    sheet.append([])
    sheet.append(["Scenario", "FAIL count"])
    for scenario, count in sorted(scenario_counts.items()):
        sheet.append([scenario, count])
    sheet.append([])
    sheet.append(["Focus 위치", "FAIL count"])
    for position, count in sorted(position_counts.items()):
        sheet.append([position, count])
    _style_sheet(sheet, {1: 28, 2: 40})
    sheet.freeze_panes = "A2"


def _add_screenshot_thumbnail(sheet: openpyxl.worksheet.worksheet.Worksheet, row_number: int, annotation: str, source: Path) -> bool:
    annotation_path = source.parent / annotation
    if not annotation or not annotation_path.is_file():
        return False
    try:
        image = ExcelImage(str(annotation_path))
    except (AttributeError, ImportError, OSError, ValueError):
        return False
    ratio = image.height / image.width if image.width else 1
    image.width = 200
    image.height = max(1, round(200 * ratio))
    sheet.add_image(image, f"F{row_number}")
    sheet.row_dimensions[row_number].height = max(155, image.height * 0.75 + 8)
    return True


def _write_checklist(sheet: openpyxl.worksheet.worksheet.Worksheet, rows: list[SourceRow], metadata: RunMetadata, source: Path) -> None:
    sheet.append(QA_REVIEW_COLUMNS)
    for index, row in enumerate(rows, start=1):
        values = [
            f"{metadata.run_id}-R{index:03d}", row.scenario_name, row.focus_target, row.approximate_position,
            row.review_description, row.screenshot, row.speech, row.visible_text, row.validator_decision,
            row.validator_comment, row.step,
        ]
        sheet.append(values)
        screenshot_cell = sheet.cell(sheet.max_row, 6)
        screenshot_path = source.parent / row.screenshot if row.screenshot_evidence_type != "No screenshot" else None
        if screenshot_path is not None and screenshot_path.is_file():
            screenshot_cell.value = f"{row.screenshot_evidence_type}\nOpen original screenshot"
            screenshot_cell.hyperlink = row.screenshot.replace("\\", "/")
            screenshot_cell.style = "Hyperlink"
        else:
            screenshot_cell.value = "Full screenshot not captured for this run"
    validation = DataValidation(type="list", formula1='"미검토,정상 발화,실제 접근성 문제,False Positive,재현 불가,추가 조사 필요"', allow_blank=False)
    validation.error = "Select a permitted validator decision."
    validation.errorTitle = "Invalid decision"
    sheet.add_data_validation(validation)
    validation.add(f"I2:I{max(2, sheet.max_row)}")
    sheet.conditional_formatting.add(f"A2:K{max(2, sheet.max_row)}", FormulaRule(formula=['$I2="실제 접근성 문제"'], fill=FAIL_FILL))
    _style_sheet(sheet, {1: 28, 2: 28, 3: 28, 4: 24, 5: 58, 6: 32, 7: 42, 8: 34, 9: 20, 10: 34, 11: 8})
    for row_number, row in enumerate(rows, start=2):
        if row.screenshot_annotation:
            _add_screenshot_thumbnail(sheet, row_number, row.screenshot_annotation, source)
    sheet.column_dimensions["K"].hidden = True


def _write_additional(sheet: openpyxl.worksheet.worksheet.Worksheet, warnings: list[dict[str, str]]) -> None:
    sheet.append(["Scenario ID", "WARN Type", "Representative Step", "Final Terminal/Result", "Count", "Actual FAIL Companion"])
    for warning in warnings:
        sheet.append([warning["scenario"], warning["warning_type"], warning["step"], warning["terminal"], int(warning["count"]), warning["actual_fail"]])
    _style_sheet(sheet, {1: 34, 2: 30, 3: 20, 4: 24, 5: 12, 6: 22})


def _write_automation_diagnostic(sheet: openpyxl.worksheet.worksheet.Worksheet, rows: list[SourceRow], source: Path) -> None:
    headers = ["Scenario", "Issue", "Reason", "Step", "Traversal State", "Recovery State", "Terminal State", "Resource ID", "Bounds", "Focus Center / Relative %", "Automatic Result", "Issue Type", "TalkBack Speech", "Visible Text", "Expected/Reference Text or Speech", "Developer Evidence", "Notes"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.scenario_name, row.classification_reason, row.mismatch_reason, row.step, row.traversal_state, row.recovery_state, row.terminal_state, row.resource_id, row.bounds, row.focus_center_relative, row.automatic_result, row.issue_type, row.speech, row.visible_text, row.expected, row.evidence, ""])
        evidence_cell = sheet.cell(sheet.max_row, 16)
        evidence_path = source.parent / row.evidence if row.evidence != "Not available" else None
        if evidence_path is not None and evidence_path.is_file():
            evidence_cell.hyperlink = row.evidence.replace("\\", "/")
            evidence_cell.style = "Hyperlink"
        else:
            evidence_cell.value = "Not available"
    _style_sheet(sheet, {1: 30, 2: 28, 3: 28, 4: 12, 5: 20, 6: 20, 7: 28, 8: 36, 9: 24, 10: 22, 11: 18, 12: 26, 13: 42, 14: 34, 15: 42, 16: 42, 17: 36})


def _review_target(source: Path) -> Path:
    return source.with_name(f"{source.stem}.review.generated.xlsx")


def _can_force_replace(target: Path) -> bool:
    return target.name.endswith(".review.generated.xlsx")


def _contains_decisions(target: Path) -> bool:
    if not target.is_file():
        return False
    workbook = openpyxl.load_workbook(target, read_only=True, data_only=True)
    if "Review Checklist" not in workbook.sheetnames:
        workbook.close()
        return False
    sheet = workbook["Review Checklist"]
    headers = [str(value or "") for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "Validator Checklist" not in headers:
        workbook.close()
        return False
    decision_column = headers.index("Validator Checklist")
    found = any(str(row[decision_column] or "미검토") != "미검토" for row in sheet.iter_rows(min_row=2, values_only=True))
    workbook.close()
    return found


def generate_review_checklist(source: Path, *, output: Path | None = None, pass_sample_rate: float = 0.0, force_regenerate: bool = False) -> Path:
    source = Path(source).resolve()
    metadata, rows, warnings = read_review_rows(source)
    qa_rows = [row for row in rows if row.review_area == "QA"]
    diagnostics = [row for row in rows if row.review_area == "AUTOMATION"]
    target = Path(output).resolve() if output else _review_target(source)
    if _contains_decisions(target):
        return target
    if target.is_file() and not force_regenerate:
        return target
    if target.is_file() and force_regenerate and not _can_force_replace(target):
        return target
    workbook = _setup_workbook()
    checklist = workbook.active
    checklist.title = "Review Checklist"
    _write_checklist(checklist, qa_rows, metadata, source)
    summary = workbook.create_sheet("Summary")
    _write_summary(summary, metadata, _source_counts(source), qa_rows, diagnostics)
    additional = workbook.create_sheet("Additional Review")
    _write_additional(additional, warnings)
    diagnostic = workbook.create_sheet("Automation Diagnostic")
    _write_automation_diagnostic(diagnostic, diagnostics, source)
    if pass_sample_rate > 0:
        audit = workbook.create_sheet("Audit Sample")
        audit.append(["Scenario ID", "Step", "Visible Text", "TalkBack Speech", "Automatic Result", "Sampling Rate"])
        for row in deterministic_pass_sample(source, read_pass_rows(source), pass_sample_rate):
            audit.append([str(row.get("scenario_id") or ""), str(row.get("step") or ""), str(row.get("visible_label") or ""), str(row.get("merged_announcement") or ""), "PASS", pass_sample_rate])
        _style_sheet(audit, {1: 34, 2: 12, 3: 42, 4: 42, 5: 18, 6: 16})
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(f"{target.suffix}.tmp")
    try:
        workbook.save(temporary)
        temporary.replace(target)
    except OSError:
        temporary.unlink(missing_ok=True)
        raise
    return target


def _detail_decisions(path: Path) -> tuple[dict[str, str], Counter[str]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Review Checklist"]
    headers = [str(value or "") for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    positions = {name: headers.index(name) for name in ("Validator Checklist", "Scenario")}
    decisions = Counter(str(row[positions["Validator Checklist"]] or "미검토") for row in sheet.iter_rows(min_row=2, values_only=True))
    metadata_sheet = workbook["Summary"]
    metadata = {str(row[0]): str(row[1] or "") for row in metadata_sheet.iter_rows(min_row=2, values_only=True) if row[0]}
    workbook.close()
    return metadata, decisions


def generate_review_summary(details: list[Path], *, output: Path | None = None) -> Path:
    if not details:
        raise ValueError("At least one detail review file is required.")
    target = Path(output).resolve() if output else details[0].parent / "talkback_review_summary.xlsx"
    workbook = _setup_workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    headers = ["Run ID", "Device", "Android", "One UI", "TalkBack", "Locale", "App Version", "FAIL Review Count", *DECISIONS, "Completion %", "Overall Human Review Status", "Detail Review File Link"]
    sheet.append(headers)
    for detail in sorted(details, key=lambda path: path.name):
        metadata, decisions = _detail_decisions(Path(detail))
        total = sum(decisions.values())
        completion = 1.0 if total == 0 else sum(decisions[name] for name in DECISIONS[1:]) / total
        status = "IN_PROGRESS" if decisions["미검토"] else ("COMPLETED_WITH_ISSUES" if decisions["실제 접근성 문제"] else ("RETEST_REQUIRED" if decisions["추가 조사 필요"] else "COMPLETED_NO_ISSUE"))
        sheet.append([metadata.get("Run ID", Path(detail).stem), metadata.get("Device", ""), metadata.get("Android / One UI", "").split(" / ")[0], metadata.get("Android / One UI", "").split(" / ")[-1], metadata.get("TalkBack", ""), metadata.get("Locale", ""), metadata.get("App version", ""), metadata.get("FAIL count", ""), *[decisions[name] for name in DECISIONS], completion, status, detail.name])
        sheet.cell(sheet.max_row, 16).hyperlink = detail.name
        sheet.cell(sheet.max_row, 16).style = "Hyperlink"
    _style_sheet(sheet, {1: 28, 2: 18, 3: 12, 4: 12, 5: 28, 6: 14, 7: 18, 8: 18, 9: 14, 10: 16, 11: 20, 12: 16, 13: 14, 14: 14, 15: 34, 16: 48})
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 14).number_format = "0%"
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target
