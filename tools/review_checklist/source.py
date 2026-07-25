import hashlib
import json
from pathlib import Path
from typing import Final

import openpyxl

from .models import RunMetadata, SourceRow
from .classification import classify_review
from .images import create_focus_annotation, create_full_screen_annotation
from .qa_context import focus_context
from .speech_status import classify_speech_status, evidence_for_row, speech_evidence_index, speech_review_instruction

_FAIL_RESULTS: Final = {"FAIL"}
_ISSUE_TYPES: Final = {
    "EMPTY_VISIBLE": "화면 텍스트 없음",
    "EMPTY_SPEECH": "TalkBack 발화 없음",
    "NEW_ACCESSIBILITY_FAILURE": "신규 접근성 failure",
}


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _int(value: object) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _nested(payload: dict[str, object], *keys: str) -> str:
    current: object = payload
    for key in keys:
        if not isinstance(current, dict):
            return ""
        current = current.get(key)
    if isinstance(current, dict) and "value" in current:
        current = current.get("value")
    return _text(current)


def _load_json(path: Path) -> dict[str, object]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _metadata(run_root: Path, source: Path) -> RunMetadata:
    summary = _load_json(run_root / "summary.json")
    profile_name = _text(_nested(summary, "environment_profile", "filename")) or f"{source.stem}.environment_profile.json"
    profile = _load_json(run_root / profile_name)
    manifest = _load_json(run_root / f"{source.stem}.evidence_manifest.json")
    manifest_values = manifest.get("manifest") if isinstance(manifest.get("manifest"), dict) else {}
    manifest_dict = manifest_values if isinstance(manifest_values, dict) else {}
    run_id = source.stem
    model = _text(summary.get("model")) or _nested(profile, "device", "model")
    return RunMetadata(
        run_id=run_id,
        batch_id=_text(summary.get("batch_id")),
        device_model=model,
        form_factor=_nested(profile, "device", "form_factor"),
        android_version=_nested(profile, "android", "release"),
        one_ui_version=_nested(profile, "android", "one_ui_version"),
        talkback=_nested(profile, "talkback", "version_name"),
        talkback_package=_nested(profile, "talkback", "package"),
        app=_nested(profile, "target_app", "version_name"),
        app_package=_nested(profile, "target_app", "package"),
        locale=_nested(profile, "locale") or _nested(profile, "environment_fingerprint", "fingerprint_source", "direct", "locale") or _text(manifest_dict.get("locale")),
        display_width=_int(profile.get("display", {}).get("logical_size", {}).get("value", {}).get("width")) if isinstance(profile.get("display"), dict) and isinstance(profile.get("display", {}).get("logical_size"), dict) and isinstance(profile.get("display", {}).get("logical_size", {}).get("value"), dict) else 0,
        display_height=_int(profile.get("display", {}).get("logical_size", {}).get("value", {}).get("height")) if isinstance(profile.get("display"), dict) and isinstance(profile.get("display", {}).get("logical_size"), dict) and isinstance(profile.get("display", {}).get("logical_size", {}).get("value"), dict) else 0,
    )


def _row_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[str], list[dict[str, object]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_text(value) for value in rows[0]]
    return headers, [dict(zip(headers, row, strict=False)) for row in rows[1:]]


def _raw_index(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[tuple[str, str], dict[str, object]]:
    headers, rows = _row_map(sheet)
    del headers
    index: dict[tuple[str, str], dict[str, object]] = {}
    for row in rows:
        key = (_text(row.get("scenario_id")), _text(row.get("step_index")))
        if key[0] and key[1]:
            index[key] = row
    return index


def _value(result: dict[str, object], raw: dict[str, object], *names: str) -> str:
    for name in names:
        value = _text(result.get(name)) or _text(raw.get(name))
        if value:
            return value
    return ""


def _focus_node_class(raw: dict[str, object]) -> str:
    value = raw.get("focus_node")
    if not isinstance(value, str):
        return ""
    try:
        node = json.loads(value)
    except json.JSONDecodeError:
        return ""
    return _text(node.get("className")) if isinstance(node, dict) else ""


def _qa_display_text(*values: str, placeholder: str) -> str:
    for value in values:
        if value:
            return value
    return placeholder


def _issue_type(result: dict[str, object]) -> str:
    mismatch = _text(result.get("mismatch_type")).upper()
    failure = _text(result.get("failure_reason")).lower()
    if mismatch in _ISSUE_TYPES:
        return _ISSUE_TYPES[mismatch]
    if "speech" in failure and "visible" in failure:
        return "화면 텍스트와 발화 불일치"
    if failure in {"move_failed", "repeat_no_progress", "terminal_not_handled", "recovery_failed"}:
        return "자동화 이동 실패와 접근성 FAIL 구분"
    return "접근성 FAIL"


def _relative_artifact(run_root: Path, value: str, fallback: str) -> str:
    if not value:
        return fallback
    candidate = Path(value)
    if not candidate.is_absolute():
        candidate = run_root / candidate
    if not candidate.is_file():
        return fallback
    try:
        return candidate.relative_to(run_root).as_posix()
    except ValueError:
        return fallback


def _artifact_path(run_root: Path, value: str) -> Path | None:
    if not value:
        return None
    candidate = Path(value)
    if not candidate.is_absolute():
        candidate = run_root / candidate
    return candidate if candidate.is_file() else None


def read_review_rows(source: Path) -> tuple[RunMetadata, list[SourceRow], list[dict[str, str]]]:
    run_root = source.parent
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    result_headers, result_rows = _row_map(workbook["result"])
    del result_headers
    raw = _raw_index(workbook["raw"]) if "raw" in workbook.sheetnames else {}
    evidence_events = speech_evidence_index(run_root / f"{source.stem}.evidence.jsonl")
    metadata = _metadata(run_root, source)
    failures: list[SourceRow] = []
    warnings: dict[tuple[str, str], dict[str, str]] = {}
    for result_row, result in enumerate(result_rows, start=2):
        status = _text(result.get("final_result")).upper()
        scenario = _text(result.get("scenario_id"))
        step = _text(result.get("step"))
        raw_row = raw.get((scenario, step), {})
        if status in _FAIL_RESULTS:
            crop_value = _value(result, raw_row, "crop_image_path", "crop_image")
            crop_path = _artifact_path(run_root, crop_value)
            full_value = _value(result, raw_row, "full_screenshot_path")
            full_path = _artifact_path(run_root, full_value)
            evidence = _relative_artifact(run_root, f"{source.stem}.evidence.jsonl", "Not available")
            visible_text = _value(result, raw_row, "visible_label", "actual_focus_visible", "focus_text")
            speech = _value(result, raw_row, "merged_announcement", "actual_focus_speech", "speech_main")
            expected_speech = _value(result, raw_row, "representative_speech")
            resource_id = _value(result, raw_row, "focus_view_id", "actual_focus_resource_id", "representative_resource_id")
            parent_text = _value(result, raw_row, "semantic_card_title", "parent_visible", "parent_text")
            representative_text = _value(result, raw_row, "representative_visible", "representative_speech")
            bounds = _value(result, raw_row, "focus_bounds", "actual_focus_bounds")
            content_description = _value(result, raw_row, "focus_content_description")
            class_name = _value(result, raw_row, "class_name", "focus_class_name") or _focus_node_class(raw_row)
            speech_status = classify_speech_status(
                speech=speech,
                visible_text=visible_text,
                content_description=content_description,
                class_name=class_name,
                evidence=evidence_for_row(evidence_events.get((scenario, step), []), resource_id),
            )
            context = focus_context(raw=raw_row, visible_text=visible_text, speech=speech, resource_id=resource_id, parent_text=parent_text, representative_text=representative_text, screen=_value(result, raw_row, "context_type", "tab_name") or scenario, bounds=bounds, crop_path=crop_path, display_width=metadata.display_width, display_height=metadata.display_height)
            if full_path is not None:
                screenshot = _relative_artifact(run_root, full_value, "Full screenshot not captured for this run")
                annotation = create_full_screen_annotation(full_path, bounds=bounds, logical_width=metadata.display_width, logical_height=metadata.display_height, target=context.target, run_root=run_root)
                evidence_type = "Full screen + focus bounds" if annotation else "Full screen"
            elif crop_path is not None:
                screenshot = _relative_artifact(run_root, crop_value, "Full screenshot not captured for this run")
                annotation = create_focus_annotation(crop_path, run_root)
                evidence_type = "Crop only (legacy run)"
            else:
                screenshot = "Full screenshot not captured for this run"
                annotation = ""
                evidence_type = "No screenshot"
            if evidence_type == "Crop only (legacy run)":
                review_description = f"이 Run에는 전체 화면 캡처가 없습니다. 첨부된 crop과 {context.approximate_position} 위치를 참고하여 {context.target} 항목의 실제 TalkBack 발화를 확인하세요."
            elif evidence_type == "No screenshot":
                review_description = f"전체 화면 증거가 없어 자동 위치 안내가 제한됩니다. {_value(result, raw_row, 'context_type', 'tab_name') or scenario} 화면의 {context.approximate_position} 근처에서 {context.target} Focus Target을 수동 탐색하여 실제 TalkBack 발화를 확인하세요."
            else:
                review_description = context.description
            review_description = f"{review_description}\n\n{speech_review_instruction(speech_status, speech)}"
            failure_reason = _value(result, raw_row, "failure_reason")
            classification = classify_review(mismatch_type=_value(result, raw_row, "mismatch_type"), failure_reason=failure_reason)
            failures.append(SourceRow(
                result_row=result_row,
                scenario_id=scenario,
                scenario_name=_value(result, raw_row, "plugin_name") or scenario,
                step=step,
                screen=_value(result, raw_row, "context_type", "tab_name"),
                automatic_result=status,
                issue_type=_issue_type(result),
                mismatch_reason=failure_reason or _value(result, raw_row, "mismatch_type"),
                visible_text=_qa_display_text(visible_text, context.target if context.target_source == "OCR(crop)" else "", placeholder="화면 텍스트 없음\n(실제 화면 표시 여부 확인)"),
                speech=_qa_display_text(expected_speech, speech, context.target if context.target_source == "OCR(crop)" else "", visible_text, placeholder="예상 발화 없음\n(새로운 Focus가 실제로 발화하는지 확인)"),
                speech_status=speech_status.label,
                speech_diagnostic=speech_status.diagnostic,
                expected=representative_text,
                resource_id=resource_id,
                class_name=class_name,
                bounds=bounds,
                screenshot=screenshot,
                screenshot_annotation=annotation,
                screenshot_evidence_type=evidence_type,
                evidence=evidence,
                source_run_id=metadata.run_id,
                focus_target=context.target,
                focus_target_source=context.target_source,
                approximate_position=context.approximate_position,
                focus_center_relative=context.center_relative,
                review_description=review_description,
                review_area=classification.area,
                classification_reason=classification.reason,
                traversal_state=_value(result, raw_row, "move_result", "row_source"),
                recovery_state=_value(result, raw_row, "overlay_recovery_status", "recovery_status"),
                terminal_state=_value(result, raw_row, "stop_reason", "failure_reason", "final_result"),
            ))
        elif status == "WARN":
            failure = _value(result, raw_row, "failure_reason", "stop_reason", "move_result") or "WARN"
            key = (scenario, failure)
            if key not in warnings:
                warnings[key] = {"scenario": scenario, "warning_type": failure, "step": step, "terminal": _value(result, raw_row, "stop_reason", "final_result"), "count": "0"}
            warnings[key]["count"] = str(int(warnings[key]["count"]) + 1)
    workbook.close()
    fail_scenarios = {row.scenario_id for row in failures}
    for warning in warnings.values():
        warning["actual_fail"] = "YES" if warning["scenario"] in fail_scenarios else "NO"
    return metadata, failures, sorted(warnings.values(), key=lambda row: (row["scenario"], row["warning_type"], row["step"]))


def deterministic_pass_sample(source: Path, rows: list[dict[str, object]], rate: float) -> list[dict[str, object]]:
    if rate <= 0:
        return []
    selected: list[dict[str, object]] = []
    for row in rows:
        key = f"{source.stem}:{_text(row.get('scenario_id'))}:{_text(row.get('step'))}"
        bucket = int(hashlib.sha256(key.encode()).hexdigest()[:8], 16) / 0xFFFFFFFF
        if bucket < rate:
            selected.append(row)
    return selected


def read_pass_rows(source: Path) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    headers, rows = _row_map(workbook["result"])
    workbook.close()
    return [row for row in rows if _text(row.get("final_result")).upper() == "PASS"]
