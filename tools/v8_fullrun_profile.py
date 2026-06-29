"""Build scenario-level performance and quality profiles from completed V8 runs."""

from __future__ import annotations

import argparse
import csv
import json
import re
import statistics
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROFILE_JSON = "v8_fullrun_profile.json"
PROFILE_MARKDOWN = "v8_fullrun_profile.md"
PROFILE_CSV = "v8_fullrun_profile.csv"

SCENARIO_FIELDS = [
    "scenario_id",
    "step_count",
    "attempted_step_count",
    "persisted_step_count",
    "suppressed_row_count",
    "suppressed_row_count_estimated",
    "first_step",
    "last_step",
    "duration_sec",
    "result_row_count",
    "fail_row_count",
    "review_row_count",
    "clean_row_count",
    "repeat_no_progress_count",
    "viewport_exhausted_eval_count",
    "local_tab_probe_success_count",
    "probe_candidate_count",
    "probe_attempted_count",
    "probe_success_count",
    "probe_failed_count",
    "probe_skipped_count",
    "screen_skipped_count",
    "scenario_filtered_count",
    "validation_match_count",
    "validation_partial_match_count",
    "promotable_count",
    "promoted_count",
    "dedup_skipped_count",
    "adaptive_extension_count",
    "adaptive_hard_limit_count",
    "adaptive_safety_stop_count",
    "soft_limit_extension_rows",
    "post_soft_limit_new_rows",
    "post_soft_limit_suppressed_rows",
]

COUNT_FIELDS = [field for field in SCENARIO_FIELDS if field.endswith("_count")]

SCENARIO_START_RE = re.compile(
    r"\[SCENARIO\]\[(?:stabilization|start|begin)\].*?\bscenario=(?:'([^']+)'|([^\s]+))"
)
SCENARIO_FIELD_RE = re.compile(r"\bscenario=(?:'([^']+)'|([^\s]+))")
PERF_SUMMARY_RE = re.compile(r"\[PERF\]\[scenario_summary\].*?\bscenario=(?:'([^']+)'|([^\s]+))")
TOTAL_RUNTIME_RE = re.compile(r"\btotal_runtime=(\d+(?:\.\d+)?)")
TOTAL_STEPS_RE = re.compile(r"\btotal_steps=(\d+)")
STEP_RE = re.compile(r"\bstep=(\d+)")
TIMESTAMP_RE = re.compile(r"^\[(\d{2}):(\d{2}):(\d{2})\]")
RUN_STEM_RE = re.compile(r"(talkback_compare_\d{8}_\d{6})")

PROBE_RESULT_FIELD_MAP = {
    "probe_candidate_count": ("candidate_count",),
    "probe_attempted_count": ("attempted_count",),
    "probe_success_count": ("success_count",),
    "probe_failed_count": ("failed_count",),
    "probe_skipped_count": ("skipped_count",),
    "screen_skipped_count": ("screen_skipped_count",),
    "scenario_filtered_count": ("scenario_filtered_count",),
}

VALIDATION_FIELD_MAP = {
    "validation_match_count": ("match_count",),
    "validation_partial_match_count": ("partial_match_count",),
    "promotable_count": ("promotable_count",),
    "promoted_count": ("promoted_count", "promoted_row_count"),
    "dedup_skipped_count": (
        "dedup_skipped_count",
        "promotion_dedup_skipped_count",
    ),
}

REPEAT_WARNING_THRESHOLD = 2
TOP_LIMIT = 5


def _empty_scenario(scenario_id: str) -> dict[str, Any]:
    row: dict[str, Any] = {
        "scenario_id": scenario_id,
        "step_count": 0,
        "attempted_step_count": 0,
        "persisted_step_count": 0,
        "suppressed_row_count": 0,
        "suppressed_row_count_estimated": False,
        "first_step": None,
        "last_step": None,
        "duration_sec": None,
    }
    for field in SCENARIO_FIELDS:
        row.setdefault(field, 0)
    return row


def _scenario(
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    scenario_id: str | None,
) -> dict[str, Any] | None:
    normalized = str(scenario_id or "").strip().strip("'\"")
    if not normalized:
        return None
    if normalized not in scenarios:
        scenarios[normalized] = _empty_scenario(normalized)
        execution_order.append(normalized)
    return scenarios[normalized]


def _match_value(match: re.Match[str] | None) -> str | None:
    if not match:
        return None
    return next((value for value in match.groups() if value), None)


def _int_value(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _first_int(mapping: dict[str, Any], keys: Iterable[str]) -> int | None:
    for key in keys:
        value = _int_value(mapping.get(key))
        if value is not None:
            return value
    summary = mapping.get("summary")
    if isinstance(summary, dict):
        for key in keys:
            value = _int_value(summary.get(key))
            if value is not None:
                return value
    return None


def _read_json(path: Path, warnings: list[str], label: str) -> dict[str, Any] | None:
    if not path.is_file():
        warnings.append(f"{label} file not found: {path}")
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        warnings.append(f"{label} could not be read: {path} ({exc})")
        return None
    if not isinstance(payload, dict):
        warnings.append(f"{label} root is not an object: {path}")
        return None
    return payload


def _aggregate_scenarios(payload: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ("scenarios", "scenario_summaries"):
        value = payload.get(key)
        if isinstance(value, list):
            return [entry for entry in value if isinstance(entry, dict)]
        if isinstance(value, dict):
            entries = []
            for scenario_id, entry in value.items():
                if isinstance(entry, dict):
                    entries.append({"scenario_id": scenario_id, **entry})
            return entries
    return []


def parse_xlsx(
    path: Path,
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    warnings: list[str],
) -> None:
    if not path.is_file():
        warnings.append(f"xlsx file not found: {path}")
        return
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        warnings.append(f"xlsx could not be read: {path} ({exc})")
        return

    try:
        if "result" not in workbook.sheetnames:
            warnings.append(f"xlsx result sheet not found: {path}")
            return
        sheet = workbook["result"]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, ())
        index = {
            str(value).strip(): position
            for position, value in enumerate(headers)
            if value is not None and str(value).strip()
        }
        if "scenario_id" not in index:
            warnings.append(f"xlsx result sheet has no scenario_id column: {path}")
            return

        for values in rows:
            scenario_id = _cell(values, index, "scenario_id")
            metric = _scenario(scenarios, execution_order, scenario_id)
            if metric is None:
                continue
            metric["result_row_count"] += 1

            final_result = _cell(values, index, "final_result").upper()
            if final_result == "FAIL":
                metric["fail_row_count"] += 1
            elif final_result in {"WARN", "REVIEW"}:
                metric["review_row_count"] += 1
            elif final_result == "PASS":
                metric["clean_row_count"] += 1

            row_source = _cell(values, index, "row_source").upper()
            if row_source == "COVERAGE_PROBE_PROMOTED":
                metric["promoted_count"] += 1

            dedup_status = _cell(values, index, "promotion_dedup_status").upper()
            if dedup_status == "SKIPPED":
                metric["dedup_skipped_count"] += 1
    finally:
        workbook.close()


def _cell(values: tuple[Any, ...], index: dict[str, int], key: str) -> str:
    position = index.get(key)
    if position is None or position >= len(values):
        return ""
    value = values[position]
    return str(value or "").strip()


def parse_probe_results(
    path: Path,
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    warnings: list[str],
) -> None:
    payload = _read_json(path, warnings, "probe results aggregate")
    if payload is None:
        return
    entries = _aggregate_scenarios(payload)
    if not entries:
        warnings.append(f"probe results aggregate has no scenario summaries: {path}")
        return
    for entry in entries:
        metric = _scenario(scenarios, execution_order, str(entry.get("scenario_id") or ""))
        if metric is None:
            continue
        for target, source_keys in PROBE_RESULT_FIELD_MAP.items():
            value = _first_int(entry, source_keys)
            if value is not None:
                metric[target] = value


def parse_probe_validation(
    path: Path,
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    warnings: list[str],
) -> None:
    payload = _read_json(path, warnings, "probe validation aggregate")
    if payload is None:
        return
    entries = _aggregate_scenarios(payload)
    if not entries:
        warnings.append(f"probe validation aggregate has no scenario summaries: {path}")
        return
    for entry in entries:
        metric = _scenario(scenarios, execution_order, str(entry.get("scenario_id") or ""))
        if metric is None:
            continue
        for target, source_keys in VALIDATION_FIELD_MAP.items():
            value = _first_int(entry, source_keys)
            if value is not None:
                metric[target] = value


def parse_log(
    path: Path,
    scenarios: dict[str, dict[str, Any]],
    execution_order: list[str],
    warnings: list[str],
) -> None:
    if not path.is_file():
        warnings.append(f"normal log file not found: {path}")
        return
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError as exc:
        warnings.append(f"normal log could not be read: {path} ({exc})")
        return

    current_scenario: str | None = None
    day_offset = 0
    previous_clock: int | None = None
    scenario_start_seconds: dict[str, int] = {}
    scenario_end_seconds: dict[str, int] = {}
    observed_step_events: dict[str, int] = {}
    attempted_steps: dict[str, set[int]] = {}
    explicit_suppressed_steps: dict[str, set[int]] = {}
    active_step_by_scenario: dict[str, int] = {}
    adaptive_soft_limit: dict[str, int] = {}
    post_soft_limit_attempted_steps: dict[str, set[int]] = {}
    post_soft_limit_suppressed_steps: dict[str, set[int]] = {}
    repeat_keys: set[tuple[str, str]] = set()

    for line in lines:
        clock = _clock_seconds(line)
        if clock is not None:
            if previous_clock is not None and clock < previous_clock:
                day_offset += 24 * 60 * 60
            previous_clock = clock
            clock += day_offset

        start_match = SCENARIO_START_RE.search(line)
        if start_match:
            current_scenario = _match_value(start_match)
            metric = _scenario(scenarios, execution_order, current_scenario)
            if metric is not None and clock is not None:
                scenario_start_seconds.setdefault(metric["scenario_id"], clock)

        perf_match = PERF_SUMMARY_RE.search(line)
        explicit_scenario = _match_value(perf_match or SCENARIO_FIELD_RE.search(line))
        scenario_id = explicit_scenario or current_scenario
        metric = _scenario(scenarios, execution_order, scenario_id)
        if metric is None:
            continue

        if "[STEP] START" in line or "[STEP] END" in line:
            step_match = STEP_RE.search(line)
            if step_match:
                step = int(step_match.group(1))
                metric["first_step"] = step if metric["first_step"] is None else min(metric["first_step"], step)
                metric["last_step"] = step if metric["last_step"] is None else max(metric["last_step"], step)
                if "[STEP] START" in line:
                    attempted_steps.setdefault(metric["scenario_id"], set()).add(step)
                    active_step_by_scenario[metric["scenario_id"]] = step
                if "[STEP] END" in line:
                    observed_step_events[metric["scenario_id"]] = observed_step_events.get(metric["scenario_id"], 0) + 1
                    active_step_by_scenario[metric["scenario_id"]] = step

        if "[STEP][row_filter]" in line:
            active_step = active_step_by_scenario.get(metric["scenario_id"])
            if active_step is not None:
                explicit_suppressed_steps.setdefault(metric["scenario_id"], set()).add(active_step)

        if "[ADAPTIVE_MAX_STEPS]" in line:
            scenario_key = metric["scenario_id"]
            current_limit_match = re.search(r"\bcurrent_limit=(\d+)", line)
            decision_match = re.search(r"\bdecision='([^']+)'", line)
            if current_limit_match:
                current_limit = int(current_limit_match.group(1))
                adaptive_soft_limit.setdefault(scenario_key, current_limit)
                for attempted_step in attempted_steps.get(scenario_key, set()):
                    if attempted_step > adaptive_soft_limit[scenario_key]:
                        post_soft_limit_attempted_steps.setdefault(scenario_key, set()).add(attempted_step)
                for suppressed_step in explicit_suppressed_steps.get(scenario_key, set()):
                    if suppressed_step > adaptive_soft_limit[scenario_key]:
                        post_soft_limit_suppressed_steps.setdefault(scenario_key, set()).add(suppressed_step)
            if decision_match:
                decision = decision_match.group(1)
                if decision == "extend":
                    metric["adaptive_extension_count"] += 1
                elif decision == "hard_stop":
                    metric["adaptive_hard_limit_count"] += 1
                elif decision == "stop":
                    metric["adaptive_safety_stop_count"] += 1

        if "[STEP][viewport_exhausted_eval]" in line:
            metric["viewport_exhausted_eval_count"] += 1

        if "[STEP][local_tab_content_entry_probe_success]" in line:
            metric["local_tab_probe_success_count"] += 1

        if "[STOP][eval]" in line and "repeat_no_progress" in line:
            step_match = STEP_RE.search(line)
            repeat_key = (metric["scenario_id"], step_match.group(1) if step_match else line)
            if repeat_key not in repeat_keys:
                repeat_keys.add(repeat_key)
                metric["repeat_no_progress_count"] += 1

        if perf_match:
            runtime_match = TOTAL_RUNTIME_RE.search(line)
            steps_match = TOTAL_STEPS_RE.search(line)
            if runtime_match:
                metric["duration_sec"] = float(runtime_match.group(1))
            if steps_match:
                metric["step_count"] = int(steps_match.group(1))
                metric["persisted_step_count"] = int(steps_match.group(1))
            if clock is not None:
                scenario_end_seconds[metric["scenario_id"]] = clock

    for scenario_id, metric in scenarios.items():
        attempted_count = len(attempted_steps.get(scenario_id, set()))
        metric["attempted_step_count"] = attempted_count
        if metric["step_count"] == 0 and observed_step_events.get(scenario_id):
            metric["step_count"] = observed_step_events[scenario_id]
            metric["persisted_step_count"] = observed_step_events[scenario_id]
        explicit_suppressed_count = len(explicit_suppressed_steps.get(scenario_id, set()))
        if explicit_suppressed_count > 0:
            metric["suppressed_row_count"] = explicit_suppressed_count
        else:
            metric["suppressed_row_count"] = max(
                0,
                int(metric["attempted_step_count"]) - int(metric["persisted_step_count"]),
            )
            metric["suppressed_row_count_estimated"] = metric["attempted_step_count"] > 0
        soft_limit = adaptive_soft_limit.get(scenario_id)
        if soft_limit is not None:
            attempted_after_soft_limit = {
                step for step in attempted_steps.get(scenario_id, set()) if step > soft_limit
            }
            suppressed_after_soft_limit = {
                step for step in explicit_suppressed_steps.get(scenario_id, set()) if step > soft_limit
            }
            metric["post_soft_limit_suppressed_rows"] = len(suppressed_after_soft_limit)
            metric["post_soft_limit_new_rows"] = max(
                0,
                len(attempted_after_soft_limit) - len(suppressed_after_soft_limit),
            )
            metric["soft_limit_extension_rows"] = metric["post_soft_limit_new_rows"]
        if metric["duration_sec"] is None:
            start = scenario_start_seconds.get(scenario_id)
            end = scenario_end_seconds.get(scenario_id)
            if start is not None and end is not None and end >= start:
                metric["duration_sec"] = float(end - start)
            elif metric["step_count"] or metric["result_row_count"]:
                warnings.append(f"duration unavailable for scenario: {scenario_id}")


def _clock_seconds(line: str) -> int | None:
    match = TIMESTAMP_RE.match(line)
    if not match:
        return None
    hour, minute, second = (int(value) for value in match.groups())
    return hour * 3600 + minute * 60 + second


def build_profile(
    *,
    xlsx: Path,
    log: Path,
    probe_results: Path,
    probe_validation: Path,
) -> dict[str, Any]:
    warnings: list[str] = []
    scenarios: dict[str, dict[str, Any]] = {}
    execution_order: list[str] = []

    parse_xlsx(xlsx, scenarios, execution_order, warnings)
    parse_log(log, scenarios, execution_order, warnings)
    parse_probe_results(probe_results, scenarios, execution_order, warnings)
    parse_probe_validation(probe_validation, scenarios, execution_order, warnings)

    scenario_rows = [scenarios[scenario_id] for scenario_id in execution_order]
    findings = build_findings(scenario_rows)
    summary = build_summary(scenario_rows)
    stem = _artifact_stem(xlsx, log, probe_results, probe_validation)
    return {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "artifact_stem": stem,
        "inputs": {
            "xlsx": str(xlsx),
            "log": str(log),
            "probe_results": str(probe_results),
            "probe_validation": str(probe_validation),
        },
        "warnings": warnings,
        "summary": summary,
        "scenarios": scenario_rows,
        "findings": findings,
    }


def build_summary(scenarios: list[dict[str, Any]]) -> dict[str, Any]:
    durations = [float(row["duration_sec"]) for row in scenarios if row["duration_sec"] is not None]
    return {
        "scenario_count": len(scenarios),
        "total_steps": sum(int(row["step_count"]) for row in scenarios),
        "total_attempted_steps": sum(int(row["attempted_step_count"]) for row in scenarios),
        "total_persisted_steps": sum(int(row["persisted_step_count"]) for row in scenarios),
        "total_suppressed_rows": sum(int(row["suppressed_row_count"]) for row in scenarios),
        "total_duration_sec": round(sum(durations), 3) if durations else None,
        "total_repeat_no_progress": sum(int(row["repeat_no_progress_count"]) for row in scenarios),
        "total_probe_candidates": sum(int(row["probe_candidate_count"]) for row in scenarios),
        "total_probe_attempted": sum(int(row["probe_attempted_count"]) for row in scenarios),
        "total_probe_success": sum(int(row["probe_success_count"]) for row in scenarios),
        "total_promoted": sum(int(row["promoted_count"]) for row in scenarios),
        "total_dedup_skipped": sum(int(row["dedup_skipped_count"]) for row in scenarios),
        "total_adaptive_extensions": sum(int(row["adaptive_extension_count"]) for row in scenarios),
        "total_adaptive_hard_limits": sum(int(row["adaptive_hard_limit_count"]) for row in scenarios),
        "total_adaptive_safety_stops": sum(int(row["adaptive_safety_stop_count"]) for row in scenarios),
        "total_post_soft_limit_new_rows": sum(int(row["post_soft_limit_new_rows"]) for row in scenarios),
        "total_post_soft_limit_suppressed_rows": sum(int(row["post_soft_limit_suppressed_rows"]) for row in scenarios),
        "slowest_scenarios": _rank(scenarios, "duration_sec", include_zero=False),
        "highest_repeat_scenarios": _rank(scenarios, "repeat_no_progress_count", include_zero=False),
        "highest_probe_failure_scenarios": _rank(scenarios, "probe_failed_count", include_zero=False),
    }


def _rank(
    scenarios: list[dict[str, Any]],
    field: str,
    *,
    include_zero: bool,
) -> list[dict[str, Any]]:
    ranked = [
        {"scenario_id": row["scenario_id"], "value": row[field]}
        for row in scenarios
        if row.get(field) is not None and (include_zero or float(row[field]) > 0)
    ]
    ranked.sort(key=lambda item: (-float(item["value"]), item["scenario_id"]))
    return ranked[:TOP_LIMIT]


def build_findings(scenarios: list[dict[str, Any]]) -> list[dict[str, str]]:
    findings: list[dict[str, str]] = []
    durations = [float(row["duration_sec"]) for row in scenarios if row["duration_sec"] is not None]
    duration_threshold = max(300.0, statistics.median(durations) * 2.0) if len(durations) >= 2 else None

    for row in scenarios:
        scenario_id = str(row["scenario_id"])
        repeat_count = int(row["repeat_no_progress_count"])
        if repeat_count > REPEAT_WARNING_THRESHOLD:
            findings.append(
                _finding(
                    scenario_id,
                    "high_repeat_no_progress",
                    f"repeat_no_progress occurred {repeat_count} times (threshold {REPEAT_WARNING_THRESHOLD}).",
                )
            )
        if int(row["probe_failed_count"]) > int(row["probe_success_count"]):
            findings.append(
                _finding(
                    scenario_id,
                    "probe_failures_exceed_successes",
                    "Probe failures exceed probe successes.",
                )
            )
        duration = row["duration_sec"]
        if duration_threshold is not None and duration is not None and float(duration) > duration_threshold:
            findings.append(
                _finding(
                    scenario_id,
                    "unusually_high_duration",
                    f"Duration {float(duration):.1f}s exceeds the heuristic threshold {duration_threshold:.1f}s.",
                )
            )
        if int(row["probe_success_count"]) > 0 and int(row["promoted_count"]) == 0:
            findings.append(
                _finding(
                    scenario_id,
                    "successful_probe_not_promoted",
                    "Probe succeeded, but no production row was promoted.",
                )
            )
    return findings


def _finding(scenario_id: str, code: str, message: str) -> dict[str, str]:
    return {
        "severity": "WARN",
        "scenario_id": scenario_id,
        "code": code,
        "message": message,
    }


def write_outputs(profile: dict[str, Any], output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / PROFILE_JSON
    markdown_path = output_dir / PROFILE_MARKDOWN
    csv_path = output_dir / PROFILE_CSV

    json_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(render_markdown(profile), encoding="utf-8")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=SCENARIO_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(profile["scenarios"])
    return {"json": json_path, "markdown": markdown_path, "csv": csv_path}


def render_markdown(profile: dict[str, Any]) -> str:
    summary = profile["summary"]
    scenarios = profile["scenarios"]
    lines = [
        "# V8 Full Run Profile",
        "",
        f"Artifact: `{profile.get('artifact_stem', '')}`",
        "",
        "## Run Summary",
        "",
        "| Metric | Value |",
        "| --- | ---: |",
    ]
    for key in (
        "scenario_count",
        "total_steps",
        "total_attempted_steps",
        "total_persisted_steps",
        "total_suppressed_rows",
        "total_duration_sec",
        "total_repeat_no_progress",
        "total_probe_candidates",
        "total_probe_attempted",
        "total_probe_success",
        "total_promoted",
        "total_dedup_skipped",
        "total_adaptive_extensions",
        "total_adaptive_hard_limits",
        "total_adaptive_safety_stops",
        "total_post_soft_limit_new_rows",
        "total_post_soft_limit_suppressed_rows",
    ):
        lines.append(f"| {key} | {_display(summary.get(key))} |")

    lines.extend(
        [
            "",
            "## Step Accounting",
            "",
            "Profiler `step_count` remains backward-compatible and reflects persisted/perf steps.",
            "Use `attempted_step_count` when evaluating adaptive max steps or loop churn.",
            "",
            "| Scenario | Attempted steps | Persisted steps | Suppressed rows | Estimated suppression | Result rows |",
            "| --- | ---: | ---: | ---: | --- | ---: |",
        ]
    )
    for row in scenarios:
        lines.append(
            f"| {row['scenario_id']} | {row['attempted_step_count']} | {row['persisted_step_count']} "
            f"| {row['suppressed_row_count']} | {row['suppressed_row_count_estimated']} | {row['result_row_count']} |"
        )

    lines.extend(
        [
            "",
            "## Scenario Table",
            "",
            "| Scenario | Steps | Duration (s) | Result rows | Fail | Review | Repeat | Probe A/S/F | Promotable | Promoted | Dedup skipped |",
            "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in scenarios:
        probe = f"{row['probe_attempted_count']}/{row['probe_success_count']}/{row['probe_failed_count']}"
        lines.append(
            f"| {row['scenario_id']} | {row['step_count']} | {_display(row['duration_sec'])} "
            f"| {row['result_row_count']} | {row['fail_row_count']} | {row['review_row_count']} "
            f"| {row['repeat_no_progress_count']} | {probe} | {row['promotable_count']} "
            f"| {row['promoted_count']} | {row['dedup_skipped_count']} |"
        )

    lines.extend(
        [
            "",
            "## Adaptive Max Steps",
            "",
            "| Scenario | Extensions | Hard limit | Safety stop | Post-soft new rows | Post-soft suppressed |",
            "| --- | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    extended_scenarios = []
    for row in scenarios:
        if int(row["adaptive_extension_count"]) > 0:
            extended_scenarios.append(row["scenario_id"])
        lines.append(
            f"| {row['scenario_id']} | {row['adaptive_extension_count']} | {row['adaptive_hard_limit_count']} "
            f"| {row['adaptive_safety_stop_count']} | {row['post_soft_limit_new_rows']} "
            f"| {row['post_soft_limit_suppressed_rows']} |"
        )
    lines.extend(
        [
            "",
            f"Scenarios extended: {', '.join(f'`{scenario}`' for scenario in extended_scenarios) if extended_scenarios else 'None'}",
            "",
        ]
    )

    lines.extend(["", "## Top Slow / Heavy Scenarios", ""])
    _append_ranking(lines, "Top by duration", summary["slowest_scenarios"], "s")
    _append_ranking(lines, "Top by step_count", _rank(scenarios, "step_count", include_zero=False), "")
    _append_ranking(lines, "Top by attempted_step_count", _rank(scenarios, "attempted_step_count", include_zero=False), "")
    _append_ranking(lines, "Top by repeat_no_progress_count", summary["highest_repeat_scenarios"], "")
    _append_ranking(lines, "Top by probe_failed_count", summary["highest_probe_failure_scenarios"], "")

    lines.extend(["", "## Findings", ""])
    if profile["findings"]:
        for finding in profile["findings"]:
            lines.append(
                f"- {finding['severity']} `{finding['scenario_id']}` "
                f"({finding['code']}): {finding['message']}"
            )
    else:
        lines.append("- No rule-based warnings were detected.")

    lines.extend(["", "## Warnings", ""])
    if profile["warnings"]:
        lines.extend(f"- {warning}" for warning in profile["warnings"])
    else:
        lines.append("- None.")
    lines.append("")
    return "\n".join(lines)


def _append_ranking(
    lines: list[str],
    title: str,
    ranking: list[dict[str, Any]],
    suffix: str,
) -> None:
    lines.append(f"### {title}")
    lines.append("")
    if not ranking:
        lines.append("- No data.")
    else:
        for item in ranking:
            lines.append(f"- `{item['scenario_id']}`: {_display(item['value'])}{suffix}")
    lines.append("")


def _display(value: Any) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, float):
        return f"{value:.3f}".rstrip("0").rstrip(".")
    return str(value)


def resolve_artifacts(artifact_dir: Path) -> dict[str, Path]:
    if not artifact_dir.is_dir():
        raise FileNotFoundError(f"artifact directory not found: {artifact_dir}")
    anchors = list(artifact_dir.rglob("talkback_compare_*.xlsx"))
    if not anchors:
        anchors = list(artifact_dir.rglob("talkback_compare_*.normal.log"))
    if not anchors:
        raise FileNotFoundError(f"no talkback_compare artifacts found under: {artifact_dir}")
    anchor = max(anchors, key=_artifact_sort_key)
    stem = _artifact_stem(anchor)
    parent = anchor.parent
    return {
        "xlsx": parent / f"{stem}.xlsx",
        "log": parent / f"{stem}.normal.log",
        "probe_results": parent / f"{stem}.coverage_probe_results.aggregate.json",
        "probe_validation": parent / f"{stem}.coverage_probe_validation.aggregate.json",
    }


def _artifact_sort_key(path: Path) -> tuple[str, float]:
    match = RUN_STEM_RE.search(path.name)
    timestamp = match.group(1).removeprefix("talkback_compare_") if match else ""
    try:
        modified = path.stat().st_mtime
    except OSError:
        modified = 0.0
    return timestamp, modified


def _artifact_stem(*paths: Path) -> str:
    for path in paths:
        match = RUN_STEM_RE.search(path.name)
        if match:
            return match.group(1)
    return "talkback_compare_unknown"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path)
    parser.add_argument("--log", type=Path)
    parser.add_argument("--probe-results", type=Path)
    parser.add_argument("--probe-validation", type=Path)
    parser.add_argument("--artifact-dir", type=Path)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args(argv)
    if not args.artifact_dir and not any(
        (args.xlsx, args.log, args.probe_results, args.probe_validation)
    ):
        parser.error("provide --artifact-dir or at least one explicit artifact path")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    detected: dict[str, Path] = {}
    if args.artifact_dir:
        try:
            detected = resolve_artifacts(args.artifact_dir)
        except FileNotFoundError as exc:
            print(f"error: {exc}", file=sys.stderr)
            return 2

    explicit_paths = [
        path
        for path in (args.xlsx, args.log, args.probe_results, args.probe_validation)
        if path is not None
    ]
    base_stem = _artifact_stem(*explicit_paths, *detected.values())
    base_dir = (
        explicit_paths[0].parent
        if explicit_paths
        else next(iter(detected.values())).parent
    )
    paths = {
        "xlsx": args.xlsx or detected.get("xlsx") or base_dir / f"{base_stem}.xlsx",
        "log": args.log or detected.get("log") or base_dir / f"{base_stem}.normal.log",
        "probe_results": args.probe_results
        or detected.get("probe_results")
        or base_dir / f"{base_stem}.coverage_probe_results.aggregate.json",
        "probe_validation": args.probe_validation
        or detected.get("probe_validation")
        or base_dir / f"{base_stem}.coverage_probe_validation.aggregate.json",
    }
    output_dir = args.output_dir or base_dir / f"v8_profile_{base_stem.removeprefix('talkback_compare_')}"
    profile = build_profile(**paths)
    outputs = write_outputs(profile, output_dir)

    summary = profile["summary"]
    print(f"profiled {summary['scenario_count']} scenarios")
    print(f"total_steps={summary['total_steps']} total_duration_sec={summary['total_duration_sec']}")
    for label, path in outputs.items():
        print(f"{label}={path}")
    if profile["warnings"]:
        print(f"warnings={len(profile['warnings'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
