import json
import os
import re
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from talkback_lib import A11yAdbClient


DEV_SERIAL = "R3CX40QFDBP"
SCRIPT_VERSION = "1.7.1"
LOG_LEVEL = os.getenv("TB_LOG_LEVEL", "NORMAL").upper()
LOG_LEVEL_ORDER = {"QUIET": 0, "NORMAL": 1, "DEBUG": 2}

OVERLAY_ENTRY_CANDIDATES = [
    {
        "resource_id": "com.samsung.android.oneconnect:id/add_menu_button",
        "label": "Add",
    },
    {
        "resource_id": "com.samsung.android.oneconnect:id/more_menu_button",
        "label": "More options",
    },
]
# Backward compatibility: legacy name kept as alias.
OVERLAY_ENTRY_ALLOWLIST = OVERLAY_ENTRY_CANDIDATES

OVERLAY_MAX_STEPS = 10
MAIN_STEP_WAIT_SECONDS = 1.2
MAIN_ANNOUNCEMENT_WAIT_SECONDS = 1.2
OVERLAY_STEP_WAIT_SECONDS = 0.8
OVERLAY_ANNOUNCEMENT_WAIT_SECONDS = 0.8
BACK_RECOVERY_WAIT_SECONDS = 0.8
CHECKPOINT_SAVE_EVERY_STEPS = 3
OVERLAY_REALIGN_MAX_STEPS = 8

TAB_CONFIGS = [
    {
        "scenario_id": "home_main",
        "tab_name": "(?i).*home.*",
        "tab_type": "b",
        "tab": {
            "resource_id_regex": "com\\.samsung\\.android\\.oneconnect:id/menu_favorites",
            "text_regex": "(?i).*home.*",
            "announcement_regex": "(?i).*(selected|선택됨)?.*home.*",
            "tie_breaker": "bottom_nav_left_to_right",
            "allow_resource_id_only": True,
        },
        "anchor_name": "(?i).*location.*qr.*code.*",
        "anchor_type": "b",
        "anchor": {
            "text_regex": "(?i).*location.*qr.*code.*",
            "announcement_regex": "(?i).*qr.*code.*",
            "tie_breaker": "top_left",
        },
        "context_verify": {
            "type": "selected_bottom_tab",
            "announcement_regex": "(?i).*(selected|선택됨).*home.*",
        },
        "max_steps": 5,
        "overlay_policy": {
            "allow_candidates": [
                {
                    "resource_id": "com.samsung.android.oneconnect:id/add_menu_button",
                    "label": "Add",
                },
                {
                    "resource_id": "com.samsung.android.oneconnect:id/more_menu_button",
                    "label": "More options",
                },
            ],
            "block_candidates": [],
        },
    },
    {
        "scenario_id": "devices_main",
        "tab_name": "(?i).*devices.*",
        "tab_type": "b",
        "tab": {
            "resource_id_regex": "com\\.samsung\\.android\\.oneconnect:id/menu_devices",
            "text_regex": "(?i).*devices.*",
            "announcement_regex": "(?i).*(selected|선택됨)?.*devices.*",
            "tie_breaker": "bottom_nav_left_to_right",
            "allow_resource_id_only": True,
        },
        "anchor_name": "(?i).*location.*qr.*code.*",
        "anchor_type": "b",
        "anchor": {
            "text_regex": "(?i).*location.*qr.*code.*",
            "announcement_regex": "(?i).*qr.*code.*",
            "tie_breaker": "top_left",
        },
        "context_verify": {
            "type": "selected_bottom_tab",
            "announcement_regex": "(?i).*(selected|선택됨).*devices.*",
        },
        "enabled": True,
        "max_steps": 5,
        "overlay_policy": {
            "allow_candidates": [
                {
                    "resource_id": "com.samsung.android.oneconnect:id/more_menu_button",
                    "label": "More options",
                }
            ],
            "block_candidates": [
                {
                    "resource_id": "com.samsung.android.oneconnect:id/add_menu_button",
                    "label": "Add",
                }
            ],
        },
    },
    {
        "scenario_id": "life_main",
        "tab_name": "(?i).*life.*",
        "tab_type": "b",
        "tab": {
            "resource_id_regex": "com\\.samsung\\.android\\.oneconnect:id/menu_services",
            "text_regex": "(?i).*life.*",
            "announcement_regex": "(?i).*(selected|선택됨)?.*life.*",
            "tie_breaker": "bottom_nav_left_to_right",
            "allow_resource_id_only": True,
        },
        "anchor_name": "(?i).*location.*qr.*code.*",
        "anchor_type": "b",
        "anchor": {
            "text_regex": "(?i).*location.*qr.*code.*",
            "announcement_regex": "(?i).*qr.*code.*",
            "tie_breaker": "top_left",
        },
        "context_verify": {
            "type": "selected_bottom_tab",
            "announcement_regex": "(?i).*(selected|선택됨).*life.*",
        },
        "enabled": True,
        "max_steps": 5,
        "overlay_policy": {
            "allow_candidates": [
                {
                    "resource_id": "com.samsung.android.oneconnect:id/more_menu_button",
                    "label": "More options",
                }
            ],
            "block_candidates": [
                {
                    "resource_id": "com.samsung.android.oneconnect:id/add_menu_button",
                    "label": "Add",
                }
            ],
        },
    },
    {
        "scenario_id": "routines_main",
        "tab_name": "(?i).*routines.*",
        "tab_type": "b",
        "tab": {
            "resource_id_regex": "com\\.samsung\\.android\\.oneconnect:id/menu_automations",
            "text_regex": "(?i).*routines.*",
            "announcement_regex": "(?i).*(selected|선택됨)?.*routines.*",
            "tie_breaker": "bottom_nav_left_to_right",
            "allow_resource_id_only": True,
        },
        "anchor_name": "(?i).*location.*qr.*code.*",
        "anchor_type": "b",
        "anchor": {
            "text_regex": "(?i).*location.*qr.*code.*",
            "announcement_regex": "(?i).*qr.*code.*",
            "tie_breaker": "top_left",
        },
        "context_verify": {
            "type": "selected_bottom_tab",
            "announcement_regex": "(?i).*(selected|선택됨).*routines.*",
        },
        "enabled": True,
        "max_steps": 5,
        "overlay_policy": {
            "allow_candidates": [
                {
                    "resource_id": "com.samsung.android.oneconnect:id/more_menu_button",
                    "label": "More options",
                }
            ],
            "block_candidates": [
                {
                    "resource_id": "com.samsung.android.oneconnect:id/add_menu_button",
                    "label": "Add",
                }
            ],
        },
    },
    {
        "scenario_id": "menu_main",
        "tab_name": "(?i).*menu.*",
        "tab_type": "b",
        "tab": {
            "resource_id_regex": "com\\.samsung\\.android\\.oneconnect:id/menu_more",
            "text_regex": "(?i).*menu.*",
            "announcement_regex": "(?i).*(selected|선택됨)?.*menu.*",
            "tie_breaker": "bottom_nav_left_to_right",
            "allow_resource_id_only": True,
        },
        "anchor_name": "(?i).*smartthings.*",
        "anchor_type": "b",
        "anchor": {
            "text_regex": "(?i).*smartthings.*",
            "announcement_regex": "(?i).*smartthings.*",
            "tie_breaker": "top_left",
        },
        "context_verify": {
            "type": "selected_bottom_tab",
            "announcement_regex": "(?i).*(selected|선택됨).*menu.*",
        },
        "enabled": True,
        "max_steps": 5,
    },
    {
        "scenario_id": "settings_entry_example",
        "tab_name": "(?i).*menu.*",
        "tab_type": "b",
        "anchor_name": "(?i).*settings.*",
        "anchor_type": "a",
        "anchor": {
            "resource_id_regex": "com.samsung.android.oneconnect:id/add_menu_button",
            "text_regex": "(?i).*qr.*",
            "tie_breaker": "top_left",
            "allow_resource_id_only": True,
        },
        "context_verify": {
            "type": "screen",
            "text_regex": "(?i).*settings.*",
            "announcement_regex": "(?i).*settings.*",
        },
        "enabled": False,
        "max_steps": 20,
    },
    {
        "scenario_id": "life_plugin_example",
        "tab_name": "(?i).*life.*",
        "tab_type": "b",
        "anchor_name": "(?i).*location.*qr.*code.*",
        "anchor_type": "b",
        "anchor": {
            "text_regex": "(?i).*location.*qr.*code.*",
            "announcement_regex": "(?i).*qr.*code.*",
            "tie_breaker": "top_left",
        },
        "context_verify": {
            "type": "plugin",
            "text_regex": "(?i).*smartthings.*energy.*",
        },
        "enabled": False,
        "max_steps": 5,
    },
    {
        "scenario_id": "resource_id_only_example",
        "tab_name": "(?i).*home.*",
        "tab_type": "b",
        "anchor_name": "com.samsung.android.oneconnect:id/add_menu_button",
        "anchor_type": "r",
        "anchor": {
            "resource_id_regex": "com\\.samsung\\.android\\.oneconnect:id/add_menu_button",
            "tie_breaker": "top_left",
            "allow_resource_id_only": True,
        },
        "enabled": False,
        "max_steps": 10,
    },
]

ENABLE_IMAGE_CROP = True
ENABLE_IMAGE_INSERT_TO_EXCEL = True
IMAGE_DIR = "output/crops"


def now_str() -> str:
    return time.strftime("%H:%M:%S")


def _should_log(level: str = "NORMAL") -> bool:
    current = LOG_LEVEL if LOG_LEVEL in LOG_LEVEL_ORDER else "NORMAL"
    return LOG_LEVEL_ORDER.get(current, 1) >= LOG_LEVEL_ORDER.get(level, 1)


def log(msg: str, level: str = "NORMAL") -> None:
    if _should_log(level):
        print(f"[{now_str()}] {msg}")
 

def generate_output_path() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"output/talkback_compare_{timestamp}.xlsx"


def to_json_text(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False)
    except Exception:
        return str(value)


def parse_bounds_str(bounds_str: str) -> tuple[int, int, int, int] | None:
    if not bounds_str:
        return None
    try:
        parts = [int(x.strip()) for x in bounds_str.split(",")]
        if len(parts) != 4:
            return None
        l, t, r, b = parts
        if r <= l or b <= t:
            return None
        return l, t, r, b
    except Exception:
        return None


def _safe_regex_search(pattern: str, value: str) -> bool:
    if not pattern:
        return False
    return bool(re.search(pattern, value or "", flags=re.IGNORECASE))


def _extract_candidate_from_node(node: dict[str, Any], index: int = -1) -> dict[str, Any]:
    text = str(node.get("text", "") or "").strip()
    description = str(node.get("contentDescription", "") or "").strip()
    announcement = str(node.get("talkbackLabel", "") or "").strip()
    if not announcement:
        announcement = f"{text} {description}".strip()
    resource_id = str(node.get("viewIdResourceName", "") or "").strip()
    class_name = str(node.get("className", "") or "").strip()
    bounds = str(node.get("boundsInScreen", "") or "").strip()
    parsed = parse_bounds_str(bounds)
    top, left, right, bottom = (parsed[1], parsed[0], parsed[2], parsed[3]) if parsed else (10**9, 10**9, -1, -1)
    return {
        "source": "dump_tree",
        "index": index,
        "text": text,
        "class_name": class_name,
        "announcement": announcement,
        "resource_id": resource_id,
        "bounds": bounds,
        "top": top,
        "left": left,
        "right": right,
        "bottom": bottom,
    }


def _extract_candidate_from_step(step: dict[str, Any]) -> dict[str, Any]:
    bounds = str(step.get("focus_bounds", "") or "").strip()
    parsed = parse_bounds_str(bounds)
    top, left, right, bottom = (parsed[1], parsed[0], parsed[2], parsed[3]) if parsed else (10**9, 10**9, -1, -1)
    return {
        "source": "focus_step",
        "index": -1,
        "text": str(step.get("visible_label", "") or "").strip(),
        "class_name": str(step.get("focus_node", {}).get("className", "") or "").strip(),
        "announcement": str(step.get("merged_announcement", "") or "").strip(),
        "resource_id": str(step.get("focus_view_id", "") or "").strip(),
        "bounds": bounds,
        "top": top,
        "left": left,
        "right": right,
        "bottom": bottom,
    }


def _resolve_anchor_cfg(tab_cfg: dict[str, Any]) -> dict[str, Any]:
    anchor_cfg = dict(tab_cfg.get("anchor", {}) or {})
    if "tie_breaker" not in anchor_cfg:
        anchor_cfg["tie_breaker"] = "top_left"
    anchor_cfg["allow_resource_id_only"] = bool(anchor_cfg.get("allow_resource_id_only", False))
    if not anchor_cfg.get("text_regex") and tab_cfg.get("anchor_name"):
        anchor_type = str(tab_cfg.get("anchor_type", "") or "").lower()
        if anchor_type in {"t", "b", "a"}:
            anchor_cfg["text_regex"] = str(tab_cfg.get("anchor_name") or "")
        if anchor_type in {"r", "a"}:
            anchor_cfg["resource_id_regex"] = str(tab_cfg.get("anchor_name") or "")
    return anchor_cfg


def _match_composite_candidate(candidate: dict[str, Any], match_cfg: dict[str, Any]) -> dict[str, Any]:
    matched_fields: list[str] = []
    score = 0

    resource_id_regex = str(match_cfg.get("resource_id_regex", "") or "").strip()
    text_regex = str(match_cfg.get("text_regex", "") or "").strip()
    announcement_regex = str(match_cfg.get("announcement_regex", "") or "").strip()
    class_name_regex = str(match_cfg.get("class_name_regex", "") or "").strip()
    bounds_regex = str(match_cfg.get("bounds_regex", "") or "").strip()
    allow_resource_id_only = bool(match_cfg.get("allow_resource_id_only", False))

    if resource_id_regex and _safe_regex_search(resource_id_regex, candidate.get("resource_id", "")):
        matched_fields.append("resource_id")
        score += 100
    if text_regex and _safe_regex_search(text_regex, candidate.get("text", "")):
        matched_fields.append("text")
        score += 40
    if announcement_regex and _safe_regex_search(announcement_regex, candidate.get("announcement", "")):
        matched_fields.append("announcement")
        score += 30
    if class_name_regex and _safe_regex_search(class_name_regex, candidate.get("class_name", "")):
        matched_fields.append("class_name")
        score += 20
    if bounds_regex and _safe_regex_search(bounds_regex, candidate.get("bounds", "")):
        matched_fields.append("bounds")
        score += 10

    has_resource_match = "resource_id" in matched_fields
    has_other_match = any(field in matched_fields for field in ("text", "announcement", "class_name"))
    matched = has_resource_match and (has_other_match or allow_resource_id_only)
    if not resource_id_regex:
        matched = bool(matched_fields)

    return {
        "matched": matched,
        "score": score,
        "matched_fields": matched_fields,
        "candidate": candidate,
        "allow_resource_id_only": allow_resource_id_only,
    }


def match_anchor(candidate: dict[str, Any], anchor_cfg: dict[str, Any]) -> dict[str, Any]:
    return _match_composite_candidate(candidate, anchor_cfg)


def normalize_tab_config(tab_cfg: dict[str, Any]) -> dict[str, Any]:
    normalized_tab_cfg = dict(tab_cfg.get("tab", {}) or {})
    fallback_to_legacy = not bool(normalized_tab_cfg)
    if fallback_to_legacy:
        tab_name = str(tab_cfg.get("tab_name", "") or "").strip()
        tab_type = str(tab_cfg.get("tab_type", "") or "").strip().lower()
        if tab_type in {"t", "b", "a"} and tab_name:
            normalized_tab_cfg["text_regex"] = tab_name
        if tab_type in {"r", "a"} and tab_name:
            normalized_tab_cfg["resource_id_regex"] = tab_name
    if "tie_breaker" not in normalized_tab_cfg:
        normalized_tab_cfg["tie_breaker"] = "bottom_nav_left_to_right"
    normalized_tab_cfg["allow_resource_id_only"] = bool(normalized_tab_cfg.get("allow_resource_id_only", False))
    normalized_tab_cfg["_fallback_to_legacy"] = fallback_to_legacy
    return normalized_tab_cfg


def match_tab_candidate(node: dict[str, Any], tab_cfg: dict[str, Any]) -> dict[str, Any]:
    candidate = _extract_candidate_from_node(node)
    return _match_composite_candidate(candidate, tab_cfg)


def choose_best_anchor_candidate(matches: list[dict[str, Any]], tie_breaker: str = "top_left") -> dict[str, Any] | None:
    if not matches:
        return None
    if tie_breaker == "top_left":
        return sorted(
            matches,
            key=lambda item: (
                -int(item.get("score", 0)),
                int(item["candidate"].get("top", 10**9)),
                int(item["candidate"].get("left", 10**9)),
            ),
        )[0]
    return sorted(matches, key=lambda item: -int(item.get("score", 0)))[0]


def choose_best_tab_candidate(matches: list[dict[str, Any]], tie_breaker: str = "first_match") -> dict[str, Any] | None:
    if not matches:
        return None
    if tie_breaker == "bottom_nav_left_to_right":
        return sorted(
            matches,
            key=lambda item: (
                -int(item.get("score", 0)),
                -int(item["candidate"].get("top", -1)),
                int(item["candidate"].get("left", 10**9)),
                int("resource_id" in item.get("matched_fields", [])) * -1,
            ),
        )[0]
    if tie_breaker == "top_left":
        return sorted(
            matches,
            key=lambda item: (
                -int(item.get("score", 0)),
                int(item["candidate"].get("top", 10**9)),
                int(item["candidate"].get("left", 10**9)),
            ),
        )[0]
    return sorted(matches, key=lambda item: -int(item.get("score", 0)))[0]


def verify_context(
    step: dict[str, Any],
    scenario_cfg: dict[str, Any],
    client: A11yAdbClient | None = None,
    dev: str = "",
) -> dict[str, Any]:
    context_cfg = dict(scenario_cfg.get("context_verify", {}) or {})
    context_type = str(context_cfg.get("type", "none") or "none").strip().lower()
    if context_type in {"", "none"}:
        return {
            "ok": True,
            "type": "none",
            "expected": "",
            "actual_text": "",
            "actual_announcement": "",
        }

    text_regex = str(context_cfg.get("text_regex", "") or "").strip()
    announcement_regex = str(context_cfg.get("announcement_regex", "") or "").strip()

    if context_type == "selected_bottom_tab":
        nodes = step.get("dump_tree_nodes", [])
        dump_source = "step_cache"
        lazy_dump_node_count = 0
        if not isinstance(nodes, list) or not nodes:
            if client and dev:
                lazy_nodes = client.dump_tree(dev=dev)
                if isinstance(lazy_nodes, list):
                    nodes = lazy_nodes
                    step["dump_tree_nodes"] = lazy_nodes
                    lazy_dump_node_count = len(lazy_nodes)
                    dump_source = "lazy_dump"
                else:
                    nodes = []
        selected_candidates: list[str] = []
        selected_values: list[str] = []
        fallback_values: list[str] = []

        for node in nodes if isinstance(nodes, list) else []:
            if not isinstance(node, dict):
                continue
            text = str(node.get("text", "") or "").strip()
            description = str(node.get("contentDescription", "") or "").strip()
            view_id = str(node.get("viewIdResourceName", "") or "").strip()
            bounds = str(node.get("boundsInScreen", "") or "").strip()
            selected_raw = node.get("selected")
            selected_state = bool(selected_raw) if not isinstance(selected_raw, str) else selected_raw.strip().lower() == "true"
            combined = ", ".join(part for part in [description, text] if part).strip()
            if not combined:
                combined = str(node.get("talkbackLabel", "") or "").strip()
            marker = f"text='{text}' desc='{description}' selected={selected_state} viewId='{view_id}' bounds='{bounds}'"
            selected_candidates.append(marker)
            if combined:
                fallback_values.append(combined)
            if selected_state or re.search(r"(selected|선택됨)", combined or "", flags=re.IGNORECASE):
                selected_values.append(combined or marker)

        actual_selected_text = selected_values[0] if selected_values else ""
        actual_text = actual_selected_text
        actual_announcement = actual_selected_text

        # selected_bottom_tab은 dump 기반 selected 후보에서만 판정한다.
        text_ok = True if not text_regex else _safe_regex_search(text_regex, actual_selected_text)
        announcement_ok = True if not announcement_regex else _safe_regex_search(announcement_regex, actual_selected_text)
        ok = text_ok and announcement_ok

        if not actual_selected_text and fallback_values:
            actual_selected_text = fallback_values[0]
            actual_text = actual_selected_text
            actual_announcement = actual_selected_text

        expected_parts = []
        if text_regex:
            expected_parts.append(f"text={text_regex}")
        if announcement_regex:
            expected_parts.append(f"announcement={announcement_regex}")

        return {
            "ok": ok,
            "type": context_type,
            "expected": " | ".join(expected_parts),
            "actual_text": actual_text,
            "actual_announcement": actual_announcement,
            "actual_selected_text": actual_selected_text,
            "selected_candidates": selected_candidates,
            "dump_source": dump_source,
            "lazy_dump_node_count": lazy_dump_node_count,
        }

    actual_text = str(step.get("visible_label", "") or "").strip()
    actual_announcement = str(step.get("merged_announcement", "") or "").strip()
    text_ok = True if not text_regex else _safe_regex_search(text_regex, actual_text)
    announcement_ok = True if not announcement_regex else _safe_regex_search(announcement_regex, actual_announcement)
    ok = text_ok and announcement_ok

    expected_parts = []
    if text_regex:
        expected_parts.append(f"text={text_regex}")
    if announcement_regex:
        expected_parts.append(f"announcement={announcement_regex}")

    return {
        "ok": ok,
        "type": context_type,
        "expected": " | ".join(expected_parts),
        "actual_text": actual_text,
        "actual_announcement": actual_announcement,
    }


def sanitize_filename(value: str) -> str:
    keep = []
    for ch in value:
        if ch.isalnum() or ch in ("_", "-", "."):
            keep.append(ch)
        else:
            keep.append("_")
    return "".join(keep).strip("_") or "item"


def capture_full_screenshot(client: A11yAdbClient, dev: str, save_path: str) -> None:
    # talkback_lib 내부 private helper 재사용
    client._take_snapshot(dev, save_path)


def crop_image_by_bounds(
    screenshot_path: str,
    bounds_str: str,
    crop_path: str,
    shrink_px: int = 0,
) -> bool:
    bounds = parse_bounds_str(bounds_str)
    if not bounds:
        return False

    l, t, r, b = bounds
    with Image.open(screenshot_path) as img:
        width, height = img.size

        l = max(0, l + shrink_px)
        t = max(0, t + shrink_px)
        r = min(width, r - shrink_px)
        b = min(height, b - shrink_px)

        if r <= l or b <= t:
            return False

        cropped = img.crop((l, t, r, b))
        Path(crop_path).parent.mkdir(parents=True, exist_ok=True)
        cropped.save(crop_path)
        cropped.close()
    return True


def maybe_capture_focus_crop(
    client: A11yAdbClient,
    dev: str,
    row: dict,
    output_base_dir: str,
) -> dict:
    row["t_before_crop"] = round(time.monotonic() - float(row.get("_step_mono_start", time.monotonic())), 3) if row.get("_step_mono_start") else 0.0
    row["crop_image_path"] = ""
    row["crop_image_saved"] = False
    row["crop_bounds"] = str(row.get("focus_bounds", "") or "").strip()
    row["crop_source"] = "focus_bounds"
    row["crop_focus_confidence_low"] = False

    if not ENABLE_IMAGE_CROP:
        row["t_after_crop"] = row["t_before_crop"]
        return row

    bounds_str = str(row.get("focus_bounds", "") or "").strip()
    if not bounds_str:
        row["t_after_crop"] = row["t_before_crop"]
        return row

    tab_name = sanitize_filename(str(row.get("tab_name", "unknown")))
    step_index = row.get("step_index", -1)
    visible_label = sanitize_filename(str(row.get("visible_label", "") or "")[:40])

    crop_dir = Path(output_base_dir) / "crops"
    crop_dir.mkdir(parents=True, exist_ok=True)

    crop_path = crop_dir / f"{tab_name}_step_{step_index}_{visible_label}.png"

    capture_started = time.perf_counter()
    screenshot_path = ""
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".png",
            prefix=f"tb_step_{step_index}_",
            delete=False,
        ) as temp_file:
            screenshot_path = temp_file.name
        capture_full_screenshot(client, dev, screenshot_path)
        ok = crop_image_by_bounds(
            screenshot_path=screenshot_path,
            bounds_str=bounds_str,
            crop_path=str(crop_path),
            shrink_px=2,
        )
        if ok:
            row["crop_image_path"] = str(crop_path)
            row["crop_image_saved"] = True
        row["screenshot_capture_elapsed"] = round(time.perf_counter() - capture_started, 3)
    except Exception as exc:
        log(f"[IMAGE] crop failed step={step_index}: {exc}")
    finally:
        if screenshot_path:
            try:
                Path(screenshot_path).unlink(missing_ok=True)
            except Exception:
                pass
        row["crop_elapsed_sec"] = round(time.perf_counter() - capture_started, 3)
        if row.get("_step_mono_start"):
            row["t_after_crop"] = round(time.monotonic() - float(row["_step_mono_start"]), 3)
        else:
            row["t_after_crop"] = row.get("t_before_crop", 0.0)
        payload_source = str(row.get("focus_payload_source", "") or "").lower()
        response_success = bool(row.get("get_focus_response_success", False))
        focus_view_id = str(row.get("focus_view_id", "") or "").strip()
        row["crop_focus_confidence_low"] = bool(
            (payload_source == "top_level" and not response_success)
            or (not focus_view_id and bool(row.get("crop_bounds", "")))
        )

    return row


def add_rule_compare(df: pd.DataFrame) -> pd.DataFrame:
    def compare_row(row) -> str:
        visible = str(row.get("normalized_visible_label", "") or "").strip()
        speech = str(row.get("normalized_announcement", "") or "").strip()

        if row.get("status") == "ANCHOR":
            return "SKIP"

        if not visible and not speech:
            return "EMPTY"
        if visible == speech:
            return "EXACT"
        if visible and speech and (visible in speech or speech in visible):
            return "PARTIAL"
        return "DIFF"

    df["rule_compare"] = df.apply(compare_row, axis=1)
    return df


def stringify_complex_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: to_json_text(x) if isinstance(x, (list, dict)) else x
        )
    return df


def insert_images_to_excel(excel_path: str, image_col_name: str = "crop_image") -> None:
    if not ENABLE_IMAGE_INSERT_TO_EXCEL:
        return

    wb = load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if image_col_name not in headers or "crop_image_path" not in headers:
        wb.save(excel_path)
        return

    image_col_idx = headers.index(image_col_name) + 1
    path_col_idx = headers.index("crop_image_path") + 1

    col_letter = ws.cell(row=1, column=image_col_idx).column_letter

    for row_idx in range(2, ws.max_row + 1):
        path_value = ws.cell(row=row_idx, column=path_col_idx).value
        if not path_value:
            continue

        img_path = Path(str(path_value))
        if not img_path.exists():
            continue

        try:
            img = XLImage(str(img_path))
            img.width = 90
            img.height = 90
            ws.add_image(img, f"{col_letter}{row_idx}")
            ws.row_dimensions[row_idx].height = 72
        except Exception as exc:
            log(f"[EXCEL] image insert failed row={row_idx}: {exc}")

    ws.column_dimensions[col_letter].width = 16
    wb.save(excel_path)


def save_excel(rows: list[dict], output_path: str, with_images: bool = True) -> None:
    df = pd.DataFrame(rows)
    if df.empty:
        log("[SAVE] skip: no rows")
        return

    df = add_rule_compare(df)
    df = stringify_complex_columns(df)

    ordered_cols = [
        "tab_name",
        "context_type",
        "parent_step_index",
        "overlay_entry_label",
        "overlay_recovery_status",
        "step_index",
        "status",
        "stop_reason",
        "step_elapsed_sec",
        "move_result",
        "visible_label",
        "normalized_visible_label",
        "merged_announcement",
        "normalized_announcement",
        "rule_compare",
        "focus_text",
        "focus_content_description",
        "focus_view_id",
        "focus_bounds",
        "crop_image",
        "crop_image_path",
        "crop_image_saved",
        "partial_announcements",
        "last_announcements",
        "last_merged_announcement",
        "focus_node",
        "dump_tree_nodes",
    ]

    existing_cols = [c for c in ordered_cols if c in df.columns]
    remaining_cols = [c for c in df.columns if c not in existing_cols]
    df = df[existing_cols + remaining_cols]

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)

    if with_images and "crop_image" in df.columns:
        insert_images_to_excel(output_path, image_col_name="crop_image")

    log(f"[SAVE] saved excel: {output_path} rows={len(df)} with_images={with_images}")


def _matches_overlay_candidate(step: dict[str, Any], entry: dict[str, Any]) -> bool:
    focus_view_id = str(step.get("focus_view_id", "") or "").strip()
    normalized_visible_label = str(step.get("normalized_visible_label", "") or "").strip()
    entry_view_id = str(entry.get("resource_id", "") or "").strip()
    entry_label = str(entry.get("label", "") or "").strip().lower()
    if entry_view_id and focus_view_id == entry_view_id:
        return True
    return bool(entry_label and normalized_visible_label == entry_label)


def _get_overlay_policy_entries(tab_cfg: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str]:
    policy = tab_cfg.get("overlay_policy")
    if isinstance(policy, dict):
        allow_candidates = list(policy.get("allow_candidates", []) or [])
        block_candidates = list(policy.get("block_candidates", []) or [])
        return allow_candidates, block_candidates, "scenario_policy"
    return OVERLAY_ENTRY_CANDIDATES, [], "global_candidates"


def is_overlay_candidate(step: dict[str, Any], tab_cfg: dict[str, Any]) -> tuple[bool, str]:
    allow_candidates, block_candidates, source = _get_overlay_policy_entries(tab_cfg)

    for entry in block_candidates:
        if _matches_overlay_candidate(step, entry):
            return False, f"blocked_by_{source}"

    for entry in allow_candidates:
        if _matches_overlay_candidate(step, entry):
            return True, f"matched_{source}"

    return False, f"not_in_{source}"


def _node_signature(step: dict[str, Any]) -> set[str]:
    signatures: set[str] = set()
    nodes = step.get("dump_tree_nodes", [])
    for node in nodes if isinstance(nodes, list) else []:
        if not isinstance(node, dict):
            continue
        view_id = str(node.get("viewIdResourceName", "") or "").strip()
        text = str(node.get("text", "") or "").strip().lower()
        desc = str(node.get("contentDescription", "") or "").strip().lower()
        marker = f"{view_id}|{text}|{desc}".strip("|")
        if marker:
            signatures.add(marker)
    return signatures


def classify_post_click_result(
    client: A11yAdbClient,
    dev: str,
    tab_cfg: dict[str, Any],
    pre_click_step: dict[str, Any],
) -> tuple[str, dict[str, Any]]:
    post_click_step = client.collect_focus_step(
        dev=dev,
        step_index=int(pre_click_step.get("step_index", 0) or 0),
        move=False,
        wait_seconds=MAIN_STEP_WAIT_SECONDS,
        announcement_wait_seconds=MAIN_ANNOUNCEMENT_WAIT_SECONDS,
    )

    pre_fp = make_main_fingerprint(pre_click_step)
    post_fp = make_main_fingerprint(post_click_step)
    if pre_fp == post_fp and any(pre_fp):
        return "unchanged", post_click_step

    pre_signature = _node_signature(pre_click_step)
    post_signature = _node_signature(post_click_step)
    overlap_ratio = 0.0
    if pre_signature and post_signature:
        overlap_ratio = len(pre_signature & post_signature) / float(max(len(pre_signature), 1))

    post_label = str(post_click_step.get("visible_label", "") or "").strip().lower()
    post_announcement = str(post_click_step.get("merged_announcement", "") or "").strip().lower()
    post_view_id = str(post_click_step.get("focus_view_id", "") or "").strip().lower()
    navigation_cues = ("navigate up", "back", "up button")
    toolbar_cues = ("toolbar", "action_bar", "appbar", "title")
    looks_like_navigation = (
        any(cue in post_label for cue in navigation_cues)
        or any(cue in post_announcement for cue in navigation_cues)
        or any(cue in post_view_id for cue in toolbar_cues)
    )

    if looks_like_navigation or (pre_signature and post_signature and overlap_ratio < 0.30):
        return "navigation", post_click_step

    if pre_signature and post_signature and 0.30 <= overlap_ratio < 0.45:
        log(
            f"[WARN] overlay classification low-confidence "
            f"overlap_ratio={overlap_ratio:.2f} "
            f"pre_label='{pre_click_step.get('visible_label', '')}' "
            f"post_label='{post_click_step.get('visible_label', '')}'"
        )

    return "overlay", post_click_step


def make_overlay_entry_fingerprint(tab_name: str, step: dict[str, Any]) -> str:
    focus_view_id = str(step.get("focus_view_id", "") or "").strip()
    normalized_visible_label = str(step.get("normalized_visible_label", "") or "").strip()
    return f"{tab_name}|{focus_view_id}|{normalized_visible_label}"


def make_main_fingerprint(step: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(step.get("normalized_visible_label", "") or "").strip(),
        str(step.get("focus_view_id", "") or "").strip(),
        str(step.get("focus_bounds", "") or "").strip(),
    )


def _bounds_changed_significantly(prev_bounds: str, curr_bounds: str) -> bool:
    prev = parse_bounds_str(prev_bounds)
    curr = parse_bounds_str(curr_bounds)
    if not prev or not curr:
        return False
    pl, pt, pr, pb = prev
    cl, ct, cr, cb = curr
    center_dx = abs(((pl + pr) / 2.0) - ((cl + cr) / 2.0))
    center_dy = abs(((pt + pb) / 2.0) - ((ct + cb) / 2.0))
    return center_dx > 80 or center_dy > 80


def detect_step_mismatch(
    row: dict[str, Any],
    previous_step: dict[str, Any] | None = None,
) -> tuple[list[str], list[str]]:
    mismatch_reasons: list[str] = []
    low_confidence_reasons: list[str] = []
    visible = str(row.get("normalized_visible_label", "") or "").strip()
    speech = str(row.get("normalized_announcement", "") or "").strip()
    focus_source = str(row.get("focus_payload_source", "") or "").strip().lower()
    response_success = bool(row.get("get_focus_response_success", False))
    top_level_suspicious = bool(row.get("get_focus_top_level_success_false", False))
    focus_view_id = str(row.get("focus_view_id", "") or "").strip()
    focus_bounds = str(row.get("focus_bounds", "") or "").strip()
    context_type = str(row.get("context_type", "") or "").strip().lower()

    if top_level_suspicious or (focus_source == "top_level" and not response_success):
        low_confidence_reasons.append("get_focus_top_level_success_false")

    visible_terms = [token for token in visible.split(" ") if token]
    speech_terms = [token for token in speech.split(" ") if token]
    speech_visible_compatible = (
        not visible
        or not speech
        or visible == speech
        or visible in speech
        or speech in visible
        or (
            bool(visible_terms)
            and bool(speech_terms)
            and (speech_terms[0] == visible_terms[0] or visible_terms[0] in speech_terms[:2])
        )
    )
    if visible and speech and not speech_visible_compatible:
        mismatch_reasons.append("speech_visible_diverged")

    if context_type == "overlay" and not focus_view_id and focus_bounds:
        low_confidence_reasons.append("overlay_bounds_only_focus")

    prev = previous_step or {}
    prev_speech = str(prev.get("normalized_announcement", "") or "").strip()
    prev_bounds = str(prev.get("focus_bounds", "") or "").strip()
    if prev_speech and speech and prev_speech == speech and _bounds_changed_significantly(prev_bounds, focus_bounds):
        mismatch_reasons.append("speech_bounds_diverged")

    if (
        context_type == "main"
        and str(row.get("overlay_recovery_status", "") or "").strip().lower().startswith("realign")
        and not focus_view_id
        and focus_bounds
        and visible
        and speech
        and not speech_visible_compatible
    ):
        mismatch_reasons.append("overlay_realign_bounds_only_then_label_mismatch")

    if bool(row.get("crop_focus_confidence_low", False)):
        low_confidence_reasons.append("crop_low_confidence")

    fallback_found = bool(row.get("get_focus_fallback_found", False))
    success_false_top_level_dump_found = bool(row.get("get_focus_success_false_top_level_dump_found", False))
    if (
        focus_source == "top_level"
        and not fallback_found
        and not success_false_top_level_dump_found
    ):
        low_confidence_reasons.append("top_level_without_fallback_dump")
    if not focus_view_id and focus_bounds:
        low_confidence_reasons.append("bounds_dependent_focus")

    return mismatch_reasons, low_confidence_reasons


def is_overlay_entry_focus(current_step: dict[str, Any], entry_step: dict[str, Any]) -> bool:
    current_view_id = str(current_step.get("focus_view_id", "") or "").strip()
    entry_view_id = str(entry_step.get("focus_view_id", "") or "").strip()
    if current_view_id and entry_view_id and current_view_id == entry_view_id:
        return True

    current_label = str(current_step.get("normalized_visible_label", "") or "").strip()
    entry_label = str(entry_step.get("normalized_visible_label", "") or "").strip()
    if current_label and entry_label and current_label == entry_label:
        return True

    current_bounds = str(current_step.get("focus_bounds", "") or "").strip()
    entry_bounds = str(entry_step.get("focus_bounds", "") or "").strip()
    return bool(current_bounds and entry_bounds and current_bounds == entry_bounds)


def get_overlay_entry_match_by(current_step: dict[str, Any], entry_step: dict[str, Any]) -> str:
    current_view_id = str(current_step.get("focus_view_id", "") or "").strip()
    entry_view_id = str(entry_step.get("focus_view_id", "") or "").strip()
    if current_view_id and entry_view_id and current_view_id == entry_view_id:
        return "view_id"
    current_label = str(current_step.get("normalized_visible_label", "") or "").strip()
    entry_label = str(entry_step.get("normalized_visible_label", "") or "").strip()
    if current_label and entry_label and current_label == entry_label:
        return "label"
    current_bounds = str(current_step.get("focus_bounds", "") or "").strip()
    entry_bounds = str(entry_step.get("focus_bounds", "") or "").strip()
    if current_bounds and entry_bounds and current_bounds == entry_bounds:
        return "bounds"
    return ""


def collect_realign_probe(
    client: A11yAdbClient,
    dev: str,
    move: bool,
    probe_idx: int = 0,
    direction: str = "next",
    wait_seconds: float = MAIN_STEP_WAIT_SECONDS,
) -> dict[str, Any]:
    probe: dict[str, Any] = {
        "move_result": None,
        "move_elapsed_sec": 0.0,
        "get_focus_elapsed_sec": 0.0,
        "focus_view_id": "",
        "focus_bounds": "",
        "visible_label": "",
        "normalized_visible_label": "",
        "realign_probe_idx": probe_idx,
        "realign_move_result": "",
        "realign_focus_source": "none",
    }

    if move:
        move_started = time.monotonic()
        try:
            if str(direction).strip().lower() == "next":
                probe["move_result"] = client.move_focus_smart(dev=dev, direction=direction)
            else:
                probe["move_result"] = "moved" if client.move_focus(dev=dev, direction=direction) else "failed"
        except Exception as exc:  # defensive
            probe["move_result"] = f"error: {exc}"
        probe["move_elapsed_sec"] = round(time.monotonic() - move_started, 3)
    probe["realign_move_result"] = str(probe.get("move_result", "") or "")

    focus_started = time.monotonic()
    try:
        focus_node = client.get_focus(dev=dev, wait_seconds=wait_seconds)
    except Exception:
        focus_node = {}
    probe["get_focus_elapsed_sec"] = round(time.monotonic() - focus_started, 3)

    safe_focus_node = focus_node if isinstance(focus_node, dict) else {}
    probe["focus_view_id"] = str(safe_focus_node.get("viewIdResourceName", "") or "").strip()
    probe["visible_label"] = client.extract_visible_label_from_focus(safe_focus_node)
    probe["normalized_visible_label"] = client.normalize_for_comparison(probe["visible_label"])

    normalize_bounds = getattr(client, "_normalize_bounds", None)
    if callable(normalize_bounds):
        probe["focus_bounds"] = str(normalize_bounds(safe_focus_node) or "").strip()
    trace = getattr(client, "last_get_focus_trace", {})
    if isinstance(trace, dict):
        probe["realign_focus_source"] = str(trace.get("focus_payload_source", "none") or "none")

    log(
        f"[OVERLAY] realign probe move={move} "
        f"move_elapsed={probe['move_elapsed_sec']:.3f}s "
        f"focus_elapsed={probe['get_focus_elapsed_sec']:.3f}s "
        f"view_id='{probe['focus_view_id']}' "
        f"label='{probe['visible_label']}'",
        level="DEBUG",
    )
    return probe


def realign_focus_after_overlay(
    client: A11yAdbClient,
    dev: str,
    entry_step: dict[str, Any],
    known_step_index_by_fingerprint: dict[tuple[str, str, str], int],
) -> dict[str, Any]:
    current_step = collect_realign_probe(
        client=client,
        dev=dev,
        move=False,
        probe_idx=0,
        wait_seconds=MAIN_STEP_WAIT_SECONDS,
    )
    current_fp = make_main_fingerprint(current_step)
    entry_idx = int(entry_step.get("step_index", 0) or 0)

    current_match_by = get_overlay_entry_match_by(current_step, entry_step)
    if current_match_by:
        if current_match_by == "bounds":
            log(
                f"[WARN] overlay realign matched by bounds only "
                f"probe_idx=0 entry_label='{entry_step.get('visible_label', '')}'",
            )
        return {
            "status": "already_on_entry",
            "steps_taken": 0,
            "entry_reached": True,
            "match_by": current_match_by,
            "current_step": current_step,
        }

    seen_idx = known_step_index_by_fingerprint.get(current_fp)
    if seen_idx is None or seen_idx >= entry_idx:
        return {
            "status": "skip_realign_not_before_entry",
            "steps_taken": 0,
            "entry_reached": False,
            "current_step": current_step,
        }

    for realign_idx in range(1, OVERLAY_REALIGN_MAX_STEPS + 1):
        probe_step = collect_realign_probe(
            client=client,
            dev=dev,
            move=True,
            probe_idx=realign_idx,
            direction="next",
            wait_seconds=MAIN_STEP_WAIT_SECONDS,
        )
        match_by = get_overlay_entry_match_by(probe_step, entry_step)
        if match_by:
            if match_by == "bounds":
                log(
                    f"[WARN] overlay realign matched by bounds only "
                    f"probe_idx={realign_idx} entry_label='{entry_step.get('visible_label', '')}'",
                )
            return {
                "status": "realign_entry_reached",
                "steps_taken": realign_idx,
                "entry_reached": True,
                "match_by": match_by,
                "current_step": probe_step,
            }

    return {
        "status": "realign_entry_not_found",
        "steps_taken": OVERLAY_REALIGN_MAX_STEPS,
        "entry_reached": False,
        "current_step": current_step,
    }


def expand_overlay(
    client: A11yAdbClient,
    dev: str,
    tab_cfg: dict[str, Any],
    entry_step: dict[str, Any],
    rows: list[dict[str, Any]],
    all_rows: list[dict[str, Any]],
    output_path: str,
    output_base_dir: str,
    skip_entry_click: bool = False,
) -> list[dict[str, Any]]:
    overlay_rows: list[dict[str, Any]] = []
    entry_label = str(entry_step.get("visible_label", "") or "").strip()
    entry_view_id = str(entry_step.get("focus_view_id", "") or "").strip()

    clicked = skip_entry_click
    if not clicked and entry_view_id:
        clicked = client.touch(
            dev=dev,
            name=f"^{re.escape(entry_view_id)}$",
            type_="r",
            wait_=3,
        )
    elif not clicked and entry_label:
        clicked = client.touch(
            dev=dev,
            name=f"^{re.escape(entry_label)}$",
            type_="a",
            wait_=3,
        )

    recovery_status = "not_attempted"
    if not clicked:
        recovery_status = "entry_click_failed"
        return overlay_rows

    if not skip_entry_click:
        time.sleep(1.0)

    parent_step_index = entry_step.get("step_index")
    overlay_prev_fingerprint = ("", "", "")
    overlay_fail_count = 0
    overlay_same_count = 0
    for overlay_step_idx in range(1, OVERLAY_MAX_STEPS + 1):
        overlay_row = client.collect_focus_step(
            dev=dev,
            step_index=overlay_step_idx,
            move=True,
            direction="next",
            wait_seconds=OVERLAY_STEP_WAIT_SECONDS,
            announcement_wait_seconds=OVERLAY_ANNOUNCEMENT_WAIT_SECONDS,
        )
        overlay_row["tab_name"] = tab_cfg["tab_name"]
        overlay_row["context_type"] = "overlay"
        overlay_row["parent_step_index"] = parent_step_index
        overlay_row["overlay_entry_label"] = entry_label
        overlay_row["overlay_recovery_status"] = ""
        overlay_row["status"] = "OK"
        overlay_row["stop_reason"] = ""
        overlay_row["crop_image"] = "IMAGE"
        overlay_row["_step_mono_start"] = time.monotonic() - float(overlay_row.get("t_step_start", 0.0) or 0.0)
        overlay_row = maybe_capture_focus_crop(client, dev, overlay_row, output_base_dir)
        overlay_row.pop("_step_mono_start", None)

        overlay_rows.append(overlay_row)
        rows.append(overlay_row)
        all_rows.append(overlay_row)

        (
            should_end_overlay,
            overlay_fail_count,
            overlay_same_count,
            overlay_reason,
            overlay_prev_fingerprint,
        ) = should_stop(
            row=overlay_row,
            prev_fingerprint=overlay_prev_fingerprint,
            fail_count=overlay_fail_count,
            same_count=overlay_same_count,
        )
        if should_end_overlay:
            overlay_row["status"] = "END"
            overlay_row["stop_reason"] = overlay_reason
            save_excel(all_rows, output_path, with_images=False)
            break
        if overlay_step_idx % CHECKPOINT_SAVE_EVERY_STEPS == 0:
            save_excel(all_rows, output_path, with_images=False)

    recovery_anchor = str(entry_step.get("normalized_visible_label", "") or "").strip()
    scenario_anchor = str(tab_cfg.get("anchor_name", "") or "").strip()
    expected_anchor: str | None = recovery_anchor or scenario_anchor or None

    recovery_result = client.press_back_and_recover_focus(
        dev=dev,
        expected_parent_anchor=expected_anchor,
        wait_seconds=BACK_RECOVERY_WAIT_SECONDS,
        retry=1,
    )
    recovery_status = str(recovery_result.get("status", "") or "")
    if recovery_status != "ok" and scenario_anchor:
        select_ok = client.select(
            dev=dev,
            name=scenario_anchor,
            type_=str(tab_cfg.get("anchor_type", "a") or "a"),
            wait_=3,
        )
        recovery_status = "ok_select_fallback" if select_ok else f"{recovery_status}_select_fallback_failed"

    if overlay_rows:
        overlay_rows[-1]["overlay_recovery_status"] = recovery_status
    save_excel(all_rows, output_path, with_images=False)
    return overlay_rows


def stabilize_anchor(
    client: A11yAdbClient,
    dev: str,
    tab_cfg: dict[str, Any],
    phase: str,
    max_retries: int = 2,
    verify_reads: int = 2,
) -> dict[str, Any]:
    anchor_cfg = _resolve_anchor_cfg(tab_cfg)
    tie_breaker = str(anchor_cfg.get("tie_breaker", "top_left") or "top_left")
    last_verify: dict[str, Any] = {}
    last_context: dict[str, Any] = {"ok": True, "type": "none", "expected": ""}
    scenario_id = str(tab_cfg.get("scenario_id", "") or "")

    for attempt in range(1, max_retries + 1):
        dump_nodes = client.dump_tree(dev=dev)
        candidates = [
            _extract_candidate_from_node(node, index=i)
            for i, node in enumerate(dump_nodes if isinstance(dump_nodes, list) else [])
        ]
        matches = [m for m in (match_anchor(c, anchor_cfg) for c in candidates) if m["matched"]]
        best = choose_best_anchor_candidate(matches, tie_breaker=tie_breaker)

        selected = False
        if best and best["candidate"].get("resource_id"):
            resource_pattern = f"^{re.escape(str(best['candidate']['resource_id']))}$"
            selected = client.select(
                dev=dev,
                name=resource_pattern,
                type_="r",
                wait_=8,
            )

        if not selected:
            selected = client.select(
                dev=dev,
                name=str(tab_cfg.get("anchor_name", "") or ""),
                type_=str(tab_cfg.get("anchor_type", "a") or "a"),
                wait_=8,
            )

        verify_match: dict[str, Any] | None = None
        context_result: dict[str, Any] = {"ok": True, "type": "none", "expected": ""}
        verify_rows: list[dict[str, Any]] = []
        for verify_idx in range(max(1, verify_reads)):
            verify_row = client.collect_focus_step(
                dev=dev,
                step_index=-(attempt * 10 + verify_idx),
                move=False,
                wait_seconds=MAIN_STEP_WAIT_SECONDS,
                announcement_wait_seconds=MAIN_ANNOUNCEMENT_WAIT_SECONDS,
            )
            verify_rows.append(verify_row)
            verify_candidate = _extract_candidate_from_step(verify_row)
            verify_match = match_anchor(verify_candidate, anchor_cfg)
            context_result = verify_context(verify_row, tab_cfg, client=client, dev=dev)
            if verify_match["matched"]:
                break

        last_verify = verify_match or {}
        last_context = context_result
        log(
            f"[ANCHOR][{phase}] attempt={attempt}/{max_retries} selected={selected} "
            f"matched={bool(last_verify.get('matched'))} "
            f"context_ok={bool(last_context.get('ok'))} "
            f"scenario='{scenario_id}' "
            f"fields={last_verify.get('matched_fields', [])} "
            f"score={last_verify.get('score', 0)} "
            f"resource='{(last_verify.get('candidate') or {}).get('resource_id', '')}' "
            f"bounds='{(last_verify.get('candidate') or {}).get('bounds', '')}'"
        )
        if str(last_context.get("type", "")) == "selected_bottom_tab":
            expected_value = (
                str(dict(tab_cfg.get("context_verify", {}) or {}).get("announcement_regex", "") or "").strip()
                or str(dict(tab_cfg.get("context_verify", {}) or {}).get("text_regex", "") or "").strip()
            )
            log(
                f"[CONTEXT][dump] scenario='{scenario_id}' type='selected_bottom_tab' "
                f"expected='{expected_value}'"
            )
            selected_candidates = last_context.get("selected_candidates", [])
            log(f"[CONTEXT][dump] selected_candidates_count={len(selected_candidates) if isinstance(selected_candidates, list) else 0}")
            log(f"[CONTEXT][dump] selected_candidates={selected_candidates}", level="DEBUG")
            log(f"[CONTEXT][dump] actual_selected_text='{last_context.get('actual_selected_text', '')}'")
            log(f"[CONTEXT][dump] source='{last_context.get('dump_source', 'step_cache')}'")
            log(f"[CONTEXT][dump] lazy_dump_node_count={int(last_context.get('lazy_dump_node_count', 0) or 0)}")
            log(f"[CONTEXT][dump] ok={bool(last_context.get('ok'))}")
        log(
            f"[CONTEXT] scenario='{scenario_id}' type='{last_context.get('type', 'none')}' "
            f"expected='{last_context.get('expected', '')}' "
            f"actual='{last_context.get('actual_selected_text', last_context.get('actual_announcement', last_context.get('actual_text', '')))}' "
            f"ok={bool(last_context.get('ok'))}"
        )
        if not bool(last_verify.get("matched")):
            log(f"[ANCHOR][{phase}] anchor mismatch scenario='{scenario_id}'")
        elif not bool(last_context.get("ok")):
            log(f"[ANCHOR][{phase}] context mismatch scenario='{scenario_id}'")
            log(f"[CONTEXT] verification failed scenario='{scenario_id}'")
        else:
            log(f"[CONTEXT] verification passed scenario='{scenario_id}'")

        if bool(last_verify.get("matched")) and bool(last_context.get("ok")):
            success_reason = "selected_and_verified" if selected else "verified_without_select"
            log(
                f"[ANCHOR][{phase}] success scenario='{scenario_id}' selected={selected} "
                f"matched=True context_ok=True reason='{success_reason}'"
            )
            return {
                "ok": True,
                "attempt": attempt,
                "selected": selected,
                "reason": success_reason,
                "verify": last_verify,
                "context": last_context,
                "verify_rows": verify_rows,
                "candidate_count": len(matches),
                "phase": phase,
            }

    return {
        "ok": False,
        "attempt": max_retries,
        "selected": False,
        "verify": last_verify,
        "context": last_context,
        "candidate_count": 0,
        "phase": phase,
    }


def stabilize_tab_selection(
    client: A11yAdbClient,
    dev: str,
    tab_cfg: dict[str, Any],
    max_retries: int = 2,
) -> dict[str, Any]:
    normalized_tab_cfg = normalize_tab_config(tab_cfg)
    tie_breaker = str(normalized_tab_cfg.get("tie_breaker", "bottom_nav_left_to_right") or "bottom_nav_left_to_right")
    scenario_id = str(tab_cfg.get("scenario_id", "") or "")
    fallback_to_legacy = bool(normalized_tab_cfg.get("_fallback_to_legacy", False))
    if fallback_to_legacy:
        log(f"[TAB][select] fallback_to_legacy=True scenario='{scenario_id}'")

    last_context: dict[str, Any] = {"ok": True, "type": "none", "expected": ""}
    last_best: dict[str, Any] = {}
    for attempt in range(1, max_retries + 1):
        dump_nodes = client.dump_tree(dev=dev)
        node_list = dump_nodes if isinstance(dump_nodes, list) else []
        matches = [m for m in (match_tab_candidate(node, normalized_tab_cfg) for node in node_list) if m.get("matched")]
        best = choose_best_tab_candidate(matches, tie_breaker=tie_breaker)
        last_best = best or {}
        log(
            f"[TAB][select][debug] scenario='{scenario_id}' candidates={len(matches)} tie_breaker='{tie_breaker}'",
            level="DEBUG",
        )

        selected = False
        if best and best.get("candidate", {}).get("resource_id"):
            resource_pattern = f"^{re.escape(str(best['candidate']['resource_id']))}$"
            selected = client.select(dev=dev, name=resource_pattern, type_="r", wait_=5)

        if not selected:
            selected = client.touch(
                dev=dev,
                name=str(tab_cfg.get("tab_name", "") or ""),
                type_=str(tab_cfg.get("tab_type", "") or ""),
                wait_=5,
            )

        verify_row = client.collect_focus_step(
            dev=dev,
            step_index=-(500 + attempt),
            move=False,
            wait_seconds=MAIN_STEP_WAIT_SECONDS,
            announcement_wait_seconds=MAIN_ANNOUNCEMENT_WAIT_SECONDS,
        )
        last_context = verify_context(verify_row, tab_cfg, client=client, dev=dev)
        log(
            f"[TAB][select] scenario='{scenario_id}' selected={selected} "
            f"matched_fields={(best or {}).get('matched_fields', [])} score={(best or {}).get('score', 0)}"
        )
        log(
            f"[TAB][verify] selected_bottom_tab ok={bool(last_context.get('ok'))} "
            f"actual='{last_context.get('actual_selected_text', '')}'"
        )
        if _should_log("DEBUG") and best:
            log(
                f"[TAB][select][debug] best_resource='{best['candidate'].get('resource_id', '')}' "
                f"best_bounds='{best['candidate'].get('bounds', '')}'"
            )
        if selected and bool(last_context.get("ok")):
            return {
                "ok": True,
                "attempt": attempt,
                "selected": selected,
                "verify_context": last_context,
                "best": best,
                "candidate_count": len(matches),
            }
        if attempt < max_retries:
            log(f"[TAB][select] retry {attempt}/{max_retries} scenario='{scenario_id}'")

    return {
        "ok": False,
        "attempt": max_retries,
        "selected": False,
        "verify_context": last_context,
        "best": last_best,
    }


def open_tab_and_anchor(client: A11yAdbClient, dev: str, tab_cfg: dict) -> bool:
    tab_stabilized = stabilize_tab_selection(
        client=client,
        dev=dev,
        tab_cfg=tab_cfg,
        max_retries=2,
    )
    if not tab_stabilized.get("ok"):
        log(f"[TAB][select] stabilization failed scenario='{tab_cfg.get('scenario_id', '')}'")
        return False

    time.sleep(1.0)
    client.reset_focus_history(dev)
    time.sleep(0.5)

    stabilize_result = stabilize_anchor(
        client=client,
        dev=dev,
        tab_cfg=tab_cfg,
        phase="scenario_start",
        max_retries=2,
        verify_reads=2,
    )
    if not stabilize_result.get("ok"):
        log(f"[ANCHOR][scenario_start] stabilization failed tab='{tab_cfg.get('tab_name', '')}'")
        return False
    time.sleep(1.0)
    return True


def should_stop(
    row: dict,
    prev_fingerprint: tuple[str, str, str],
    fail_count: int,
    same_count: int,
) -> tuple[bool, int, int, str, tuple[str, str, str]]:
    move_result = str(row.get("move_result", "") or "")
    visible_label = str(row.get("visible_label", "") or "").strip()
    merged_announcement = str(row.get("merged_announcement", "") or "").strip()
    normalized_visible_label = str(row.get("normalized_visible_label", "") or "").strip()
    focus_view_id = str(row.get("focus_view_id", "") or "").strip()
    focus_bounds = str(row.get("focus_bounds", "") or "").strip()
    current_fingerprint = (
        normalized_visible_label,
        focus_view_id,
        focus_bounds,
    )

    reason = ""

    if move_result == "failed":
        fail_count += 1
    else:
        fail_count = 0

    if all(current_fingerprint) and current_fingerprint == prev_fingerprint:
        same_count += 1
    else:
        same_count = 0

    if fail_count >= 2:
        reason = "move_failed_twice"
        return True, fail_count, same_count, reason, current_fingerprint

    if same_count >= 3:
        reason = "same_fingerprint_repeated"
        return True, fail_count, same_count, reason, current_fingerprint

    if not visible_label and not merged_announcement:
        reason = "empty_visible_and_speech"
        return True, fail_count, same_count, reason, current_fingerprint

    return False, fail_count, same_count, reason, current_fingerprint


def collect_tab_rows(
    client: A11yAdbClient,
    dev: str,
    tab_cfg: dict,
    all_rows: list[dict],
    output_path: str,
    output_base_dir: str,
) -> list[dict]:
    rows: list[dict] = []

    opened = open_tab_and_anchor(client, dev, tab_cfg)
    if not opened:
        row = {
            "tab_name": tab_cfg["tab_name"],
            "step_index": -1,
            "status": "TAB_OPEN_FAILED",
            "stop_reason": "tab_or_anchor_failed",
            "crop_image": "",
            "crop_image_path": "",
            "crop_image_saved": False,
        }
        rows.append(row)
        all_rows.append(row)
        save_excel(all_rows, output_path, with_images=False)
        return rows

    anchor_start = time.perf_counter()
    anchor_row = client.collect_focus_step(
        dev=dev,
        step_index=0,
        move=False,
        wait_seconds=MAIN_STEP_WAIT_SECONDS,
        announcement_wait_seconds=MAIN_ANNOUNCEMENT_WAIT_SECONDS,
    )
    anchor_elapsed = time.perf_counter() - anchor_start

    anchor_row["tab_name"] = tab_cfg["tab_name"]
    anchor_row["context_type"] = "main"
    anchor_row["parent_step_index"] = ""
    anchor_row["overlay_entry_label"] = ""
    anchor_row["overlay_recovery_status"] = ""
    anchor_row["status"] = "ANCHOR"
    anchor_row["stop_reason"] = ""
    anchor_row["step_elapsed_sec"] = round(anchor_elapsed, 3)
    anchor_row["crop_image"] = "IMAGE"
    anchor_row["_step_mono_start"] = time.monotonic() - float(anchor_row.get("t_step_start", 0.0) or 0.0)
    anchor_row = maybe_capture_focus_crop(client, dev, anchor_row, output_base_dir)
    anchor_row.pop("_step_mono_start", None)

    rows.append(anchor_row)
    all_rows.append(anchor_row)
    save_excel(all_rows, output_path, with_images=False)

    prev_fingerprint = make_main_fingerprint(anchor_row)
    previous_step_row: dict[str, Any] | None = anchor_row
    fail_count = 0
    same_count = 0
    expanded_overlay_entries: set[str] = set()
    main_step_index_by_fingerprint: dict[tuple[str, str, str], int] = {
        prev_fingerprint: 0,
    }

    for step_idx in range(1, tab_cfg["max_steps"] + 1):
        log(f"[STEP] START tab='{tab_cfg['tab_name']}' step={step_idx}")
        step_start = time.perf_counter()

        row = client.collect_focus_step(
            dev=dev,
            step_index=step_idx,
            move=True,
            direction="next",
            wait_seconds=MAIN_STEP_WAIT_SECONDS,
            announcement_wait_seconds=MAIN_ANNOUNCEMENT_WAIT_SECONDS,
        )
        step_elapsed = time.perf_counter() - step_start

        row["tab_name"] = tab_cfg["tab_name"]
        row["context_type"] = "main"
        row["parent_step_index"] = ""
        row["overlay_entry_label"] = ""
        row["overlay_recovery_status"] = ""
        row["status"] = "OK"
        row["stop_reason"] = ""
        row["step_elapsed_sec"] = round(step_elapsed, 3)
        row["crop_image"] = "IMAGE"
        row["_step_mono_start"] = time.monotonic() - float(row.get("t_step_start", 0.0) or 0.0)
        row = maybe_capture_focus_crop(client, dev, row, output_base_dir)
        row.pop("_step_mono_start", None)
        row["step_total_elapsed_sec"] = round(time.perf_counter() - step_start, 3)

        move_result = str(row.get("move_result", "") or "")
        visible_label = str(row.get("visible_label", "") or "").strip()
        merged_announcement = str(row.get("merged_announcement", "") or "").strip()

        log(
            f"[STEP] END tab='{tab_cfg['tab_name']}' step={step_idx} "
            f"elapsed={step_elapsed:.2f}s move_result='{move_result}' "
            f"visible='{visible_label}' speech='{merged_announcement}' "
            f"crop='{row.get('crop_image_path', '')}' "
            f"timing(move={row.get('move_elapsed_sec', 0):.3f}s "
            f"ann={row.get('announcement_elapsed_sec', 0):.3f}s "
            f"get_focus={row.get('get_focus_elapsed_sec', 0):.3f}s "
            f"get_focus_fallback_dump={row.get('get_focus_fallback_dump_elapsed_sec', 0):.3f}s "
            f"step_dump={row.get('step_dump_tree_elapsed_sec', 0):.3f}s "
            f"crop={row.get('crop_elapsed_sec', 0):.3f}s total={row.get('step_total_elapsed_sec', 0):.3f}s) "
            f"focus_reason='{row.get('get_focus_empty_reason', '')}' "
            f"fallback_used={row.get('get_focus_fallback_used', False)} "
            f"fallback_found={row.get('get_focus_fallback_found', False)} "
            f"step_dump_used={row.get('step_dump_tree_used', False)} "
            f"step_dump_reason='{row.get('step_dump_tree_reason', '')}' "
            f"req_id='{row.get('get_focus_req_id', '')}'"
        )
        mismatch_reasons, low_confidence_reasons = detect_step_mismatch(row=row, previous_step=previous_step_row)
        if mismatch_reasons:
            log(
                f"[MISMATCH] step={step_idx} tab='{tab_cfg['tab_name']}' "
                f"reason='{','.join(mismatch_reasons)}' "
                f"speech='{merged_announcement}' visible='{visible_label}' "
                f"focus_bounds='{row.get('focus_bounds', '')}' source='{row.get('focus_payload_source', '')}'"
            )
        elif low_confidence_reasons:
            log(
                f"[LOW_CONFIDENCE] step={step_idx} tab='{tab_cfg['tab_name']}' "
                f"reason='{','.join(low_confidence_reasons)}' "
                f"speech='{merged_announcement}' visible='{visible_label}' "
                f"focus_bounds='{row.get('focus_bounds', '')}' source='{row.get('focus_payload_source', '')}'"
            )
        elif _should_log("DEBUG"):
            log(
                f"[DEBUG][diag] step={step_idx} speech_count={row.get('announcement_count', 0)} "
                f"window={row.get('announcement_window_sec', 0)} "
                f"focus_source='{row.get('focus_payload_source', '')}' "
                f"response_success={row.get('get_focus_response_success', False)} "
                f"t(after_move={row.get('t_after_move', 0)} "
                f"after_ann={row.get('t_after_ann', 0)} "
                f"after_focus={row.get('t_after_get_focus', 0)} "
                f"before_crop={row.get('t_before_crop', 0)} after_crop={row.get('t_after_crop', 0)})",
                level="DEBUG",
            )

        stop, fail_count, same_count, reason, prev_fingerprint = should_stop(
            row=row,
            prev_fingerprint=prev_fingerprint,
            fail_count=fail_count,
            same_count=same_count,
        )

        if stop:
            row["status"] = "END"
            row["stop_reason"] = reason

        rows.append(row)
        all_rows.append(row)
        row_fingerprint = make_main_fingerprint(row)
        if all(row_fingerprint):
            main_step_index_by_fingerprint[row_fingerprint] = step_idx
        if stop or (step_idx % CHECKPOINT_SAVE_EVERY_STEPS == 0):
            save_excel(all_rows, output_path, with_images=False)

        is_candidate, candidate_reason = is_overlay_candidate(row, tab_cfg)
        if is_candidate:
            fingerprint = make_overlay_entry_fingerprint(tab_cfg["tab_name"], row)
            if fingerprint not in expanded_overlay_entries:
                log(
                    f"[OVERLAY] candidate matched scenario='{tab_cfg.get('scenario_id', '')}' "
                    f"tab='{tab_cfg.get('tab_name', '')}' step={row.get('step_index')} "
                    f"view_id='{row.get('focus_view_id', '')}' label='{row.get('visible_label', '')}' "
                    f"reason='{candidate_reason}'"
                )
                clicked = False
                row_view_id = str(row.get("focus_view_id", "") or "").strip()
                row_label = str(row.get("visible_label", "") or "").strip()
                if row_view_id:
                    clicked = client.touch(
                        dev=dev,
                        name=f"^{re.escape(row_view_id)}$",
                        type_="r",
                        wait_=3,
                    )
                elif row_label:
                    clicked = client.touch(
                        dev=dev,
                        name=f"^{re.escape(row_label)}$",
                        type_="a",
                        wait_=3,
                    )
                if not clicked:
                    log(
                        f"[OVERLAY] post_click classification='unchanged' scenario='{tab_cfg.get('scenario_id', '')}' "
                        f"tab='{tab_cfg.get('tab_name', '')}' step={row.get('step_index')} "
                        f"view_id='{row_view_id}' label='{row_label}' reason='entry_click_failed'"
                    )
                else:
                    time.sleep(0.8)
                    classification, post_click_step = classify_post_click_result(
                        client=client,
                        dev=dev,
                        tab_cfg=tab_cfg,
                        pre_click_step=row,
                    )
                    log(
                        f"[OVERLAY] post_click classification='{classification}' "
                        f"scenario='{tab_cfg.get('scenario_id', '')}' tab='{tab_cfg.get('tab_name', '')}' "
                        f"entry_view_id='{row_view_id}' entry_label='{row_label}' "
                        f"post_view_id='{post_click_step.get('focus_view_id', '')}' "
                        f"post_label='{post_click_step.get('visible_label', '')}'"
                    )

                    if classification == "overlay":
                        expand_overlay(
                            client=client,
                            dev=dev,
                            tab_cfg=tab_cfg,
                            entry_step=row,
                            rows=rows,
                            all_rows=all_rows,
                            output_path=output_path,
                            output_base_dir=output_base_dir,
                            skip_entry_click=True,
                        )
                        expanded_overlay_entries.add(fingerprint)

                        realign_result = realign_focus_after_overlay(
                            client=client,
                            dev=dev,
                            entry_step=row,
                            known_step_index_by_fingerprint=main_step_index_by_fingerprint,
                        )
                        log(
                            f"[OVERLAY] realign status='{realign_result.get('status')}' "
                            f"entry_reached={realign_result.get('entry_reached')} "
                            f"steps_taken={realign_result.get('steps_taken')} "
                            f"match_by='{realign_result.get('match_by', '')}'"
                        )
                        if realign_result.get("entry_reached"):
                            post_overlay_stabilized = stabilize_anchor(
                                client=client,
                                dev=dev,
                                tab_cfg=tab_cfg,
                                phase="overlay_realign",
                                max_retries=2,
                                verify_reads=1,
                            )
                            if not post_overlay_stabilized.get("ok"):
                                log(
                                    f"[ANCHOR][overlay_realign] stabilization failed "
                                    f"tab='{tab_cfg.get('tab_name', '')}'"
                                )
                    elif classification == "navigation":
                        log(
                            f"[OVERLAY] overlay routine skipped (navigation) scenario='{tab_cfg.get('scenario_id', '')}' "
                            f"step={row.get('step_index')}"
                        )
                    else:
                        log(
                            f"[OVERLAY] overlay routine skipped (unchanged) scenario='{tab_cfg.get('scenario_id', '')}' "
                            f"step={row.get('step_index')}"
                        )
            else:
                log(f"[OVERLAY] skip already expanded entry fingerprint='{fingerprint}'")
        elif "blocked" in candidate_reason:
            log(
                f"[OVERLAY] blocked by scenario policy scenario='{tab_cfg.get('scenario_id', '')}' "
                f"tab='{tab_cfg.get('tab_name', '')}' step={row.get('step_index')} "
                f"view_id='{row.get('focus_view_id', '')}' label='{row.get('visible_label', '')}'"
            )

        if stop:
            log(f"[INFO] stop tab={tab_cfg['tab_name']} step={step_idx} reason={reason}")
            break
        previous_step_row = row

    return rows


def main():
    log(f"[MAIN] script start (version={SCRIPT_VERSION}, log_level={LOG_LEVEL})")
    client = A11yAdbClient(dev_serial=DEV_SERIAL)

    all_rows: list[dict] = []
    output_path = generate_output_path()
    output_base_dir = str(Path(output_path).with_suffix(""))

    log(f"[MAIN] output file: {output_path}")
    log(f"[MAIN] image dir base: {output_base_dir}")

    try:
        for tab_cfg in TAB_CONFIGS:
            if not bool(tab_cfg.get("enabled", True)):
                log(
                    f"[MAIN] skip disabled scenario_id='{tab_cfg.get('scenario_id', '')}' "
                    f"tab='{tab_cfg.get('tab_name', '')}'"
                )
                continue
            collect_tab_rows(
                client,
                DEV_SERIAL,
                tab_cfg,
                all_rows,
                output_path,
                output_base_dir,
            )

    except Exception as exc:
        log(f"[FATAL] script interrupted: {exc}")
        save_excel(all_rows, output_path, with_images=False)
        raise

    finally:
        save_excel(all_rows, output_path, with_images=True)
        log("[MAIN] final save complete")

    log("[MAIN] script end")


if __name__ == "__main__":
    main()
    
