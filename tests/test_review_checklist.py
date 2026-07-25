from pathlib import Path
import json

import openpyxl
from PIL import Image
from openpyxl.worksheet.datavalidation import DataValidation

from tools.review_checklist import generate_review_checklist, generate_review_summary
from tools.review_checklist.qa_context import _resource_label


QA_HEADERS = [
    "Review ID", "Scenario", "Focus Target", "Approximate Position", "Review Description",
    "Screenshot", "Speech Status", "Visible Text", "Validator Checklist", "Validator Comment",
    "Step",
]


def _write_source_run(run_root: Path) -> Path:
    run_root.mkdir()
    source = run_root / "talkback_compare_20260724_124302.xlsx"
    workbook = openpyxl.Workbook()
    result = workbook.active
    result.title = "result"
    result.append(
        [
            "plugin_group",
            "plugin_name",
            "scenario_id",
            "step",
            "context_type",
            "visible_label",
            "merged_announcement",
            "representative_visible",
            "mismatch_type",
            "final_result",
            "failure_reason",
            "review_note",
            "focus_view_id",
        ]
    )
    result.append(["Life", "Food", "life_food_plugin", 1, "main", "", "", "", "EMPTY_VISIBLE", "FAIL", "speech_visible_diverged", "검토", "view-1"])
    result.append(["Life", "Food", "life_food_plugin", 2, "main", "Visible", "Speech", "Expected", "EXACT_MATCH", "PASS", "", "", "view-2"])
    result.append(["Life", "Food", "life_food_plugin", 3, "main", "Visible", "Speech", "Expected", "EXACT_MATCH", "WARN", "move_failed", "", "view-3"])
    raw = workbook.create_sheet("raw")
    raw.append(["scenario_id", "step_index", "focus_bounds", "crop_image_path", "crop_image_saved", "actual_focus_resource_id", "focus_node"])
    raw.append(["life_food_plugin", 1, "1,2,3,4", "missing.png", False, "", ""])
    workbook.create_sheet("filtered")
    workbook.create_sheet("summary")
    workbook.save(source)
    (run_root / "summary.json").write_text(json.dumps({"batch_id": "batch_test", "serial": "serial-1", "model": "SM-Test", "xlsx_path": str(source)}), encoding="utf-8")
    return source


def test_generate_review_checklist_selects_failures_and_preserves_missing_evidence(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")

    output = generate_review_checklist(source)

    workbook = openpyxl.load_workbook(output, data_only=False)
    checklist = workbook["Review Checklist"]
    assert checklist.max_row == 2
    assert [cell.value for cell in checklist[1]] == QA_HEADERS
    assert checklist.column_dimensions["K"].hidden is True
    assert checklist.cell(2, 2).value == "Food"
    assert checklist.cell(2, 6).value == "Full screenshot not captured for this run"
    assert checklist.cell(2, 7).value == "Unknown"
    assert checklist.cell(2, 8).value == "화면 텍스트 없음\n(실제 화면 표시 여부 확인)"
    assert any(isinstance(item, DataValidation) for item in checklist.data_validations.dataValidation)
    assert "실제 접근성 문제" in checklist.data_validations.dataValidation[0].formula1
    assert workbook["Additional Review"].max_row == 2
    assert workbook["Summary"].cell(3, 2).value == "batch_test"
    summary_values = {row[0]: row[1] for row in workbook["Summary"].iter_rows(min_row=2, values_only=True) if row[0]}
    assert summary_values["QA 예상 검토 시간"] == "약 2분"
    assert summary_values["Unknown Target count"] == 0
    assert summary_values["Resource-derived Target count"] == 1
    assert workbook.calculation.fullCalcOnLoad is True


def test_review_generation_is_deterministic_and_does_not_overwrite_reviewed_file(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    first = generate_review_checklist(source)
    first_values = list(openpyxl.load_workbook(first, data_only=False)["Review Checklist"].values)
    reviewed = source.with_name(f"{source.stem}.reviewed.xlsx")
    reviewed.write_bytes(b"reviewed")

    second = generate_review_checklist(source)

    assert first == second
    assert list(openpyxl.load_workbook(second, data_only=False)["Review Checklist"].values) == first_values
    assert reviewed.read_bytes() == b"reviewed"


def test_force_regenerate_replaces_only_unreviewed_generated_file(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    output = generate_review_checklist(source)
    workbook = openpyxl.load_workbook(output)
    workbook["Review Checklist"]["B2"] = "stale"
    workbook.save(output)

    generate_review_checklist(source, force_regenerate=True)

    regenerated = openpyxl.load_workbook(output)["Review Checklist"]
    assert regenerated["B2"].value == "Food"
    regenerated["I2"] = "정상 발화"
    regenerated.parent.save(output)
    generate_review_checklist(source, force_regenerate=True)
    assert openpyxl.load_workbook(output)["Review Checklist"]["I2"].value == "정상 발화"


def test_generate_review_summary_uses_detail_decisions(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    detail = generate_review_checklist(source)
    workbook = openpyxl.load_workbook(detail)
    workbook["Review Checklist"]["I2"] = "실제 접근성 문제"
    workbook.save(detail)

    output = generate_review_summary([detail])
    summary = openpyxl.load_workbook(output, data_only=True)["Summary"]
    assert summary.cell(2, 11).value == 1
    assert summary.cell(2, 15).value == 1


def test_review_checklist_prioritizes_human_focus_target_and_crop_screenshot(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    crop = source.parent / "crop.png"
    Image.new("RGB", (200, 100), "white").save(crop)
    (source.parent / f"{source.stem}.evidence.jsonl").write_text("{}\n", encoding="utf-8")
    workbook = openpyxl.load_workbook(source)
    raw = workbook["raw"]
    raw.append(["life_food_plugin", 1, "800,0,1000,200", "crop.png", True, "resource-1"])
    raw_headers = [cell.value for cell in raw[1]]
    raw.cell(raw.max_row, raw_headers.index("focus_node") + 1, json.dumps({"contentDescription": "Settings"}))
    workbook.save(source)
    (source.parent / f"{source.stem}.environment_profile.json").write_text(
        json.dumps({"display": {"logical_size": {"value": {"width": 1000, "height": 1000}}}}),
        encoding="utf-8",
    )

    output = generate_review_checklist(source)

    checklist = openpyxl.load_workbook(output, data_only=False)["Review Checklist"]
    headers = [cell.value for cell in checklist[1]]
    values = {header: checklist.cell(2, index + 1).value for index, header in enumerate(headers)}
    assert values["Focus Target"] == "Settings"
    assert values["Approximate Position"] == "Top Right (90%, 10%)"
    assert values["Screenshot"] == "Crop only (legacy run)\nOpen original screenshot"
    assert (source.parent / "review_annotations" / "crop.focus.png").is_file()
    assert "Settings" in values["Review Description"]


def test_review_checklist_uses_ocr_only_after_text_sources_are_empty(tmp_path: Path, monkeypatch) -> None:
    source = _write_source_run(tmp_path / "device_run")
    crop = source.parent / "crop.png"
    crop.write_bytes(b"crop")
    workbook = openpyxl.load_workbook(source)
    raw = workbook["raw"]
    raw.cell(2, 4, "crop.png")
    raw.cell(2, 5, True)
    workbook["result"].cell(2, 13, "")
    workbook.save(source)
    monkeypatch.setattr("tools.review_checklist.qa_context.extract_crop_text", lambda _path: "Low Battery")

    output = generate_review_checklist(source)

    checklist = openpyxl.load_workbook(output, data_only=False)["Review Checklist"]
    headers = [cell.value for cell in checklist[1]]
    values = {header: checklist.cell(2, index + 1).value for index, header in enumerate(headers)}
    assert values["Focus Target"] == "Low Battery"
    assert values["Speech Status"] == "Unknown"


def test_review_checklist_uses_position_when_screenshot_is_missing(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    workbook = openpyxl.load_workbook(source)
    raw = workbook["raw"]
    raw.cell(2, 3, "800,0,1000,200")
    workbook.save(source)
    (source.parent / f"{source.stem}.environment_profile.json").write_text(
        json.dumps({"display": {"logical_size": {"value": {"width": 1000, "height": 1000}}}}),
        encoding="utf-8",
    )

    output = generate_review_checklist(source)

    checklist = openpyxl.load_workbook(output, data_only=False)["Review Checklist"]
    headers = [cell.value for cell in checklist[1]]
    values = {header: checklist.cell(2, index + 1).value for index, header in enumerate(headers)}
    assert values["Screenshot"] == "Full screenshot not captured for this run"
    assert "Top Right" in values["Review Description"]


def test_review_checklist_uses_resource_id_before_unknown_target(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    workbook = openpyxl.load_workbook(source)
    workbook["result"].cell(2, 13, "lowBattery")
    workbook.save(source)

    output = generate_review_checklist(source)

    checklist = openpyxl.load_workbook(output, data_only=False)["Review Checklist"]
    assert checklist["C2"].value == "Low Battery"


def test_review_checklist_normalizes_resource_ids_for_qa(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    workbook = openpyxl.load_workbook(source)
    workbook["result"].cell(2, 13, "com.example:id/shm_setting_button")
    workbook.save(source)

    output = generate_review_checklist(source)

    checklist = openpyxl.load_workbook(output, data_only=False)["Review Checklist"]
    assert checklist["C2"].value == "Settings Button"
    assert "Settings Button" in checklist["E2"].value
    assert "Shm" not in checklist["E2"].value
    summary = {row[0]: row[1] for row in openpyxl.load_workbook(output, data_only=False)["Summary"].iter_rows(min_row=2, values_only=True) if row[0]}
    assert summary["Resource-derived Target count"] == 1
    assert _resource_label("home_monitor_setting") == "Home Monitor Settings"
    assert _resource_label("btn_menu") == "Menu Button"
    assert _resource_label("sensor_card") == "Sensor Card"


def test_review_checklist_separates_automation_diagnostics_from_manual_review(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    workbook = openpyxl.load_workbook(source)
    workbook["result"].append(["Life", "Food", "life_food_plugin", 4, "main", "", "", "", "EMPTY_VISIBLE", "FAIL", "terminal_not_handled", "", "view-4"])
    workbook.save(source)

    output = generate_review_checklist(source)

    workbook = openpyxl.load_workbook(output, data_only=False)
    assert workbook["Review Checklist"].max_row == 2
    diagnostic = workbook["Automation Diagnostic"]
    assert diagnostic.max_row == 2
    assert diagnostic.cell(2, 2).value == "terminal_not_handled"
    summary_values = {row[0]: row[1] for row in workbook["Summary"].iter_rows(min_row=2, values_only=True) if row[0]}
    assert summary_values["QA Review Count"] == 1
    assert summary_values["Automation Diagnostic Count"] == 1


def test_review_checklist_distinguishes_speech_statuses_from_evidence_provenance(tmp_path: Path) -> None:
    source = _write_source_run(tmp_path / "device_run")
    workbook = openpyxl.load_workbook(source)
    result = workbook["result"]
    result.append(["Devices", "Water Leak Sensor", "device_water_leak_sensor_plugin", 2, "main", "", "", "", "EMPTY_VISIBLE", "FAIL", "speech_visible_diverged", "", "lowBattery"])
    result.append(["Devices", "Motion Sensor", "device_motion_sensor_plugin", 2, "main", "", "", "", "EMPTY_VISIBLE", "FAIL", "speech_visible_diverged", "", "lowBattery"])
    result.append(["Life", "Home Monitor", "life_home_monitor_plugin", 1, "main", "", "", "", "EMPTY_SPEECH", "FAIL", "speech_visible_diverged", "", "shm_setting_button"])
    result.append(["Life", "Observed", "life_observed_plugin", 1, "main", "Settings", "Settings", "", "SPEECH_MISMATCH", "FAIL", "speech_visible_diverged", "", "settings"])
    result.append(["Life", "Missing", "life_missing_plugin", 1, "main", "", "", "", "EMPTY_SPEECH", "FAIL", "speech_visible_diverged", "", "missing_label"])
    workbook.save(source)
    workbook = openpyxl.load_workbook(source)
    raw = workbook["raw"]
    raw.cell(1, 8, "focus_class_name")
    raw.append(["life_home_monitor_plugin", 1, "", "", False, "shm_setting_button", "", "android.widget.Button"])
    workbook.save(source)
    evidence = [
        {"scenario_id": "device_water_leak_sensor_plugin", "step_index": 2, "event_type": "POST_FOCUS_OBSERVED", "payload": {"observation": {"resource_id": "lowBattery"}}},
        {"scenario_id": "device_motion_sensor_plugin", "step_index": 2, "event_type": "POST_FOCUS_OBSERVED", "payload": {"observation": {"resource_id": "lowBattery"}}},
        {"scenario_id": "life_home_monitor_plugin", "step_index": 1, "event_type": "ACCESSIBILITY_FOCUS_EVENT", "payload": {"focus": {"resource_id": "shm_setting_button"}}},
        {"scenario_id": "life_missing_plugin", "step_index": 1, "event_type": "ACCESSIBILITY_FOCUS_EVENT", "payload": {"focus": {"resource_id": "missing_label"}}},
    ]
    (source.parent / f"{source.stem}.evidence.jsonl").write_text("\n".join(json.dumps(item) for item in evidence), encoding="utf-8")

    output = generate_review_checklist(source)

    checklist = openpyxl.load_workbook(output, data_only=False)["Review Checklist"]
    headers = [cell.value for cell in checklist[1]]
    statuses = {checklist.cell(row, headers.index("Scenario") + 1).value: checklist.cell(row, headers.index("Speech Status") + 1).value for row in range(2, checklist.max_row + 1)}
    assert statuses["Water Leak Sensor"] == "Speech Unobserved"
    assert statuses["Motion Sensor"] == "Speech Unobserved"
    assert statuses["Home Monitor"] == "Role-only Speech"
    assert statuses["Observed"] == "Speech Observed"
    assert statuses["Missing"] == "Speech Missing"
    assert statuses["Food"] == "Unknown"
    summary = {row[0]: row[1] for row in openpyxl.load_workbook(output, data_only=False)["Summary"].iter_rows(min_row=2, values_only=True) if row[0]}
    assert summary["Speech Unobserved count"] == 2
    assert summary["Role-only Speech count"] == 1
    assert summary["Speech Observed count"] == 1
    assert summary["Speech Missing count"] == 1
    assert summary["Unknown Speech Status count"] == 1
