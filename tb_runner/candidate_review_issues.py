from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any, Mapping, assert_never

import openpyxl

from tb_runner.review_classification import (
    ReviewClassification,
    ReviewDomain,
    classify_review_domain,
)
from tb_runner.review_signature import (
    review_source_signature,
    review_source_signature_digest,
)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _workbook_path(run_root: Path, summary: Mapping[str, Any]) -> Path | None:
    configured = _text(summary.get("xlsx_path"))
    configured_path = Path(configured) if configured else None
    candidates = []
    if configured_path is not None:
        candidates.extend(
            [
                configured_path,
                run_root / configured_path.name,
            ]
        )
    candidates.extend(
        path
        for path in sorted(run_root.glob("talkback_compare*.xlsx"))
        if ".review." not in path.name
    )
    root = run_root.resolve()
    for candidate in candidates:
        path = candidate if candidate.is_absolute() else candidate
        if not path.is_file():
            path = run_root / candidate
        resolved = path.resolve()
        if resolved.is_file() and resolved.is_relative_to(root):
            return resolved
    return None


def _result_rows(path: Path | None) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    indexed: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    if path is None:
        return indexed
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "result" not in workbook.sheetnames:
            return indexed
        sheet = workbook["result"]
        values = sheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(values, ())]
        for row_number, values_row in enumerate(values, 2):
            row = dict(zip(headers, values_row))
            key = (
                _text(row.get("scenario_id")),
                _text(row.get("step")),
                _text(row.get("mismatch_type")).upper(),
            )
            row["source_row"] = row_number
            indexed[key].append(row)
    finally:
        workbook.close()
    return indexed


def _evidence_transactions(path: Path | None) -> dict[tuple[str, str, str], str]:
    indexed: dict[tuple[str, str, str], str] = {}
    if path is None or not path.is_file():
        return indexed
    with path.open(encoding="utf-8") as stream:
        for line in stream:
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                continue
            if event.get("event_type") != "POST_FOCUS_OBSERVED":
                continue
            payload = event.get("payload")
            payload = payload if isinstance(payload, Mapping) else {}
            observation = payload.get("observation")
            observation = observation if isinstance(observation, Mapping) else {}
            key = (
                _text(event.get("scenario_id")),
                _text(event.get("step_index")),
                _text(observation.get("resource_id")),
            )
            transaction_id = _text(event.get("transaction_id"))
            if transaction_id:
                indexed.setdefault(key, transaction_id)
    return indexed


def _evidence_reference(
    run_root: Path,
    evidence_path: Path | None,
    transaction_id: str,
) -> str:
    if evidence_path is None or not transaction_id:
        return ""
    return (
        f"qa-run://{run_root.parent.name}/{run_root.name}/{evidence_path.name}"
        f"#transaction={transaction_id}"
    )


def _source_row(
    item: Mapping[str, Any], source_matches: list[dict[str, Any]]
) -> dict[str, Any]:
    """Select the workbook row that best corresponds to a summary signal."""
    if not source_matches:
        return {}
    failure_reason = _text(item.get("failure_reason")).lower()
    resource_id = _text(item.get("resource_id") or item.get("focus_view_id"))
    for source in source_matches:
        source_failure = _text(source.get("failure_reason")).lower()
        source_resource = _text(source.get("focus_view_id"))
        if failure_reason and source_failure and source_failure != failure_reason:
            continue
        if resource_id and source_resource and source_resource != resource_id:
            continue
        return source
    return source_matches[0]


def _sort_component(value: Any) -> tuple[int, int | str]:
    value_text = _text(value)
    if not value_text:
        return (0, "")
    if value_text.isdigit():
        return (1, int(value_text))
    return (2, value_text)


def _issue_sort_key(item: Mapping[str, Any]) -> tuple[Any, ...]:
    signature = item.get("source_signature")
    signature = signature if isinstance(signature, Mapping) else {}
    return (
        _text(item.get("scenario_id")),
        _sort_component(item.get("step") or signature.get("step")),
        _text(item.get("mismatch_type") or item.get("code")).upper(),
        _text(item.get("failure_reason")).lower(),
        _text(item.get("resource_id") or signature.get("resource_id")),
        _text(item.get("source_transaction_id") or signature.get("transaction_id")),
        _sort_component(signature.get("source_row")),
        _text(item.get("source_signature_digest")),
    )


def _malformed_automation_diagnostic() -> dict[str, Any]:
    signature = review_source_signature(
        scenario_id="",
        step="",
        mismatch_type="",
        resource_id="",
        transaction_id="",
        source_row=None,
    )
    return {
        "code": "MALFORMED_AUTOMATION_DIAGNOSTIC",
        "mismatch_type": "",
        "category": "AUTOMATION_DIAGNOSTIC",
        "review_domain": ReviewDomain.UNKNOWN.value,
        "scenario_id": "",
        "step": "",
        "raw_result": None,
        "failure_reason": "malformed_automation_diagnostic",
        "classification_reason": "malformed_automation_diagnostic",
        "resource_id": "",
        "source_transaction_id": "",
        "evidence_reference": "",
        "source_row": None,
        "source_signature": signature,
        "source_signature_digest": review_source_signature_digest(signature),
        "review_status": "UNREVIEWED",
    }


def classify_candidate_review_issues(
    run_root: Path,
    summary: Mapping[str, Any],
) -> tuple[
    tuple[dict[str, Any], ...], tuple[dict[str, Any], ...], tuple[dict[str, Any], ...]
]:
    workbook_path = _workbook_path(run_root, summary)
    result_rows = _result_rows(workbook_path)
    evidence_path = (
        workbook_path.with_suffix(".evidence.jsonl")
        if workbook_path is not None
        else None
    )
    evidence_transactions = _evidence_transactions(evidence_path)
    qa: list[dict[str, Any]] = []
    automation: list[dict[str, Any]] = []
    unknown: list[dict[str, Any]] = []
    projected_signatures: dict[str, ReviewDomain] = {}
    source_items: list[tuple[Any, ReviewDomain | None]] = []
    issues = summary.get("quality_issues")
    if isinstance(issues, list):
        source_items.extend((item, None) for item in issues)
    # Newer summaries persist the already classified automation collection
    # separately.  Keep quality_issues as a legacy fallback because older
    # summaries carried both review domains in that collection.
    diagnostics = summary.get("automation_diagnostics")
    if isinstance(diagnostics, list):
        source_items.extend(
            (item, ReviewDomain.AUTOMATION_ENGINE) for item in diagnostics
        )

    for item, expected_domain in source_items:
        if not isinstance(item, Mapping):
            if expected_domain is ReviewDomain.AUTOMATION_ENGINE:
                malformed = _malformed_automation_diagnostic()
                digest = malformed["source_signature_digest"]
                if projected_signatures.get(digest) is None:
                    projected_signatures[digest] = ReviewDomain.UNKNOWN
                    unknown.append(malformed)
            continue
        mismatch_type = _text(item.get("mismatch_type")).upper()
        if not mismatch_type:
            if expected_domain is not ReviewDomain.AUTOMATION_ENGINE:
                continue
            classification = ReviewClassification(
                ReviewDomain.UNKNOWN,
                "malformed_automation_diagnostic",
                _text(item.get("failure_reason")).lower(),
                mismatch_type,
            )
        else:
            classification = None
        scenario_id = _text(item.get("scenario_id"))
        step = _text(item.get("step"))
        source_matches = result_rows.get((scenario_id, step, mismatch_type), [])
        source = _source_row(item, source_matches)
        failure_reason = _text(
            item.get("failure_reason") or source.get("failure_reason")
        )
        resource_id = _text(
            item.get("resource_id")
            or item.get("focus_view_id")
            or source.get("focus_view_id")
        )
        transaction_id = _text(item.get("transaction_id")) or evidence_transactions.get(
            (scenario_id, step, resource_id),
            "",
        )
        evidence_reference = _text(
            item.get("evidence_reference")
        ) or _evidence_reference(
            run_root,
            evidence_path,
            transaction_id,
        )
        if classification is None:
            classification = classify_review_domain(
                mismatch_type=mismatch_type,
                failure_reason=failure_reason,
            )
        if (
            expected_domain is ReviewDomain.AUTOMATION_ENGINE
            and classification.review_domain is not ReviewDomain.AUTOMATION_ENGINE
        ):
            classification = ReviewClassification(
                ReviewDomain.UNKNOWN,
                "automation_diagnostic_domain_mismatch",
                failure_reason.lower(),
                mismatch_type,
            )
        code = mismatch_type or (
            "MALFORMED_AUTOMATION_DIAGNOSTIC"
            if expected_domain is ReviewDomain.AUTOMATION_ENGINE
            else ""
        )
        signature = review_source_signature(
            scenario_id=scenario_id,
            step=step,
            mismatch_type=mismatch_type,
            resource_id=resource_id,
            transaction_id=transaction_id,
            source_row=source.get("source_row"),
        )
        provenance = {
            "code": code,
            "mismatch_type": mismatch_type,
            "category": (
                "OBSERVED_RESULT_LIMITATION"
                if classification.review_domain is ReviewDomain.QA_ACCESSIBILITY
                else "AUTOMATION_DIAGNOSTIC"
            ),
            "review_domain": classification.review_domain.value,
            "scenario_id": scenario_id,
            "step": step,
            "raw_result": (
                item.get("final_result")
                or item.get("raw_final_result")
                or source.get("final_result")
            ),
            "failure_reason": failure_reason,
            "classification_reason": classification.classification_reason,
            "resource_id": resource_id,
            "source_transaction_id": transaction_id,
            "evidence_reference": evidence_reference,
            "source_row": source.get("source_row"),
            "source_signature": signature,
            "source_signature_digest": review_source_signature_digest(signature),
        }
        digest = provenance["source_signature_digest"]
        previous_domain = projected_signatures.get(digest)
        if previous_domain is classification.review_domain:
            continue
        projected_signatures[digest] = classification.review_domain
        match classification.review_domain:
            case ReviewDomain.QA_ACCESSIBILITY:
                qa.append({**provenance, "review_status": "UNREVIEWED"})
            case ReviewDomain.AUTOMATION_ENGINE:
                automation.append(
                    {**provenance, "acknowledgment_status": "UNACKNOWLEDGED"}
                )
            case ReviewDomain.UNKNOWN:
                unknown.append({**provenance, "review_status": "UNREVIEWED"})
            case _ as unreachable:
                assert_never(unreachable)
    qa.sort(key=_issue_sort_key)
    automation.sort(key=_issue_sort_key)
    unknown.sort(key=_issue_sort_key)
    return tuple(qa), tuple(automation), tuple(unknown)


__all__ = ["classify_candidate_review_issues"]
