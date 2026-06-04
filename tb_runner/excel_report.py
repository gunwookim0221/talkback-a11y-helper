import os
from io import BytesIO
from pathlib import Path
import json
import re

import pandas as pd

from tb_runner.diagnostics import is_placeholder_row, normalize_move_result
from tb_runner.image_utils import create_excel_thumbnail, insert_images_to_excel
from tb_runner.logging_utils import get_recent_logs, log
from tb_runner.utils import to_json_text

EXCEL_REPORT_VERSION = "1.4.4"

_DEBUG_LOG_KEYWORDS = (
    "[STEP]",
    "[ROW]",
    "[STOP]",
    "[MISMATCH]",
    "[LOW_CONFIDENCE]",
    "get_focus",
    "announcement",
    "mismatch",
    "req_id",
    "fingerprint",
)
_DEBUG_LOG_FAILURE_REASONS = {
    "repeat_no_progress",
    "terminal_not_handled",
    "move_failed",
    "speech_visible_diverged",
}
_ANN_REQUIRED_TAGS = ("[ANN][baseline]", "[ANN][poll]", "[ANN][stable]", "[ANN][select]")

RESULT_SHEET_COLUMNS = [
    "plugin_group",
    "plugin_name",
    "scenario_id",
    "step",
    "context_type",
    "visible_label",
    "merged_announcement",
    "representative_visible",
    "mismatch_type",
    "final_result",
    "failure_reason",
    "review_note",
    "focus_view_id",
    "focus_confidence",
    "result_crop_thumbnail",
]

RESULT_REPEAT_METADATA_COLUMNS = [
    "repeat_count",
    "first_step",
    "last_step",
    "steps",
    "is_repeated_issue_group",
]

_REPEATED_ISSUE_FAILURE_REASONS = {
    "repeat_no_progress",
    "move_failed",
    "focus_realign_fail",
}

PLUGIN_REPORT_METADATA: dict[str, dict[str, str]] = {
    "global_nav_main": {"group": "Global", "name": "Global Navigation"},
    "home_main": {"group": "Global", "name": "Home"},
    "devices_main": {"group": "Global", "name": "Devices"},
    "life_main": {"group": "Global", "name": "Life"},
    "routines_main": {"group": "Global", "name": "Routines"},
    "menu_main": {"group": "Global", "name": "Menu"},
    "settings_entry_example": {"group": "Settings", "name": "Settings"},
    "life_food_plugin": {"group": "Life", "name": "Food"},
    "life_air_care_plugin": {"group": "Life", "name": "Air Care"},
    "life_home_care_plugin": {"group": "Life", "name": "Home Care"},
    "life_energy_plugin": {"group": "Life", "name": "Energy"},
    "life_pet_care_plugin": {"group": "Life", "name": "Pet Care"},
    "life_family_care_plugin": {"group": "Life", "name": "Family Care"},
    "life_plant_care_plugin": {"group": "Life", "name": "Plant Care"},
    "life_clothing_care_plugin": {"group": "Life", "name": "Clothing Care"},
    "life_find_plugin": {"group": "Life", "name": "Find"},
    "life_video_plugin": {"group": "Life", "name": "Video"},
    "life_home_monitor_plugin": {"group": "Life", "name": "Home Monitor"},
    "life_music_sync_plugin": {"group": "Life", "name": "Music Sync"},
    "device_smoke_sensor_plugin": {"group": "Devices", "name": "연기센서"},
    "device_water_leak_sensor_plugin": {"group": "Devices", "name": "누수센서"},
    "device_motion_sensor_plugin": {"group": "Devices", "name": "모션센서"},
    "device_door_lock_plugin": {"group": "Devices", "name": "도어락"},
    "device_air_purifier_plugin": {"group": "Devices", "name": "공기청정기"},
    "device_humidity_sensor_plugin": {"group": "Devices", "name": "습도센서"},
    "device_temperature_humidity_sensor_plugin": {"group": "Devices", "name": "온습도센서"},
    "device_tv_plugin": {"group": "Devices", "name": "TV"},
    "device_washer_plugin": {"group": "Devices", "name": "세탁기"},
    "device_audio_plugin": {"group": "Devices", "name": "Audio"},
    "device_camera_plugin": {"group": "Devices", "name": "Camera"},
    "device_home_camera_plugin": {"group": "Devices", "name": "홈카메라 360"},
}

_KEY_VALUE_PATTERNS = {
    "req_id": (r"req_id='([^']*)'", r'req_id="([^"]*)"', r"req_id=([^\s,]+)"),
    "tab": (r"tab='([^']*)'", r'tab="([^"]*)"', r"tab=([^\s,]+)", r"tab_name='([^']*)'", r'tab_name="([^"]*)"', r"tab_name=([^\s,]+)"),
    "scenario": (r"scenario='([^']*)'", r'scenario="([^"]*)"', r"scenario=([^\s,]+)", r"scenario_id='([^']*)'", r'scenario_id="([^"]*)"', r"scenario_id=([^\s,]+)"),
    "step": (r"step=([0-9]+)", r"step_index=([0-9]+)"),
}


def _overlay_first_row_debug_enabled() -> bool:
    return str(os.getenv("TB_OVERLAY_FIRST_ROW_DEBUG", "") or "").strip().lower() in {"1", "true", "t", "yes", "y"}


def _normalize_step_value(step_value: object) -> str:
    text = str(step_value or "").strip()
    if not text:
        return ""
    try:
        numeric = float(text)
        if numeric.is_integer():
            return str(int(numeric))
    except (TypeError, ValueError):
        return text
    return text


def _sanitize_filename_part(value: object, default: str) -> str:
    text = str(value or "").strip()
    if not text:
        return default
    return re.sub(r"[^0-9a-zA-Z._-]+", "_", text)[:80] or default


def _display_name_for_path(path_text: object) -> str:
    normalized = str(path_text or "").replace("\\", "/").rstrip("/")
    return os.path.basename(normalized) or str(path_text or "")


def _is_debug_log_target(row) -> bool:
    def _row_get(key: str, default: object = "") -> object:
        if isinstance(row, dict):
            return row.get(key, default)
        if hasattr(row, "get"):
            try:
                return row.get(key, default)
            except Exception:
                pass
        return getattr(row, key, default)

    final_result = str(_row_get("final_result", "") or "").strip().upper()
    failure_reason = str(_row_get("failure_reason", "") or "").strip().lower()
    mismatch_reason = "speech_visible_diverged" in failure_reason
    if final_result in {"WARN", "FAIL"}:
        return True
    if mismatch_reason:
        return True
    if any(reason in failure_reason for reason in _DEBUG_LOG_FAILURE_REASONS):
        return True
    focus_confidence = str(_row_get("focus_confidence", "") or "").strip().upper()
    return final_result == "WARN" and focus_confidence == "LOW"


def _extract_field_value(line: str, field: str) -> str:
    for pattern in _KEY_VALUE_PATTERNS.get(field, ()):
        match = re.search(pattern, line)
        if match:
            return str(match.group(1) or "").strip().strip("'\"")
    return ""


def _line_matches_req_id_scope(line: str, *, req_id: str, tab: str, scenario: str) -> bool:
    req_text = str(req_id or "").strip()
    tab_text = str(tab or "").strip()
    scenario_text = str(scenario or "").strip()

    if req_text:
        line_req = _extract_field_value(line, "req_id")
        if line_req:
            if line_req != req_text:
                return False
        elif req_text not in line:
            return False
    elif tab_text or scenario_text:
        # req_id가 없는 경우에만 보조 스코프 허용
        pass
    else:
        return False

    if tab_text:
        line_tab = _extract_field_value(line, "tab")
        if line_tab and line_tab != tab_text:
            return False
    if scenario_text:
        line_scenario = _extract_field_value(line, "scenario")
        if line_scenario and line_scenario != scenario_text:
            return False
    return True


def _line_matches_step_scope(line: str, *, step_str: str) -> bool:
    normalized_step = _normalize_step_value(step_str)
    if not normalized_step:
        return True
    line_step = _extract_field_value(line, "step")
    if line_step:
        return _normalize_step_value(line_step) == normalized_step
    return True


def _extract_debug_lines_for_row(
    recent_logs: list[str],
    *,
    req_id: str,
    tab: str,
    scenario: str,
    step_str: str,
) -> list[str]:
    extracted: list[str] = []
    for line in recent_logs:
        if not _line_matches_req_id_scope(line, req_id=req_id, tab=tab, scenario=scenario):
            continue
        if not _line_matches_step_scope(line, step_str=step_str):
            continue
        extracted.append(line)
    return extracted


def _build_source_row_index(source_df: pd.DataFrame | None) -> dict[tuple[str, str, str], dict]:
    if source_df is None or source_df.empty:
        return {}
    index: dict[tuple[str, str, str], dict] = {}
    for row in source_df.itertuples(index=False):
        scenario = str(getattr(row, "scenario_id", "") or "").strip()
        tab = str(getattr(row, "tab_name", "") or getattr(row, "tab", "") or "").strip()
        step = _normalize_step_value(getattr(row, "step_index", getattr(row, "step", "")))
        if not scenario or not step:
            continue
        index[(scenario, tab, step)] = row._asdict()
    return index


def _build_debug_log_sections(
    row,
    source_row: dict[str, object] | None,
    prev_source_row: dict[str, object] | None,
) -> dict[str, str]:
    def _row_get(key: str, default: object = "") -> object:
        if isinstance(row, dict):
            return row.get(key, default)
        if hasattr(row, "get"):
            try:
                return row.get(key, default)
            except Exception:
                pass
        return getattr(row, key, default)

    req_id = str(_row_get("_req_id", "") or _row_get("req_id", "") or "").strip()
    scenario_id = str(_row_get("scenario_id", "") or (source_row.get("scenario_id", "") if source_row else "") or "").strip()
    tab_name = str(_row_get("_tab", "") or _row_get("tab", "") or (source_row or {}).get("tab_name", "") or (source_row or {}).get("tab", "")).strip()
    step_str = _normalize_step_value(_row_get("step", "") or (source_row or {}).get("step_index", ""))
    recent_logs = get_recent_logs(limit=260)
    scoped_logs = _extract_debug_lines_for_row(
        recent_logs,
        req_id=req_id,
        tab=tab_name,
        scenario=scenario_id,
        step_str=step_str,
    )

    ann_lines: list[str] = []
    step_lines: list[str] = []
    scroll_lines: list[str] = []
    focus_lines: list[str] = []
    decision_lines: list[str] = []

    for line in scoped_logs:
        lower = line.lower()
        if any(tag in line for tag in _ANN_REQUIRED_TAGS) or "[ANN]" in line:
            ann_lines.append(line)
        if (
            "[STEP] START" in line
            or "[STEP] END" in line
            or "[STOP][eval]" in line
            or "[STOP][triggered]" in line
            or "smart_next" in lower
        ):
            step_lines.append(line)
        if "[END_CHECK][SCROLL]" in line or "[scroll" in lower or " scrolled" in lower or "smart_next" in lower:
            scroll_lines.append(line)
        if "get_focus" in lower or "focus_payload" in lower or "fallback" in lower or "[focus" in lower:
            focus_lines.append(line)

    source = source_row or {}
    prev_source = prev_source_row or {}
    baseline_text = str(prev_source.get("merged_announcement", "") or "")
    baseline_source = "prev_step_row"
    baseline_empty_reason = ""
    if not prev_source:
        baseline_empty_reason = "no_prev_step_row"
        baseline_source = "source_row_missing"
    elif not baseline_text:
        baseline_empty_reason = "prev_step_speech_empty"
    elif not baseline_text.strip():
        baseline_empty_reason = "baseline_not_available"
    if not baseline_empty_reason and not baseline_text:
        baseline_empty_reason = "baseline_not_available"

    current_speech = str(source.get("merged_announcement", "") or _row_get("merged_announcement", "") or _row_get("speech", "") or "")
    partial_ann = source.get("partial_announcements", [])
    ann_count = int(source.get("announcement_count", 0) or 0)
    poll_source = "row_snapshot"
    poll_observed_at = ""
    if ann_lines:
        poll_source = "recent_log"
        first_line = ann_lines[0]
        timestamp_match = re.match(r"\[([0-9]{2}:[0-9]{2}:[0-9]{2})\]", first_line)
        if timestamp_match:
            poll_observed_at = timestamp_match.group(1)
    normalized_candidate = _normalize_text(current_speech)
    normalized_baseline = _normalize_text(baseline_text)
    differs_from_baseline = bool(normalized_candidate and normalized_candidate != normalized_baseline)
    changed = bool(current_speech and current_speech != baseline_text)

    trim_considered = bool(source.get("trim_considered", False))
    trim_applied = bool(source.get("trim_applied", False))
    trim_before = str(source.get("trim_before", current_speech) or current_speech)
    trim_after = str(source.get("trim_after", current_speech) or current_speech)
    trim_reject_reason = str(source.get("trim_reject_reason", "") or "")
    trim_reason = str(source.get("trim_reason", "") or "")
    if not trim_considered and not trim_reason:
        trim_reason = "trim_stage_not_reached"
    if trim_considered and not trim_applied and not trim_reject_reason:
        trim_reject_reason = "no_trim_rule_matched"

    stable_selected = current_speech
    stable_reason = str(source.get("announcement_stable_reason", "") or "result_row_snapshot")
    stable_source = str(source.get("announcement_stable_source", "") or "result_row")
    used_snapshot = stable_reason == "result_row_snapshot"
    snapshot_reason = str(source.get("snapshot_reason", "") or ("no_better_recent_poll_candidate" if used_snapshot else "not_used"))

    ann_summary_lines = [
        f"[ANN][baseline] req_id={req_id} text='{baseline_text}' source='{baseline_source}' empty_reason='{baseline_empty_reason}'",
        f"[ANN][poll] req_id={req_id} idx=0 raw='{current_speech}' normalized='{normalized_candidate}' changed={str(changed).lower()} differs_from_baseline={str(differs_from_baseline).lower()} source='{poll_source}' observed_at='{poll_observed_at}' count={ann_count}",
        f"[ANN][trim] req_id={req_id} considered={str(trim_considered).lower()} applied={str(trim_applied).lower()} before='{trim_before}' after='{trim_after}' reject_reason='{trim_reject_reason}' reason='{trim_reason}'",
        f"[ANN][stable] req_id={req_id} selected='{stable_selected}' reason='{stable_reason}' source='{stable_source}'",
        f"[ANN][select] req_id={req_id} used_snapshot={str(used_snapshot).lower()} snapshot_reason='{snapshot_reason}' previous='{baseline_text}' current='{current_speech}' final='{stable_selected}'",
    ]
    ann_lines = ann_summary_lines + ann_lines

    if not step_lines:
        step_lines.extend(
            [
                f"[STEP] START scenario='{scenario_id}' tab='{tab_name}' step={step_str} req_id='{req_id}'",
                f"[STEP] END scenario='{scenario_id}' tab='{tab_name}' step={step_str} req_id='{req_id}' move_result='{_row_get('_move_result', '') or _row_get('move_result', '')}'",
                f"[STOP][eval] scenario='{scenario_id}' step={step_str} reason='{_row_get('failure_reason', '')}'",
            ]
        )

    smart_result = str(source.get("last_smart_nav_result", "") or "")
    smart_detail = str(source.get("last_smart_nav_detail", "") or "")
    smart_terminal = bool(source.get("last_smart_nav_terminal", False))
    smart_requested_view_id = str(source.get("smart_nav_requested_view_id", "") or "")
    smart_resolved_view_id = str(source.get("smart_nav_resolved_view_id", "") or "")
    smart_actual_view_id = str(source.get("smart_nav_actual_view_id", "") or "")
    post_move_verdict_source = str(source.get("post_move_verdict_source", "") or "")
    if smart_result or smart_detail:
        step_lines.append(
            f"[SMART_NEXT] req_id='{req_id}' tab='{tab_name}' step={step_str} result='{smart_result}' detail='{smart_detail}' terminal={str(smart_terminal).lower()}"
        )
    if smart_requested_view_id or smart_resolved_view_id or smart_actual_view_id or post_move_verdict_source:
        step_lines.append(
            f"[SMART_NEXT][verdict] req_id='{req_id}' requested_view_id='{smart_requested_view_id}' "
            f"resolved_view_id='{smart_resolved_view_id}' actual_view_id='{smart_actual_view_id}' "
            f"post_move_verdict_source='{post_move_verdict_source}'"
        )
    step_lines.append(
        f"[STEP][summary] req_id={req_id} tab='{tab_name}' step={step_str} move_result='{_row_get('_move_result', '') or _row_get('move_result', '')}' terminal={str(smart_terminal).lower()} no_progress={str('repeat_no_progress' in str(_row_get('failure_reason', '') or '')).lower()}"
    )
    if not scroll_lines:
        move_result = str(source.get("move_result", "") or _row_get("_move_result", "") or _row_get("move_result", "") or "")
    move_result = str(source.get("move_result", "") or _row_get("_move_result", "") or _row_get("move_result", "") or "")
    scroll_related_log_found = bool(scroll_lines)
    inferred_scroll = move_result in {"scrolled"} or "scrolled" in smart_detail.lower()
    inference_reason = "no_scroll_signal_found"
    if move_result == "scrolled":
        inference_reason = "move_result_is_scrolled"
    elif "scrolled" in smart_detail.lower():
        inference_reason = "detail_contains_scrolled"
    scroll_lines.append(
        f"[SCROLL] req_id={req_id} scroll_related_log_found={str(scroll_related_log_found).lower()} move_result='{move_result}' smart_result='{smart_result}' detail='{smart_detail}' inferred_scroll={str(inferred_scroll).lower()} inference_reason='{inference_reason}'"
    )

    focus_source = str(source.get("focus_payload_source", "") or source.get("get_focus_final_payload_source", "") or "")
    response_success = bool(source.get("get_focus_response_success", False))
    payload_sufficient = bool(source.get("get_focus_top_level_payload_sufficient", False))
    fallback_used = bool(source.get("get_focus_fallback_used", False))
    fallback_found = bool(source.get("get_focus_fallback_found", False))
    dump_attempted = bool(source.get("get_focus_success_false_top_level_dump_attempted", False))
    final_payload_source = str(source.get("get_focus_final_payload_source", "") or "")
    focus_reason = str(source.get("get_focus_final_focus_reason", "") or "")
    missing_reason = ""
    if not focus_lines and not focus_source and not final_payload_source and not source.get("get_focus_req_id", ""):
        missing_reason = "no_direct_get_focus_trace_found"
        focus_source = "inferred_from_step_end"
    focus_lines.append(
        f"[FOCUS] req_id={req_id} get_focus_req_id='{source.get('get_focus_req_id', '')}' response_success={str(response_success).lower()} payload_source='{focus_source}' payload_sufficient={str(payload_sufficient).lower()} fallback_used={str(fallback_used).lower()} fallback_found={str(fallback_found).lower()} dump_attempted={str(dump_attempted).lower()} final_payload_source='{final_payload_source}' focus_reason='{focus_reason}' missing_reason='{missing_reason}'"
    )
    if isinstance(partial_ann, list) and partial_ann:
        ann_lines.append(f"[ANN][poll] req_id={req_id} partial_announcements={partial_ann[:3]}")

    decision_lines.extend(
        [
            f"baseline_empty={str(not bool(baseline_text)).lower()}",
            f"first_candidate_contaminated={str(differs_from_baseline and bool(current_speech)).lower()}",
            f"trim_considered={str(trim_considered).lower()}",
            f"trim_applied={str(trim_applied).lower()}",
            f"stable_reason='{stable_reason}'",
            f"used_snapshot={str(used_snapshot).lower()}",
            f"snapshot_reason='{snapshot_reason}'",
            f"scroll_evidence='{'strong' if scroll_related_log_found else 'weak'}'",
            f"focus_evidence='{'strong' if not missing_reason else 'weak'}'",
            f"notes='final speech appears to come from snapshot candidate path'",
        ]
    )

    return {
        "ann": "\n".join(ann_lines[-60:]).strip(),
        "step": "\n".join(step_lines[-40:]).strip(),
        "scroll": "\n".join(scroll_lines[-40:]).strip(),
        "focus": "\n".join(focus_lines[-40:]).strip(),
        "decision": "\n".join(decision_lines).strip(),
    }


def _populate_result_debug_logs(
    result_df: pd.DataFrame,
    output_path: str,
    source_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    result_df["_debug_log_path"] = ""
    result_df["_debug_log_name"] = ""
    if result_df.empty:
        return result_df

    debug_dir = Path(output_path).with_suffix("").parent / "debug_logs"
    debug_dir.mkdir(parents=True, exist_ok=True)
    write_failures = 0
    write_fail_reason = ""

    source_row_index = _build_source_row_index(source_df)

    for row_index, row in result_df.iterrows():
        if not _is_debug_log_target(row):
            continue
        req_id = str(row.get("_req_id", "") or "").strip()
        scenario_id = _sanitize_filename_part(row.get("scenario_id", ""), "scenario")
        step_normalized = _normalize_step_value(row.get("step", ""))
        step_part = _sanitize_filename_part(step_normalized, "na")
        req_part = _sanitize_filename_part(req_id, "no_req")
        file_name = f"{scenario_id}_step_{step_part}_req_{req_part}.log"
        file_path = debug_dir / file_name
        scenario_raw = str(row.get("scenario_id", "") or "").strip()
        tab_raw = str(row.get("_tab", "") or "").strip()
        source_row = source_row_index.get((scenario_raw, tab_raw, step_normalized), {})
        prev_step = ""
        try:
            prev_step = str(int(step_normalized) - 1)
        except (TypeError, ValueError):
            prev_step = ""
        prev_source_row = source_row_index.get((scenario_raw, tab_raw, prev_step), {}) if prev_step else {}
        sections = _build_debug_log_sections(row, source_row, prev_source_row)
        if not any(sections.values()):
            continue
        scenario_fallback = scenario_raw or str(source_row.get("scenario_id", "") or "")
        row_context = (
            "[ROW_CONTEXT]\n"
            f"scenario={scenario_fallback}\n"
            f"tab={tab_raw}\n"
            f"step={row.get('step', '')}\n"
            f"req_id={req_id}\n"
            f"visible={row.get('visible_label', '')}\n"
            f"speech={row.get('merged_announcement', '')}\n"
            f"final_result={row.get('final_result', '')}\n"
            f"failure_reason={row.get('failure_reason', '')}\n\n"
            f"review_note={row.get('review_note', '')}\n"
            f"noise={str(source_row.get('is_noise_step', ''))}\n"
            f"move_result={row.get('_move_result', '')}\n"
            f"debug_source_row_found={str(bool(source_row)).lower()}\n\n"
            "[ANN_TRACE]\n"
        )
        try:
            content = (
                f"{row_context}{sections.get('ann', '')}\n\n"
                "[STEP_TRACE]\n"
                f"{sections.get('step', '')}\n\n"
                "[SCROLL_TRACE]\n"
                f"{sections.get('scroll', '')}\n\n"
                "[FOCUS_TRACE]\n"
                f"{sections.get('focus', '')}\n\n"
                "[DECISION_TRACE]\n"
                f"{sections.get('decision', '')}\n"
            )
            file_path.write_text(content, encoding="utf-8")
            result_df.at[row_index, "_debug_log_path"] = str(file_path)
            result_df.at[row_index, "_debug_log_name"] = file_name
        except Exception as exc:
            write_failures += 1
            if not write_fail_reason:
                write_fail_reason = str(exc) or exc.__class__.__name__

    if write_failures:
        log(
            f"[WARN][debug_log] snippet extraction skipped count={write_failures} reason='{write_fail_reason}'"
        )
    return result_df


def _excel_col_to_name(col_idx: int) -> str:
    col_num = col_idx + 1
    letters: list[str] = []
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def add_rule_compare(df: pd.DataFrame) -> pd.DataFrame:
    def compare_row(row) -> str:
        visible = str(row.get("normalized_visible_label", "") or "").strip()
        speech = str(row.get("normalized_announcement", "") or "").strip()

        if row.get("status") == "ANCHOR":
            return "SKIP"

        if not visible and not speech:
            return "EMPTY"
        if visible == speech:
            return "EXACT"
        if visible and speech and (visible in speech or speech in visible):
            return "PARTIAL"
        return "DIFF"

    df["rule_compare"] = df.apply(compare_row, axis=1)
    return df


def stringify_complex_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: to_json_text(x) if isinstance(x, (list, dict)) else x
        )
    return df


def _fallback_plugin_group(scenario_type: object) -> str:
    normalized = str(scenario_type or "").strip().lower()
    if normalized == "global_nav":
        return "Global"
    if normalized == "content":
        return "Unknown"
    return str(scenario_type or "").strip() or "Unknown"


def annotate_plugin_metadata(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        if "plugin_group" not in df.columns:
            df["plugin_group"] = pd.Series(dtype=object)
        if "plugin_name" not in df.columns:
            df["plugin_name"] = pd.Series(dtype=object)
        return df

    result = df.copy()
    scenario_ids = (
        result["scenario_id"].fillna("").astype(str).str.strip()
        if "scenario_id" in result.columns
        else pd.Series([""] * len(result), index=result.index, dtype=object)
    )
    scenario_types = (
        result["scenario_type"].fillna("").astype(str).str.strip()
        if "scenario_type" in result.columns
        else pd.Series([""] * len(result), index=result.index, dtype=object)
    )

    plugin_groups: list[str] = []
    plugin_names: list[str] = []
    for scenario_id, scenario_type in zip(scenario_ids.tolist(), scenario_types.tolist()):
        metadata = PLUGIN_REPORT_METADATA.get(scenario_id, {})
        plugin_groups.append(metadata.get("group", _fallback_plugin_group(scenario_type)))
        plugin_names.append(metadata.get("name", scenario_id or "Unknown"))

    result["plugin_group"] = plugin_groups
    result["plugin_name"] = plugin_names
    return result


def _normalize_text(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _extract_status_tokens(text: object) -> tuple[str, list[str]]:
    normalized = _normalize_text(text)
    if not normalized:
        return "", []

    status_tokens: list[str] = []
    main_parts: list[str] = []
    for raw_part in normalized.split(","):
        part = " ".join(raw_part.strip().split())
        if not part:
            continue
        is_status = (
            "new content available" in part
            or "새 콘텐츠 사용 가능" in part
            or part in {"selected", "선택됨"}
            or ("tab " in part and " of " in part)
            or ("탭" in part and "중" in part and "번째" in part)
            or "expanded" in part
            or "collapsed" in part
            or "battery" in part
            or "percent" in part
        )
        if is_status:
            status_tokens.append(part)
        else:
            main_parts.append(part)
    return ", ".join(main_parts), status_tokens


def add_status_columns(df: pd.DataFrame) -> pd.DataFrame:
    def _resolve_source_series(priorities: list[str]) -> pd.Series:
        for col in priorities:
            if col in df.columns:
                return df[col]
        return pd.Series([""] * len(df), index=df.index, dtype=object)

    speech_series = _resolve_source_series(
        ["merged_announcement", "speech", "announcement", "normalized_announcement"]
    )
    visible_series = _resolve_source_series(
        ["visible_label", "normalized_visible_label", "text"]
    )

    speech_split = speech_series.apply(_extract_status_tokens)
    visible_split = visible_series.apply(_extract_status_tokens)

    df["speech_main"] = speech_split.apply(lambda item: item[0])
    df["speech_status_tokens"] = speech_split.apply(lambda item: item[1])
    df["visible_main"] = visible_split.apply(lambda item: item[0])
    df["visible_status_tokens"] = visible_split.apply(lambda item: item[1])
    return df


def make_filtered_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    def _bool_series(col: str) -> pd.Series:
        if col not in raw_df.columns:
            return pd.Series([False] * len(raw_df), index=raw_df.index, dtype=bool)
        return raw_df[col].fillna(False).astype(bool)

    def _text_series(col: str) -> pd.Series:
        if col not in raw_df.columns:
            return pd.Series([""] * len(raw_df), index=raw_df.index, dtype=object)
        return raw_df[col].fillna("").astype(str).str.strip()

    def _normalized_text_series(col: str) -> pd.Series:
        return _text_series(col).str.lower()

    is_noise = _bool_series("is_noise_step")
    is_duplicate = _bool_series("is_duplicate_step")
    is_recent_duplicate = _bool_series("is_recent_duplicate_step")

    final_result = _normalized_text_series("final_result").str.upper()
    failure_reason = _normalized_text_series("failure_reason")
    review_note = _normalized_text_series("review_note")
    visible = _normalized_text_series("visible_label")
    speech = _normalized_text_series("merged_announcement")
    normalized_visible = _normalized_text_series("normalized_visible_label")
    normalized_speech = _normalized_text_series("normalized_announcement")
    rule_compare = _normalized_text_series("rule_compare")
    speech_match_result = _normalized_text_series("speech_match_result")
    if "mismatch_reasons" in raw_df.columns:
        mismatch_reasons = raw_df["mismatch_reasons"]
    else:
        mismatch_reasons = pd.Series([""] * len(raw_df), index=raw_df.index, dtype=object)

    is_warn_or_fail = final_result.isin({"WARN", "FAIL"})
    has_review_fields = (review_note != "") | (failure_reason != "")
    has_visible_speech_mismatch = (
        ((normalized_visible != "") & (normalized_speech != "") & (normalized_visible != normalized_speech))
        | ((visible != "") & (speech != "") & (visible != speech))
        | (rule_compare == "diff")
        | (speech_match_result.str.contains("mismatch", na=False))
        | mismatch_reasons.apply(lambda value: bool(value) if isinstance(value, list) else bool(str(value or "").strip()))
    )

    keep_for_review = is_warn_or_fail | has_review_fields | has_visible_speech_mismatch
    keep_mask = (~(is_noise | is_duplicate | is_recent_duplicate)) | keep_for_review
    return raw_df[keep_mask].copy()


def make_summary_df(raw_df: pd.DataFrame, filtered_df: pd.DataFrame) -> pd.DataFrame:
    def _bool_series(col: str) -> pd.Series:
        if col not in raw_df.columns:
            return pd.Series([False] * len(raw_df), index=raw_df.index, dtype=bool)
        return raw_df[col].fillna(False).astype(bool)

    if "mismatch_reasons" in raw_df.columns:
        mismatch_count = int(
            raw_df["mismatch_reasons"].apply(
                lambda x: bool(x) if isinstance(x, list) else bool(str(x or "").strip())
            ).sum()
        )
    else:
        mismatch_count = int((raw_df.get("rule_compare", pd.Series(index=raw_df.index, dtype=str)) == "DIFF").sum())
    rows: list[dict[str, object]] = [
        {"section": "overall", "metric": "total_rows", "value": int(len(raw_df))},
        {"section": "overall", "metric": "raw_rows", "value": int(len(raw_df))},
        {"section": "overall", "metric": "filtered_rows", "value": int(len(filtered_df))},
        {"section": "overall", "metric": "noise_count", "value": int(_bool_series("is_noise_step").sum())},
        {"section": "overall", "metric": "duplicate_count", "value": int(_bool_series("is_duplicate_step").sum())},
        {"section": "overall", "metric": "recent_duplicate_count", "value": int(_bool_series("is_recent_duplicate_step").sum())},
        {"section": "overall", "metric": "mismatch_count", "value": mismatch_count},
    ]

    if "scenario_id" in raw_df.columns:
        scenario_df = (
            raw_df.groupby("scenario_id", dropna=False)
            .size()
            .reset_index(name="value")
            .rename(columns={"scenario_id": "metric"})
        )
        scenario_df["section"] = "scenario_id"
        rows.extend(scenario_df[["section", "metric", "value"]].to_dict("records"))

    return pd.DataFrame(rows)


def _normalize_compare_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"[\.,!?;:\-_/()\[\]{}\"'`]+", " ", text)
    return " ".join(text.split())


def _normalize_result_match_text(value: object) -> str:
    text = _normalize_compare_text(value)
    if not text:
        return ""
    suffixes = (
        " clear",
        " dry",
        " pause",
        " paused",
        " connected",
        " offline",
        " locked",
        " unlocked",
        " off",
        " on",
        " vibration detected",
        " motion detected",
        " no smoke detected",
        " no leak detected",
        " water leak detected",
        " no motion detected",
        " open",
        " closed",
        " armed",
        " armed away",
        " disarmed",
    )
    changed = True
    while changed:
        changed = False
        for suffix in suffixes:
            if text.endswith(suffix):
                text = text[: -len(suffix)].strip()
                changed = True
                break
    return text


def _step_sort_value(value: object) -> int:
    if isinstance(value, bool):
        return -1
    try:
        return int(value)
    except (TypeError, ValueError):
        return -1


def _format_repeat_steps(values: list[object]) -> str:
    steps: list[int] = []
    for value in values:
        step = _step_sort_value(value)
        if step >= 0:
            steps.append(step)
    return ",".join(str(step) for step in sorted(dict.fromkeys(steps)))


def _collapse_repeated_issue_groups(result: pd.DataFrame) -> pd.DataFrame:
    if result.empty:
        result = result.copy()
        result["repeat_count"] = pd.Series(dtype=int)
        result["first_step"] = pd.Series(dtype=object)
        result["last_step"] = pd.Series(dtype=object)
        result["steps"] = pd.Series(dtype=object)
        result["is_repeated_issue_group"] = pd.Series(dtype=bool)
        return result

    collapsed = result.copy()
    duplicate_flags = []
    recent_duplicate_flags = []
    for source_col in ("is_duplicate_step", "is_recent_duplicate_step"):
        if source_col not in collapsed.columns:
            collapsed[source_col] = False
    duplicate_flags = collapsed["is_duplicate_step"].apply(lambda value: str(value).strip().lower() in {"true", "1", "yes"})
    recent_duplicate_flags = collapsed["is_recent_duplicate_step"].apply(
        lambda value: str(value).strip().lower() in {"true", "1", "yes"}
    )

    collapsed["_repeat_norm_visible"] = collapsed["visible_label"].apply(_normalize_compare_text)
    collapsed["_repeat_norm_speech"] = collapsed["merged_announcement"].apply(_normalize_compare_text)
    collapsed["_repeat_failure"] = collapsed["failure_reason"].fillna("").astype(str).str.strip().str.lower()
    collapsed["_repeat_focus_id"] = collapsed["focus_view_id"].fillna("").astype(str).str.strip()
    collapsed["_repeat_scenario"] = collapsed["scenario_id"].fillna("").astype(str).str.strip()

    repeated_reason = collapsed["_repeat_failure"].isin(_REPEATED_ISSUE_FAILURE_REASONS)
    candidate_mask = repeated_reason | duplicate_flags | recent_duplicate_flags
    group_cols = [
        "_repeat_scenario",
        "_repeat_focus_id",
        "_repeat_norm_visible",
        "_repeat_norm_speech",
        "_repeat_failure",
    ]
    group_sizes = collapsed.groupby(group_cols, dropna=False)["_repeat_failure"].transform("size")
    collapse_mask = candidate_mask & (group_sizes > 1)

    collapsed["repeat_count"] = 1
    collapsed["first_step"] = collapsed["step"]
    collapsed["last_step"] = collapsed["step"]
    collapsed["steps"] = collapsed["step"].apply(lambda value: _format_repeat_steps([value]))
    collapsed["is_repeated_issue_group"] = False

    drop_indexes: set[int] = set()
    for _, group in collapsed[collapse_mask].groupby(group_cols, dropna=False, sort=False):
        group_idx = group.index.tolist()
        if len(group_idx) <= 1:
            continue
        step_values = group["step"].tolist()
        sorted_steps = sorted(step for step in (_step_sort_value(value) for value in step_values) if step >= 0)
        representative_idx = group_idx[0]
        collapsed.at[representative_idx, "repeat_count"] = len(group_idx)
        collapsed.at[representative_idx, "first_step"] = sorted_steps[0] if sorted_steps else group.at[representative_idx, "step"]
        collapsed.at[representative_idx, "last_step"] = sorted_steps[-1] if sorted_steps else group.at[representative_idx, "step"]
        collapsed.at[representative_idx, "steps"] = _format_repeat_steps(step_values)
        collapsed.at[representative_idx, "is_repeated_issue_group"] = True
        if "review_note" in collapsed.columns:
            note = str(collapsed.at[representative_idx, "review_note"] or "").strip()
            repeated_note = f"반복 이슈 group ({len(group_idx)} rows)"
            collapsed.at[representative_idx, "review_note"] = f"{note}; {repeated_note}" if note else repeated_note
        drop_indexes.update(group_idx[1:])

    if drop_indexes:
        collapsed = collapsed.drop(index=sorted(drop_indexes))

    return collapsed.drop(
        columns=[
            "_repeat_norm_visible",
            "_repeat_norm_speech",
            "_repeat_failure",
            "_repeat_focus_id",
            "_repeat_scenario",
        ],
        errors="ignore",
    )


def make_result_df(filtered_df: pd.DataFrame) -> pd.DataFrame:
    status_series = (
        filtered_df["status"].fillna("").astype(str).str.strip().str.upper()
        if "status" in filtered_df.columns
        else pd.Series([""] * len(filtered_df), index=filtered_df.index, dtype=object)
    )
    if "step_index" in filtered_df.columns:
        step_source = filtered_df["step_index"]
    elif "step" in filtered_df.columns:
        step_source = filtered_df["step"]
    else:
        step_source = pd.Series([-1] * len(filtered_df), index=filtered_df.index, dtype=object)
    step_series = pd.to_numeric(step_source, errors="coerce").fillna(-1)
    move_series = (
        filtered_df["move_result"].fillna("").astype(str).str.strip().str.lower()
        if "move_result" in filtered_df.columns
        else pd.Series([""] * len(filtered_df), index=filtered_df.index, dtype=object)
    )
    skip_anchor_mask = (status_series == "ANCHOR") | ((step_series == 0) & (move_series == ""))
    skipped_anchor_count = int(skip_anchor_mask.sum())
    if skipped_anchor_count > 0:
        log(f"[RESULT] skipped anchor rows count={skipped_anchor_count}")
    result_source_df = filtered_df.loc[~skip_anchor_mask].copy()
    if not result_source_df.empty:
        placeholder_mask = result_source_df.apply(lambda row: is_placeholder_row(row.to_dict()), axis=1)
        skipped_placeholder_count = int(placeholder_mask.sum())
        if skipped_placeholder_count > 0:
            log(f"[RESULT] skipped placeholder rows count={skipped_placeholder_count}")
            result_source_df = result_source_df.loc[~placeholder_mask].copy()

    if result_source_df.empty:
        return pd.DataFrame(
            columns=RESULT_SHEET_COLUMNS
            + [
                "_tab",
                "_req_id",
                "_move_result",
                "_fallback_used",
                "_step_dump_used",
                "_crop_image_path",
                "_debug_log_path",
                "_debug_log_name",
                "_traversal_result",
                "_speech_match_result",
                "_row_source",
                "_representative_row_source",
            ]
        )

    result = pd.DataFrame(index=result_source_df.index)

    def _pick_col(target: str, candidates: list[str], default: object = "") -> None:
        for col in candidates:
            if col in result_source_df.columns:
                result[target] = result_source_df[col]
                return
        result[target] = default

    _pick_col("plugin_group", ["plugin_group"])
    _pick_col("plugin_name", ["plugin_name"])
    _pick_col("scenario_id", ["scenario_id"])
    _pick_col("_tab", ["tab", "tab_name"])
    _pick_col("step", ["step", "step_index"])
    _pick_col("context_type", ["context_type"])

    _pick_col("visible_label", ["visible_label", "normalized_visible_label", "text"])
    _pick_col(
        "merged_announcement",
        ["merged_announcement", "speech", "announcement", "normalized_announcement"],
    )
    _pick_col("representative_visible", ["representative_visible"], default="")
    _pick_col("_move_result", ["move_result"])
    _pick_col("focus_view_id", ["focus_view_id", "resource_id"])
    _pick_col("focus_bounds", ["focus_bounds", "bounds"])
    _pick_col("_fallback_used", ["fallback_used"], default=False)
    _pick_col("_step_dump_used", ["step_dump_used"], default=False)
    _pick_col("_req_id", ["req_id", "get_focus_req_id"])
    _pick_col("last_smart_nav_result", ["last_smart_nav_result"])
    _pick_col("smart_nav_success", ["smart_nav_success"], default=False)
    _pick_col("smart_nav_requested_view_id", ["smart_nav_requested_view_id"])
    _pick_col("smart_nav_resolved_view_id", ["smart_nav_resolved_view_id"])
    _pick_col("smart_nav_actual_view_id", ["smart_nav_actual_view_id"])
    _pick_col("post_move_verdict_source", ["post_move_verdict_source"])
    _pick_col("_crop_image_path", ["crop_image_path", "crop_path", "result_crop"])
    _pick_col("_traversal_result", ["traversal_result"])
    _pick_col("_speech_match_result", ["speech_match_result"])
    _pick_col("_raw_final_result", ["final_result"])
    _pick_col("failure_reason", ["failure_reason"])
    _pick_col("is_duplicate_step", ["is_duplicate_step"], default=False)
    _pick_col("is_recent_duplicate_step", ["is_recent_duplicate_step"], default=False)
    _pick_col("recent_duplicate_distance", ["recent_duplicate_distance"], default=0)
    _pick_col("_row_source", ["row_source"], default="")
    _pick_col("_representative_row_source", ["representative_row_source"], default="")
    _pick_col("_representative_resource_id", ["representative_resource_id"], default="")
    _pick_col("_focus_node", ["focus_node"], default="")

    for text_col in (
        "visible_label",
        "merged_announcement",
        "representative_visible",
        "focus_view_id",
        "_representative_row_source",
        "_representative_resource_id",
    ):
        result[text_col] = result[text_col].fillna("")

    def _focus_node_label(value: object) -> str:
        node = value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                node = {}
            else:
                try:
                    node = json.loads(text)
                except Exception:
                    node = {}
        if not isinstance(node, dict):
            return ""
        for key in ("talkbackLabel", "mergedLabel", "contentDescription", "text"):
            candidate = str(node.get(key, "") or "").strip()
            if candidate:
                return candidate
        return ""

    def _representative_semantics_available(row: pd.Series) -> bool:
        return (
            str(row.get("_representative_row_source", "") or "").strip() == "representative"
            and bool(str(row.get("representative_visible", "") or "").strip())
        )

    representative_mask = result.apply(_representative_semantics_available, axis=1)
    visible_empty_mask = result["visible_label"].fillna("").astype(str).str.strip() == ""
    speech_empty_mask = result["merged_announcement"].fillna("").astype(str).str.strip() == ""
    focus_id_empty_mask = result["focus_view_id"].fillna("").astype(str).str.strip() == ""
    if representative_mask.any():
        result.loc[representative_mask & visible_empty_mask, "visible_label"] = result.loc[
            representative_mask & visible_empty_mask,
            "representative_visible",
        ]
        result.loc[representative_mask & speech_empty_mask, "merged_announcement"] = result.loc[
            representative_mask & speech_empty_mask,
            "representative_visible",
        ]
        representative_id_mask = representative_mask & (
            focus_id_empty_mask
            | (result["visible_label"].fillna("").astype(str).str.strip() == result["representative_visible"].fillna("").astype(str).str.strip())
        )
        result.loc[representative_id_mask, "focus_view_id"] = result.loc[
            representative_id_mask,
            "_representative_resource_id",
        ].where(
            result.loc[representative_id_mask, "_representative_resource_id"].fillna("").astype(str).str.strip() != "",
            result.loc[representative_id_mask, "focus_view_id"],
        )

    focus_node_labels = result["_focus_node"].apply(_focus_node_label)
    visible_empty_mask = result["visible_label"].fillna("").astype(str).str.strip() == ""
    speech_empty_mask = result["merged_announcement"].fillna("").astype(str).str.strip() == ""
    focus_node_label_mask = focus_node_labels.astype(str).str.strip() != ""
    result.loc[visible_empty_mask & focus_node_label_mask, "visible_label"] = focus_node_labels.loc[
        visible_empty_mask & focus_node_label_mask
    ]
    result.loc[speech_empty_mask & focus_node_label_mask, "merged_announcement"] = focus_node_labels.loc[
        speech_empty_mask & focus_node_label_mask
    ]

    result["_move_result"] = result.apply(
        lambda row: normalize_move_result(
            {
                "move_result": row["_move_result"],
                "last_smart_nav_result": row["last_smart_nav_result"],
                "smart_nav_success": row["smart_nav_success"],
            }
        ),
        axis=1,
    )
    result["_fallback_used"] = result["_fallback_used"].fillna(False).astype(bool)
    result["_step_dump_used"] = result["_step_dump_used"].fillna(False).astype(bool)

    result["_norm_visible"] = result["visible_label"].apply(_normalize_compare_text)
    result["_norm_speech"] = result["merged_announcement"].apply(_normalize_compare_text)
    result["_norm_visible_match"] = result["visible_label"].apply(_normalize_result_match_text)
    result["_norm_speech_match"] = result["merged_announcement"].apply(_normalize_result_match_text)
    result["_norm_representative"] = result["representative_visible"].apply(_normalize_result_match_text)
    if "mismatch_reasons" in result_source_df.columns:
        result["_mismatch_reasons"] = result_source_df["mismatch_reasons"]
    else:
        result["_mismatch_reasons"] = ""

    def _same_topic(a: str, b: str) -> bool:
        if not a or not b:
            return False
        a_tokens = [tok for tok in a.split(" ") if len(tok) > 1]
        b_tokens = [tok for tok in b.split(" ") if len(tok) > 1]
        if not a_tokens or not b_tokens:
            return False
        a_set = set(a_tokens)
        b_set = set(b_tokens)
        return len(a_set & b_set) >= max(1, min(len(a_set), len(b_set)) // 2)

    def _speech_match(row) -> str:
        if str(row.get("post_move_verdict_source", "") or "").strip().lower().startswith("smart_nav_result"):
            return "PASS_SMART_NAV"
        visible = row["_norm_visible"]
        speech = row["_norm_speech"]
        mismatch_reasons = str(row.get("_mismatch_reasons", "") or "")
        if "speech_visible_diverged" in mismatch_reasons:
            return "FAIL_MISMATCH"
        if visible and speech and visible == speech:
            return "PASS_EXACT"
        if visible and speech and (visible in speech or speech in visible):
            if _same_topic(visible, speech):
                length_gap = abs(len(visible.split(" ")) - len(speech.split(" ")))
                return "WARN_CONTEXT_ADDED" if length_gap >= 2 else "PASS_CONTAINS"
            return "PASS_CONTAINS"
        if _same_topic(visible, speech):
            return "WARN_CONTEXT_ADDED"
        return "FAIL_MISMATCH"

    computed_speech_match = result.apply(_speech_match, axis=1)
    existing_speech_match = result["_speech_match_result"].fillna("").astype(str).str.strip()
    result["_speech_match_result"] = existing_speech_match.where(existing_speech_match != "", computed_speech_match)

    group_keys = [k for k in ["scenario_id", "_tab", "context_type"] if k in result.columns]
    if not group_keys:
        group_keys = [result.index]

    for _, group in result.groupby(group_keys, dropna=False, sort=False):
        group_idx = group.index.tolist()
        moved_idx = [idx for idx in group_idx if result.at[idx, "_move_result"] == "moved"]
        last_moved_idx = moved_idx[-1] if moved_idx else None

        traversal = []
        failure = []
        for idx in group_idx:
            move_result = result.at[idx, "_move_result"]
            is_followup_noise = False
            if last_moved_idx is not None and idx > last_moved_idx:
                same_visible = result.at[idx, "_norm_visible"] == result.at[last_moved_idx, "_norm_visible"]
                same_speech = result.at[idx, "_norm_speech"] == result.at[last_moved_idx, "_norm_speech"]
                is_repeat_stop = "repeat_no_progress" in str(result.at[idx, "_req_id"] or "")
                is_followup_noise = move_result == "failed" and (same_visible or same_speech or is_repeat_stop)

            if idx == last_moved_idx and any(i > last_moved_idx for i in group_idx if result.at[i, "_move_result"] == "failed"):
                traversal.append("WARN_TERMINAL_BY_REPEAT_STOP")
                failure.append("terminal_not_handled")
            elif move_result == "moved":
                traversal.append("PASS_MOVED")
                failure.append("")
            elif is_followup_noise:
                traversal.append("WARN_TERMINAL_BY_REPEAT_STOP")
                failure.append("terminal_followup_noise")
            elif move_result == "failed":
                if moved_idx:
                    traversal.append("FAIL_STUCK")
                    failure.append("repeat_no_progress")
                else:
                    traversal.append("FAIL_MOVE")
                    failure.append("move_failed")
            else:
                traversal.append("FAIL_MOVE")
                failure.append("move_failed")

        existing_traversal = result.loc[group_idx, "_traversal_result"].fillna("").astype(str).str.strip()
        existing_failure = result.loc[group_idx, "failure_reason"].fillna("").astype(str).str.strip()
        computed_traversal = pd.Series(traversal, index=group_idx, dtype=object)
        computed_failure = pd.Series(failure, index=group_idx, dtype=object)
        result.loc[group_idx, "_traversal_result"] = existing_traversal.where(existing_traversal != "", computed_traversal)
        result.loc[group_idx, "failure_reason"] = existing_failure.where(existing_failure != "", computed_failure)

    def _focus_confidence(row) -> str:
        if row["_fallback_used"] or row["_step_dump_used"]:
            return "LOW"
        if row["_speech_match_result"] in {"PASS_EXACT", "PASS_CONTAINS"} and row["_traversal_result"] == "PASS_MOVED":
            return "HIGH"
        if row["_speech_match_result"] == "PASS_SMART_NAV" and row["_traversal_result"] == "PASS_MOVED":
            return "HIGH"
        if row["_speech_match_result"] == "FAIL_MISMATCH" or row["_traversal_result"].startswith("FAIL"):
            return "LOW"
        return "MEDIUM"

    result["focus_confidence"] = result.apply(_focus_confidence, axis=1)

    def _mismatch_type(row) -> str:
        visible_raw = str(row.get("visible_label", "") or "").strip()
        speech_raw = str(row.get("merged_announcement", "") or "").strip()
        visible = row["_norm_visible"]
        speech = row["_norm_speech"]
        visible_match = row["_norm_visible_match"]
        speech_match = row["_norm_speech_match"]
        representative = row["_norm_representative"]
        if not visible_raw:
            return "EMPTY_VISIBLE"
        if not speech_raw:
            return "EMPTY_SPEECH"
        if row.get("_row_source", "") == "actual_focus" and representative and representative != visible_match:
            return "REPRESENTATIVE_CONTEXT"
        if visible == speech:
            return "EXACT_MATCH"
        if visible_match and speech_match and visible_match == speech_match:
            return "NORMALIZED_MATCH"
        if visible and speech and (visible in speech or speech in visible):
            return "PARTIAL_MATCH"
        state_diff_markers = ("clear", "dry", "pause", "connected", "offline", "locked", "off", "on", "detected")
        if visible_match == speech_match and any(token in (visible + " " + speech) for token in state_diff_markers):
            return "SPEECH_ONLY_STATE_DIFF"
        if row["_speech_match_result"] == "FAIL_MISMATCH":
            return "LABEL_MISMATCH"
        return "UNKNOWN"

    result["mismatch_type"] = result.apply(_mismatch_type, axis=1)

    _ACCESSIBILITY_PASS_MATCH_TYPES = {
        "EXACT_MATCH",
        "NORMALIZED_MATCH",
        "PARTIAL_MATCH",
        "REPRESENTATIVE_CONTEXT",
    }
    _TERMINAL_PRESENTATION_REASONS = {
        "repeat_no_progress",
        "terminal_not_handled",
        "terminal_followup_noise",
        "move_terminal",
    }

    def _final_result(row) -> str:
        mismatch_type = str(row.get("mismatch_type", "") or "").strip().upper()
        failure_reason = str(row.get("failure_reason", "") or "").strip().lower()
        traversal_result = str(row.get("_traversal_result", "") or "").strip().upper()
        speech_match_result = str(row.get("_speech_match_result", "") or "").strip().upper()
        focus_confidence = str(row.get("focus_confidence", "") or "").strip().upper()
        accessibility_match_ok = mismatch_type in _ACCESSIBILITY_PASS_MATCH_TYPES

        if mismatch_type in {"LABEL_MISMATCH", "EMPTY_SPEECH", "EMPTY_VISIBLE"} or speech_match_result == "FAIL_MISMATCH":
            return "FAIL"
        if traversal_result in {"FAIL_MOVE", "FAIL_STUCK"}:
            return "WARN" if accessibility_match_ok else "FAIL"
        if traversal_result == "WARN_TERMINAL_BY_REPEAT_STOP":
            return "WARN"
        if speech_match_result == "WARN_CONTEXT_ADDED" or focus_confidence == "LOW":
            return "WARN"
        if accessibility_match_ok and traversal_result in {"PASS_MOVED", "PASS_SCROLLED"}:
            return "PASS"
        if accessibility_match_ok and failure_reason in _TERMINAL_PRESENTATION_REASONS:
            return "WARN"
        if traversal_result in {"PASS_MOVED", "PASS_SCROLLED"} and speech_match_result == "PASS_SMART_NAV":
            return "PASS"
        return "WARN"

    computed_final_result = result.apply(_final_result, axis=1)
    result["final_result"] = computed_final_result
    result["failure_reason"] = result.apply(
        lambda row: row["failure_reason"]
        if row["failure_reason"]
        else ("speech_visible_diverged" if row["_speech_match_result"] == "FAIL_MISMATCH" else ("fallback_dependent" if row["focus_confidence"] == "LOW" and row["final_result"] != "FAIL" else "")),
        axis=1,
    )

    def _review_note(row) -> str:
        final_result = str(row.get("final_result", "") or "").strip().upper()
        mismatch_type = str(row.get("mismatch_type", "") or "").strip().upper()
        failure_reason = str(row.get("failure_reason", "") or "").strip().lower()
        traversal_result = str(row.get("_traversal_result", "") or "").strip().upper()
        speech_match_result = str(row.get("_speech_match_result", "") or "").strip().upper()

        if final_result == "PASS":
            return "정상 이동 및 발화 일치"
        if mismatch_type in _ACCESSIBILITY_PASS_MATCH_TYPES and (
            traversal_result == "WARN_TERMINAL_BY_REPEAT_STOP" or failure_reason in _TERMINAL_PRESENTATION_REASONS
        ):
            return "발화 일치, 탐색 종료 reason 있음"
        if speech_match_result == "WARN_CONTEXT_ADDED":
            return "정상 이동, speech에 상위 문맥 포함"
        if traversal_result == "FAIL_STUCK":
            return "동일 항목 반복 후 종료"
        if speech_match_result == "FAIL_MISMATCH":
            return "speech와 visible 불일치"
        return "이동/발화 결과 재검토 필요"

    result["review_note"] = result.apply(_review_note, axis=1)
    result["_debug_log_path"] = ""
    result["_debug_log_name"] = ""
    result["result_crop_thumbnail"] = ""
    result = _collapse_repeated_issue_groups(result)

    public_columns = [*RESULT_SHEET_COLUMNS, *RESULT_REPEAT_METADATA_COLUMNS]
    helper_columns = [
        "_tab",
        "_req_id",
        "_move_result",
        "_fallback_used",
        "_step_dump_used",
        "_crop_image_path",
        "_debug_log_path",
        "_debug_log_name",
        "_traversal_result",
        "_speech_match_result",
        "_row_source",
        "_representative_row_source",
        "_mismatch_reasons",
    ]
    return result[[*public_columns, *helper_columns]]


def _apply_result_crop_hyperlinks(writer: pd.ExcelWriter, result_df: pd.DataFrame) -> None:
    if "result" not in writer.sheets or "_crop_image_path" not in result_df.columns:
        return

    valid_exts = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}

    def _to_hyperlink_target(path_text: str, *, xlsxwriter_mode: bool) -> tuple[str | None, str | None]:
        normalized = path_text.strip()
        if not normalized:
            return None, "empty_path"

        path_obj = Path(normalized)
        if path_obj.suffix and path_obj.suffix.lower() not in valid_exts:
            return None, "unsupported_extension"

        if normalized.startswith(("http://", "https://", "file://", "mailto:", "internal:", "external:")):
            return normalized, None

        abs_path = os.path.abspath(os.path.normpath(normalized))
        if not abs_path:
            return None, "unsupported_path_format"

        if not xlsxwriter_mode:
            return abs_path, None
        safe_path = abs_path.replace("\\", "/")
        if not safe_path:
            return None, "unsupported_path_format"
        return f"external:{safe_path}", None

    ws = writer.sheets["result"]
    if "result_crop_thumbnail" not in result_df.columns:
        return
    crop_col_idx = result_df.columns.get_loc("result_crop_thumbnail")
    is_openpyxl_sheet = hasattr(ws, "cell")
    is_xlsxwriter_sheet = hasattr(ws, "write_url")
    skipped_reasons: dict[str, int] = {}

    for row_idx, crop_path in enumerate(result_df["_crop_image_path"].tolist(), start=2):
        path_text = str(crop_path or "").strip()
        if not path_text:
            continue
        display_text = _display_name_for_path(path_text)
        target, skip_reason = _to_hyperlink_target(path_text, xlsxwriter_mode=is_xlsxwriter_sheet)
        if skip_reason:
            skipped_reasons[skip_reason] = skipped_reasons.get(skip_reason, 0) + 1
            if is_openpyxl_sheet:
                ws.cell(row=row_idx, column=crop_col_idx + 1).value = display_text
            elif hasattr(ws, "write"):
                ws.write(row_idx - 1, crop_col_idx, display_text)
            continue
        try:
            if is_openpyxl_sheet:
                cell = ws.cell(row=row_idx, column=crop_col_idx + 1)
                cell.value = display_text
                cell.hyperlink = target
                cell.style = "Hyperlink"
            elif is_xlsxwriter_sheet:
                ws.write_url(row_idx - 1, crop_col_idx, target, string=display_text)
            else:
                ws.write(row_idx - 1, crop_col_idx, display_text)
        except Exception as exc:
            skipped_reasons[str(exc)] = skipped_reasons.get(str(exc), 0) + 1
            try:
                if is_openpyxl_sheet:
                    ws.cell(row=row_idx, column=crop_col_idx + 1).value = display_text
                elif hasattr(ws, "write"):
                    ws.write(row_idx - 1, crop_col_idx, display_text)
            except Exception:
                continue

    skipped_total = sum(skipped_reasons.values())
    if skipped_total:
        reason_summary = ", ".join(
            f"{reason} x{count}" for reason, count in sorted(skipped_reasons.items(), key=lambda item: item[0])
        )
        log(
            f"[WARN][excel] result crop hyperlink skipped for {skipped_total} rows reasons='{reason_summary}'"
        )


def _apply_result_debug_log_hyperlinks(writer: pd.ExcelWriter, result_df: pd.DataFrame) -> None:
    return


def _apply_result_visual_enhancements(
    writer: pd.ExcelWriter,
    result_df: pd.DataFrame,
    *,
    with_images: bool,
) -> None:
    if "result" not in writer.sheets or result_df.empty or "final_result" not in result_df.columns:
        return

    ws = writer.sheets["result"]
    is_openpyxl_sheet = hasattr(ws, "cell")
    is_xlsxwriter_sheet = hasattr(ws, "conditional_format")
    display_columns = RESULT_SHEET_COLUMNS
    final_col_idx = display_columns.index("final_result")
    thumb_col_idx = display_columns.index("result_crop_thumbnail") if "result_crop_thumbnail" in display_columns else -1
    color_map = {
        "PASS": "C6EFCE",
        "WARN": "FFEB9C",
        "FAIL": "FFC7CE",
    }

    def _has_mismatch(row_obj) -> bool:
        mismatch = str(getattr(row_obj, "mismatch_type", "") or "").strip().upper()
        return mismatch not in {"", "EXACT_MATCH"}

    if is_openpyxl_sheet:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import PatternFill

        max_col = len(display_columns)
        thumb_failures = 0
        thumb_fail_reasons: dict[str, int] = {}
        for row_idx, (_, row) in enumerate(result_df.iterrows(), start=2):
            final_result = str(row.get("final_result", "") or "").upper()
            fill_color = color_map.get(final_result)
            if fill_color:
                fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
                for col in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

            if thumb_col_idx < 0:
                continue
            if not with_images:
                thumb_cell = ws.cell(row=row_idx, column=thumb_col_idx + 1)
                thumb_cell.value = ""
                thumb_cell.hyperlink = None
                continue

            mismatch_exists = _has_mismatch(row)
            needs_thumb = final_result in {"WARN", "FAIL"} or mismatch_exists
            crop_path = str(row.get("_crop_image_path", "") or "").strip()
            thumb_cell = ws.cell(row=row_idx, column=thumb_col_idx + 1)
            if not needs_thumb:
                thumb_cell.value = ""
                thumb_cell.hyperlink = None
                continue
            if not crop_path or not Path(crop_path).exists():
                thumb_cell.value = _display_name_for_path(crop_path) if crop_path else ""
                continue
            thumb_cell.value = _display_name_for_path(crop_path)
            try:
                thumb_data = create_excel_thumbnail(crop_path, as_bytes=True)
                if not thumb_data:
                    thumb_failures += 1
                    thumb_fail_reasons["thumbnail_create_failed"] = thumb_fail_reasons.get("thumbnail_create_failed", 0) + 1
                    continue
                img = XLImage(BytesIO(thumb_data))
                ws.add_image(img, thumb_cell.coordinate)
                row_height = (float(getattr(img, "height", 0) or 0) * 0.75) + 6.0
                ws.row_dimensions[row_idx].height = max(float(ws.row_dimensions[row_idx].height or 0), row_height)
            except Exception as exc:
                thumb_failures += 1
                reason = str(exc) or exc.__class__.__name__
                thumb_fail_reasons[reason] = thumb_fail_reasons.get(reason, 0) + 1
        if thumb_col_idx >= 0:
            ws.column_dimensions[ws.cell(row=1, column=thumb_col_idx + 1).column_letter].width = 24
        if thumb_failures:
            top_reason = max(thumb_fail_reasons.items(), key=lambda item: item[1])[0] if thumb_fail_reasons else "unknown"
            log(
                f"[WARN][excel] thumbnail insertion skipped count={thumb_failures} reason='{top_reason}'"
            )
        return

    if is_xlsxwriter_sheet:
        workbook = writer.book
        max_row = len(result_df)
        max_col = len(display_columns) - 1
        row_range = f"A2:{_excel_col_to_name(max_col)}{max_row + 1}"
        pass_fmt = workbook.add_format({"bg_color": "#C6EFCE"})
        warn_fmt = workbook.add_format({"bg_color": "#FFEB9C"})
        fail_fmt = workbook.add_format({"bg_color": "#FFC7CE"})
        result_col_letter = _excel_col_to_name(final_col_idx)
        ws.conditional_format(row_range, {"type": "formula", "criteria": f'=${result_col_letter}2="PASS"', "format": pass_fmt})
        ws.conditional_format(row_range, {"type": "formula", "criteria": f'=${result_col_letter}2="WARN"', "format": warn_fmt})
        ws.conditional_format(row_range, {"type": "formula", "criteria": f'=${result_col_letter}2="FAIL"', "format": fail_fmt})
        if thumb_col_idx < 0:
            return

        thumb_failures = 0
        thumb_fail_reasons: dict[str, int] = {}
        for row_number, (_, row) in enumerate(result_df.iterrows(), start=2):
            final_result = str(row.get("final_result", "") or "").upper()
            mismatch_exists = _has_mismatch(row)
            needs_thumb = final_result in {"WARN", "FAIL"} or mismatch_exists
            crop_path = str(row.get("_crop_image_path", "") or "").strip()
            display_name = _display_name_for_path(crop_path) if crop_path else ""
            if not needs_thumb:
                ws.write(row_number - 1, thumb_col_idx, "")
                continue
            if not with_images:
                ws.write(row_number - 1, thumb_col_idx, "")
                continue
            ws.write(row_number - 1, thumb_col_idx, display_name)
            if not crop_path or not Path(crop_path).exists():
                continue
            try:
                thumb_data = create_excel_thumbnail(crop_path, as_bytes=True)
                if not thumb_data:
                    thumb_failures += 1
                    thumb_fail_reasons["thumbnail_create_failed"] = thumb_fail_reasons.get("thumbnail_create_failed", 0) + 1
                    continue
                ws.insert_image(
                    row_number - 1,
                    thumb_col_idx,
                    f"excel_thumb_{row_number}.png",
                    {"object_position": 1, "image_data": BytesIO(thumb_data)},
                )
                ws.set_row(row_number - 1, 78)
            except Exception as exc:
                thumb_failures += 1
                reason = str(exc) or exc.__class__.__name__
                thumb_fail_reasons[reason] = thumb_fail_reasons.get(reason, 0) + 1
        ws.set_column(thumb_col_idx, thumb_col_idx, 24)
        if thumb_failures:
            top_reason = max(thumb_fail_reasons.items(), key=lambda item: item[1])[0] if thumb_fail_reasons else "unknown"
            log(
                f"[WARN][excel] thumbnail insertion skipped count={thumb_failures} reason='{top_reason}'"
            )


def save_excel(rows: list[dict], output_path: str, with_images: bool = True) -> None:
    df = pd.DataFrame(rows)
    if df.empty:
        log("[SAVE] skip: no rows")
        return

    df = annotate_plugin_metadata(df)
    df = add_rule_compare(df)
    df = add_status_columns(df)
    raw_df = df.copy()
    filtered_df = make_filtered_df(raw_df)
    summary_df = make_summary_df(raw_df, filtered_df)
    result_df = make_result_df(filtered_df)
    result_df = _populate_result_debug_logs(result_df, output_path, source_df=filtered_df)
    if _overlay_first_row_debug_enabled():
        tracked_keys: list[str] = []
        for row in rows:
            row_key = str((row or {}).get("_overlay_first_row_key", "") or "").strip() if isinstance(row, dict) else ""
            if row_key and row_key not in tracked_keys:
                tracked_keys.append(row_key)
        if tracked_keys:
            recent_logs = get_recent_logs(limit=600)
            for tracked_key in tracked_keys:
                raw_matches = [row for row in rows if str((row or {}).get("_overlay_first_row_key", "") or "").strip() == tracked_key]
                found_in_raw_rows = bool(raw_matches)
                found_in_filtered_rows = (
                    "_overlay_first_row_key" in filtered_df.columns
                    and bool((filtered_df["_overlay_first_row_key"].astype(str) == tracked_key).any())
                )
                found_in_result_rows = (
                    "_overlay_first_row_key" in result_df.columns
                    and bool((result_df["_overlay_first_row_key"].astype(str) == tracked_key).any())
                )
                found_in_debug_logs = any(tracked_key in line for line in recent_logs)
                sample_row = raw_matches[0] if raw_matches else {}
                placeholder_eval = is_placeholder_row(sample_row) if sample_row else False
                filtered_out = found_in_raw_rows and not found_in_filtered_rows
                drop_reason = ""
                if found_in_raw_rows and not found_in_filtered_rows:
                    drop_reason = "dropped_before_filtered"
                elif found_in_filtered_rows and not found_in_result_rows:
                    drop_reason = "dropped_before_result"
                elif found_in_result_rows:
                    drop_reason = "not_dropped"
                else:
                    drop_reason = "not_found_in_export_frames"
                log(
                    f"[OVERLAY][FIRSTROW][pre_result_export] row_key='{tracked_key}' "
                    f"found_in_raw_rows={str(found_in_raw_rows).lower()} "
                    f"found_in_result_rows={str(found_in_result_rows).lower()} "
                    f"found_in_debug_logs={str(found_in_debug_logs).lower()} "
                    f"drop_reason='{drop_reason}' placeholder_eval={str(placeholder_eval).lower()} "
                    f"filtered_out={str(filtered_out).lower()}",
                    level="DEBUG",
                )

    log("[SAVE] writing sheets raw/filtered/summary/result")
    log(f"[SAVE] filtered rows={len(filtered_df)}, raw rows={len(raw_df)}")

    ordered_cols = [
        "plugin_group",
        "plugin_name",
        "scenario_id",
        "tab_name",
        "context_type",
        "parent_step_index",
        "overlay_entry_label",
        "overlay_recovery_status",
        "step_index",
        "status",
        "stop_reason",
        "step_elapsed_sec",
        "move_result",
        "visible_label",
        "normalized_visible_label",
        "merged_announcement",
        "normalized_announcement",
        "rule_compare",
        "focus_text",
        "focus_content_description",
        "focus_view_id",
        "focus_bounds",
        "crop_image",
        "crop_image_path",
        "crop_image_saved",
        "partial_announcements",
        "last_announcements",
        "last_merged_announcement",
        "focus_node",
        "dump_tree_nodes",
        "fingerprint",
        "is_duplicate_step",
        "is_recent_duplicate_step",
        "recent_duplicate_distance",
        "recent_duplicate_of_step",
        "is_noise_step",
        "noise_reason",
        "speech_main",
        "speech_status_tokens",
        "visible_main",
        "visible_status_tokens",
    ]

    def _reorder_columns(frame: pd.DataFrame) -> pd.DataFrame:
        existing_cols = [c for c in ordered_cols if c in frame.columns]
        remaining_cols = [c for c in frame.columns if c not in existing_cols]
        return frame[existing_cols + remaining_cols]

    def _move_column_to_end(frame: pd.DataFrame, col_name: str) -> pd.DataFrame:
        if col_name not in frame.columns:
            return frame
        return frame[[c for c in frame.columns if c != col_name] + [col_name]]

    raw_df = _reorder_columns(raw_df)
    filtered_df = _reorder_columns(filtered_df)
    raw_df = _move_column_to_end(raw_df, "crop_image")
    raw_df = stringify_complex_columns(raw_df)
    filtered_df = stringify_complex_columns(filtered_df)
    summary_df = stringify_complex_columns(summary_df)
    result_df = stringify_complex_columns(result_df)
    result_export_df = result_df[RESULT_SHEET_COLUMNS].copy()

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path) as writer:
        raw_df.to_excel(writer, sheet_name="raw", index=False)
        filtered_df.to_excel(writer, sheet_name="filtered", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        result_export_df.to_excel(writer, sheet_name="result", index=False)
        _apply_result_crop_hyperlinks(writer, result_df)
        _apply_result_debug_log_hyperlinks(writer, result_df)
        _apply_result_visual_enhancements(writer, result_df, with_images=with_images)

    if with_images and "crop_image" in raw_df.columns:
        insert_images_to_excel(output_path, image_col_name="crop_image", sheet_name="raw")

    log(f"[SAVE] saved excel: {output_path} rows={len(raw_df)} with_images={with_images}")
