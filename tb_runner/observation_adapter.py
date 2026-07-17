"""Read-only artifact resolution and observation-set reconstruction."""

from __future__ import annotations

import hashlib
import json
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping

from tb_runner.canonical_json import canonical_sha256
from tb_runner.comparator_schema import ComparatorInput
from tb_runner.observation_normalizer import (
    build_observation,
    parse_bounds,
    parse_json_object,
)
from tb_runner.observation_schema import (
    OBSERVATION_SET_SCHEMA_VERSION,
    CanonicalObservation,
    ObservationAvailability,
    ObservationSet,
)


_EVIDENCE_SCHEMA = "evidence-event-v1"
_COVERAGE_SCHEMA = "audit-v7-focusable-coverage-v1"
_INVENTORY_SCHEMA = "audit-v7-focusable-inventory-v1"
_OBSERVATION_TYPES = (
    "evidence_ledger",
    "xlsx",
    "focusable_coverage",
    "focusable_inventory",
    "run_summary",
    "evidence_manifest",
)
_NODE_EVENT_PRIORITY = {
    "TARGET_RESOLVED": 1,
    "DELAYED_OBSERVATION": 2,
    "POST_ACTION_OBSERVATION": 3,
    "POST_FOCUS_OBSERVED": 4,
}


class ObservationArtifactError(ValueError):
    def __init__(self, status: ObservationAvailability, code: str, **details: Any):
        super().__init__(code)
        self.status = status
        self.code = code
        self.details = details


def _step_index(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        prefix = str(value).split(":", 1)[0]
        try:
            return int(prefix)
        except ValueError:
            return None

    def to_dict(self) -> dict[str, Any]:
        return {"code": self.code, "details": self.details}


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _artifact_entries(source: ComparatorInput) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for group in ("required", "optional"):
        values = source.artifacts.get(group)
        if not isinstance(values, list):
            continue
        for item in values:
            if isinstance(item, Mapping) and item.get("artifact_type"):
                result[str(item["artifact_type"])] = dict(item)
    declared = source.artifacts.get("optional_observations") or {}
    if isinstance(declared, Mapping):
        for artifact_type, status in declared.items():
            if (
                artifact_type in result
                and isinstance(status, Mapping)
                and status.get("status") not in {None, "AVAILABLE"}
            ):
                result[artifact_type]["availability"] = "UNAVAILABLE"
    return result


def _resolve_reference(
    reference: str,
    *,
    qa_runs_root: Path | None,
    artifact_root: Path | None,
) -> tuple[Path | None, str]:
    if reference.startswith("artifact://sha256/"):
        digest = reference.rsplit("/", 1)[-1]
        if artifact_root is None or len(digest) != 64:
            return None, "PINNED_STORE_UNAVAILABLE"
        path = artifact_root / "sha256" / digest[:2] / digest / "payload"
        return (path, "PINNED") if path.is_file() else (None, "PINNED_PAYLOAD_MISSING")
    if reference.startswith("qa-run://"):
        if qa_runs_root is None:
            return None, "QA_RUN_ROOT_UNAVAILABLE"
        parts = reference.removeprefix("qa-run://").split("/")
        if len(parts) < 3:
            return None, "INVALID_LOGICAL_REFERENCE"
        batch_id, device_token, filename = parts[0], parts[1], parts[-1]
        if not filename or Path(filename).name != filename:
            return None, "INVALID_LOGICAL_REFERENCE"
        batch_root = qa_runs_root / batch_id
        if not batch_root.is_dir():
            return None, "SOURCE_BATCH_MISSING"
        if device_token == "device":
            candidates = [
                path / filename
                for path in sorted(batch_root.glob("device_*"))
                if (path / filename).is_file()
            ]
        else:
            candidate = batch_root / device_token / filename
            candidates = [candidate] if candidate.is_file() else []
        if len(candidates) == 1:
            return candidates[0], "LOCAL_LOGICAL"
        if len(candidates) > 1:
            return None, "AMBIGUOUS_LOCAL_SOURCE"
        return None, "LOCAL_SOURCE_MISSING"
    return None, "UNSUPPORTED_REFERENCE"


def _resolve_artifact(
    entry: Mapping[str, Any],
    *,
    qa_runs_root: Path | None,
    artifact_root: Path | None,
) -> tuple[Path | None, dict[str, Any]]:
    artifact_type = str(entry.get("artifact_type") or "")
    reference = str(entry.get("reference") or "")
    expected_digest = str(entry.get("digest") or "").lower()
    public = {
        "artifact_type": artifact_type,
        "logical_reference": reference or None,
        "artifact_digest": expected_digest or None,
        "schema_version": entry.get("schema_version"),
        "availability": "UNAVAILABLE",
        "resolution": None,
    }
    if entry.get("availability") != "AVAILABLE" or not reference or not expected_digest:
        public["reason"] = "MANIFEST_ARTIFACT_UNAVAILABLE"
        return None, public
    path, resolution = _resolve_reference(
        reference,
        qa_runs_root=qa_runs_root,
        artifact_root=artifact_root,
    )
    public["resolution"] = resolution
    if path is None:
        public["reason"] = resolution
        return None, public
    try:
        actual_digest = _sha256_file(path)
    except OSError:
        public["availability"] = "CORRUPT"
        public["reason"] = "ARTIFACT_READ_FAILED"
        return None, public
    if actual_digest != expected_digest:
        public["availability"] = "CORRUPT"
        public["reason"] = "ARTIFACT_DIGEST_MISMATCH"
        public["actual_digest"] = actual_digest
        return None, public
    public["availability"] = "AVAILABLE"
    return path, public


def _event_payload(event: Mapping[str, Any]) -> dict[str, Any]:
    value = event.get("payload")
    return dict(value) if isinstance(value, Mapping) else {}


def _node_from_event(event_type: str, payload: Mapping[str, Any]) -> dict[str, Any]:
    if event_type == "TARGET_RESOLVED":
        value = payload.get("resolvedTarget")
    else:
        value = payload.get("observation")
    return dict(value) if isinstance(value, Mapping) else {}


def _read_evidence(path: Path) -> tuple[list[dict[str, Any]], str]:
    transactions: dict[str, dict[str, Any]] = {}
    scenario_terminal: dict[str, dict[str, Any]] = {}
    schema: str | None = None
    try:
        handle = path.open("r", encoding="utf-8")
    except OSError as exc:
        raise ObservationArtifactError(
            ObservationAvailability.CORRUPT, "EVIDENCE_READ_FAILED"
        ) from exc
    with handle:
        for line_number, line in enumerate(handle, 1):
            try:
                event = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ObservationArtifactError(
                    ObservationAvailability.CORRUPT,
                    "EVIDENCE_JSONL_CORRUPT",
                    line=line_number,
                ) from exc
            if not isinstance(event, Mapping):
                continue
            event_schema = str(event.get("schema_version") or "")
            schema = schema or event_schema
            if event_schema != _EVIDENCE_SCHEMA:
                raise ObservationArtifactError(
                    ObservationAvailability.UNSUPPORTED_SCHEMA,
                    "EVIDENCE_SCHEMA_UNSUPPORTED",
                    actual=event_schema,
                )
            event_type = str(event.get("event_type") or "")
            scenario_id = str(event.get("scenario_id") or "")
            if event_type == "SCENARIO_TERMINAL":
                scenario_terminal[scenario_id] = _event_payload(event)
                continue
            transaction_id = str(event.get("transaction_id") or "")
            if not transaction_id:
                continue
            item = transactions.setdefault(
                transaction_id,
                {
                    "scenario_id": scenario_id,
                    "step_index": _step_index(event.get("step_index")),
                    "transaction_id": transaction_id,
                    "request_id": str(event.get("logical_action_id") or ""),
                    "announcements": [],
                    "node_priority": 0,
                },
            )
            payload = _event_payload(event)
            if event_type in _NODE_EVENT_PRIORITY:
                priority = _NODE_EVENT_PRIORITY[event_type]
                node = _node_from_event(event_type, payload)
                if node and priority >= int(item.get("node_priority") or 0):
                    item["node"] = node
                    item["node_priority"] = priority
            elif event_type == "ANNOUNCEMENT_OBSERVED":
                text = payload.get("text")
                if text:
                    item["announcements"].append(str(text))
            elif event_type == "ACTION_API_RESULT":
                item["action_type"] = str(payload.get("action") or "")
                item["action_success"] = payload.get("success")
                item["action_reason"] = str(payload.get("reason") or "")
            elif event_type == "SHADOW_ACTION_REDUCED_V2":
                item["identity_verdict"] = str(payload.get("verdict") or "")
            elif event_type == "PRODUCTION_TRAVERSAL_GATE_DECIDED":
                item["progress_verdict"] = str(
                    payload.get("progress_verdict")
                    or payload.get("verdict")
                    or payload.get("decision")
                    or ""
                )
            elif event_type == "VISIT_DECIDED":
                item["visit_verdict"] = str(
                    payload.get("visit_verdict")
                    or payload.get("decision")
                    or payload.get("verdict")
                    or ""
                )
            elif event_type == "RECOVERY_CANDIDATE_RESULT":
                item["recovery_result"] = str(
                    payload.get("result") or payload.get("recovery_result") or ""
                )
    result: list[dict[str, Any]] = []
    by_scenario: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in transactions.values():
        node = item.get("node")
        if not isinstance(node, Mapping):
            continue
        announcements = []
        for value in item.get("announcements") or []:
            if not announcements or announcements[-1] != value:
                announcements.append(value)
        item["announcement"] = ", ".join(announcements)
        item["talkback_speech"] = item["announcement"]
        item["package"] = (
            node.get("packageName") or node.get("package") or node.get("package_name")
        )
        item["resource_id"] = (
            node.get("viewIdResourceName")
            or node.get("resourceId")
            or node.get("resource_id")
        )
        item["class_name"] = node.get("className") or node.get("class_name")
        item["bounds"] = node.get("boundsInScreen") or node.get("bounds")
        item["content_description"] = (
            node.get("contentDescription") or node.get("content_description")
        )
        item["visible_text"] = node.get("text") or node.get("talkbackLabel")
        item["parent_signature"] = node.get("parentPath") or node.get("parent_path")
        item["ancestor_signature"] = node.get("nodePath") or node.get("node_path")
        by_scenario[item["scenario_id"]].append(item)
        result.append(item)
    for scenario_id, terminal in scenario_terminal.items():
        values = by_scenario.get(scenario_id, [])
        if values:
            last = max(
                values,
                key=lambda item: (
                    _step_index(item.get("step_index")) or -1,
                    str(item.get("transaction_id") or ""),
                ),
            )
            last["terminal"] = True
            last["stop_reason"] = str(
                terminal.get("stop_reason") or terminal.get("reason") or ""
            )
    return result, schema or ""


def _sheet_rows(workbook: Any, sheet_name: str) -> tuple[list[str], Iterable[tuple[Any, ...]]]:
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(item or "") for item in next(rows)]
    return headers, rows


def _read_xlsx(
    path: Path,
    evidence_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ObservationArtifactError(
            ObservationAvailability.UNSUPPORTED_SCHEMA,
            "XLSX_READER_UNAVAILABLE",
        ) from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ObservationArtifactError(
            ObservationAvailability.CORRUPT, "XLSX_CORRUPT"
        ) from exc
    try:
        if not {"raw", "result"}.issubset(workbook.sheetnames):
            raise ObservationArtifactError(
                ObservationAvailability.UNSUPPORTED_SCHEMA,
                "XLSX_REQUIRED_SHEETS_MISSING",
            )
        result_headers, result_rows = _sheet_rows(workbook, "result")
        result_index = {name: index for index, name in enumerate(result_headers)}
        quality: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
        for values in result_rows:
            scenario = str(values[result_index.get("scenario_id", 0)] or "")
            step_value = values[result_index.get("step", 0)]
            try:
                step = _step_index(step_value)
            except (TypeError, ValueError):
                continue
            if step is None:
                continue
            quality[(scenario, step)].append(
                {
                    name: values[index]
                    for name, index in result_index.items()
                    if name
                    in {
                        "visible_label",
                        "merged_announcement",
                        "mismatch_type",
                        "final_result",
                        "failure_reason",
                        "focus_view_id",
                        "shadow_verdict",
                    }
                }
            )
        evidence_by_key: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
        for item in evidence_rows:
            try:
                step = _step_index(item.get("step_index"))
            except (TypeError, ValueError):
                continue
            if step is None:
                continue
            evidence_by_key[(str(item.get("scenario_id") or ""), step)].append(item)
        used_transactions: set[str] = set()
        raw_headers, raw_rows = _sheet_rows(workbook, "raw")
        raw_index = {name: index for index, name in enumerate(raw_headers)}
        observations: list[dict[str, Any]] = []
        for row_number, values in enumerate(raw_rows, 2):
            def value(name: str, default: Any = None) -> Any:
                index = raw_index.get(name)
                return values[index] if index is not None and index < len(values) else default

            scenario_id = str(value("scenario_id") or "")
            try:
                step_index = _step_index(value("step_index"))
            except (TypeError, ValueError):
                continue
            if step_index is None:
                continue
            resource_id = str(
                value("actual_focus_resource_id") or value("focus_view_id") or ""
            )
            node = parse_json_object(value("focus_node"))
            candidates = evidence_by_key.get((scenario_id, step_index), [])
            evidence = next(
                (
                    item
                    for item in candidates
                    if item.get("transaction_id") not in used_transactions
                    and (
                        not resource_id
                        or str(item.get("resource_id") or "") == resource_id
                    )
                ),
                None,
            )
            if evidence is None:
                evidence = next(
                    (
                        item
                        for item in candidates
                        if item.get("transaction_id") not in used_transactions
                    ),
                    {},
                )
            if evidence.get("transaction_id"):
                used_transactions.add(str(evidence["transaction_id"]))
            quality_candidates = quality.get((scenario_id, step_index), [])
            quality_row = next(
                (
                    item
                    for item in quality_candidates
                    if not resource_id
                    or str(item.get("focus_view_id") or "") == resource_id
                ),
                quality_candidates[0] if quality_candidates else {},
            )
            raw = dict(evidence)
            raw.update(
                {
                    "scenario_id": scenario_id,
                    "step_index": step_index,
                    "request_id": str(
                        value("smart_nav_req_id")
                        or evidence.get("request_id")
                        or ""
                    ),
                    "action_type": str(
                        evidence.get("action_type")
                        or ("SMART_NEXT" if value("smart_nav_req_id") else value("move_result") or "")
                    ),
                    "terminal": bool(value("stop_triggered") or value("last_smart_nav_terminal")),
                    "node": node or evidence.get("node") or {},
                    "resource_id": resource_id or evidence.get("resource_id"),
                    "class_name": value("focus_class_name")
                    or node.get("className")
                    or evidence.get("class_name"),
                    "bounds": value("actual_focus_bounds")
                    or value("focus_bounds")
                    or evidence.get("bounds"),
                    "visible_text": quality_row.get("visible_label")
                    if "visible_label" in quality_row
                    else value("visible_label"),
                    "content_description": value("focus_content_description")
                    or node.get("contentDescription"),
                    "talkback_speech": quality_row.get("merged_announcement")
                    if "merged_announcement" in quality_row
                    else value("merged_announcement"),
                    "announcement": quality_row.get("merged_announcement")
                    if "merged_announcement" in quality_row
                    else value("merged_announcement")
                    or evidence.get("announcement"),
                    "mismatch_type": quality_row.get("mismatch_type") or "",
                    "raw_result": quality_row.get("final_result")
                    or value("final_result")
                    or "",
                    "progress_verdict": value("traversal_result")
                    or evidence.get("progress_verdict")
                    or "",
                    "stop_reason": value("stop_reason") or evidence.get("stop_reason") or "",
                    "focusable": value("focus_focusable")
                    if value("focus_focusable") is not None
                    else node.get("focusable"),
                    "clickable": value("focus_clickable")
                    if value("focus_clickable") is not None
                    else node.get("clickable"),
                    "enabled": node.get("enabled"),
                    "selected": node.get("selected"),
                    "checked": node.get("checked"),
                    "scrollable": node.get("scrollable"),
                    "accessibility_focused": node.get("accessibilityFocused"),
                    "duplicate_of_step": (
                        value("recent_semantic_duplicate_of_step")
                        if value("is_recent_semantic_duplicate_step")
                        else value("recent_duplicate_of_step")
                        if value("is_recent_duplicate_step") or value("is_duplicate_step")
                        else None
                    ),
                    "xlsx_row_number": row_number,
                }
            )
            if (
                raw.get("resource_id")
                or raw.get("visible_text")
                or raw.get("talkback_speech")
                or raw.get("bounds")
            ):
                observations.append(raw)
        for evidence in evidence_rows:
            if str(evidence.get("transaction_id") or "") not in used_transactions:
                observations.append(dict(evidence))
        return observations, "talkback-xlsx-raw-result-v1"
    finally:
        workbook.close()


def _read_coverage(path: Path) -> tuple[list[dict[str, Any]], str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ObservationArtifactError(
            ObservationAvailability.CORRUPT, "COVERAGE_CORRUPT"
        ) from exc
    if not isinstance(payload, Mapping) or payload.get("schema_version") != _COVERAGE_SCHEMA:
        raise ObservationArtifactError(
            ObservationAvailability.UNSUPPORTED_SCHEMA,
            "COVERAGE_SCHEMA_UNSUPPORTED",
            actual=payload.get("schema_version") if isinstance(payload, Mapping) else None,
        )
    records = payload.get("records")
    return (
        [dict(item) for item in records if isinstance(item, Mapping)]
        if isinstance(records, list)
        else [],
        _COVERAGE_SCHEMA,
    )


def _read_inventory(path: Path) -> tuple[list[dict[str, Any]], str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ObservationArtifactError(
            ObservationAvailability.CORRUPT, "INVENTORY_CORRUPT"
        ) from exc
    if not isinstance(payload, Mapping) or payload.get("schema_version") != _INVENTORY_SCHEMA:
        raise ObservationArtifactError(
            ObservationAvailability.UNSUPPORTED_SCHEMA,
            "INVENTORY_SCHEMA_UNSUPPORTED",
            actual=payload.get("schema_version") if isinstance(payload, Mapping) else None,
        )
    items = payload.get("items")
    return (
        [dict(item) for item in items if isinstance(item, Mapping)]
        if isinstance(items, list)
        else [],
        _INVENTORY_SCHEMA,
    )


def _coverage_signature(record: Mapping[str, Any]) -> str:
    return canonical_sha256(
        {
            "scenario_id": record.get("scenario_id"),
            "canonical_id": record.get("canonical_id"),
            "taxonomy": record.get("taxonomy"),
        }
    )


def _attach_coverage(
    rows: list[dict[str, Any]],
    coverage: list[dict[str, Any]],
) -> None:
    by_scenario_resource: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in coverage:
        by_scenario_resource[
            (
                str(record.get("scenario_id") or ""),
                str(record.get("view_id") or ""),
            )
        ].append(record)
    for row in rows:
        key = (
            str(row.get("scenario_id") or ""),
            str(row.get("resource_id") or ""),
        )
        candidates = by_scenario_resource.get(key, [])
        if not candidates:
            continue
        row_bounds = parse_bounds(row.get("bounds"))
        match = next(
            (
                item
                for item in candidates
                if row_bounds is not None and parse_bounds(item.get("bounds")) == row_bounds
            ),
            candidates[0],
        )
        if match.get("taxonomy") != "IGNORE":
            row["coverage_signature"] = _coverage_signature(match)
            row["coverage_status"] = str(match.get("coverage_status") or "")


def _coverage_only_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "scenario_id": item.get("scenario_id"),
            "step_index": item.get("matched_step"),
            "resource_id": item.get("view_id"),
            "class_name": item.get("class_name"),
            "bounds": item.get("bounds"),
            "visible_text": item.get("label"),
            "coverage_signature": _coverage_signature(item),
            "coverage_status": item.get("coverage_status"),
            "raw_result": "",
        }
        for item in records
        if item.get("taxonomy") != "IGNORE"
    ]


def _limitation_only_rows(source: ComparatorInput) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in source.reviewed_limitations:
        signature = item.get("match_signature")
        signature = signature if isinstance(signature, Mapping) else {}
        rows.append(
            {
                "scenario_id": item.get("scenario_id"),
                "resource_id": signature.get("resource_id") or item.get("resource_id"),
                "class_name": signature.get("class") or item.get("class"),
                "bounds": signature.get("bounds") or item.get("bounds"),
                "mismatch_type": signature.get("mismatch_type"),
                "raw_result": "FAIL" if item.get("raw_fail_retained") else "",
            }
        )
    return rows


def load_observation_set(
    source: ComparatorInput,
    *,
    qa_runs_root: str | Path | None = None,
    artifact_root: str | Path | None = None,
    viewport: tuple[int, int] = (1080, 2640),
    dynamic_device_names: Iterable[str] = (),
) -> ObservationSet:
    qa_root = Path(qa_runs_root) if qa_runs_root is not None else None
    cas_root = Path(artifact_root) if artifact_root is not None else None
    entries = _artifact_entries(source)
    paths: dict[str, Path] = {}
    public_artifacts: list[dict[str, Any]] = []
    diagnostics: list[dict[str, Any]] = []
    corrupt_status: ObservationAvailability | None = None
    for artifact_type in _OBSERVATION_TYPES:
        entry = entries.get(artifact_type)
        if entry is None:
            continue
        path, public = _resolve_artifact(
            entry,
            qa_runs_root=qa_root,
            artifact_root=cas_root,
        )
        public_artifacts.append(public)
        if path is not None:
            paths[artifact_type] = path
        elif public.get("availability") == "CORRUPT":
            corrupt_status = ObservationAvailability.CORRUPT
            diagnostics.append(
                {
                    "code": public.get("reason"),
                    "artifact_type": artifact_type,
                }
            )

    evidence_rows: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []
    inventory_rows: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    schemas: dict[str, str] = {}
    try:
        if "evidence_ledger" in paths:
            evidence_rows, schemas["evidence_ledger"] = _read_evidence(
                paths["evidence_ledger"]
            )
        if "xlsx" in paths:
            rows, schemas["xlsx"] = _read_xlsx(paths["xlsx"], evidence_rows)
        elif evidence_rows:
            rows = evidence_rows
        if "focusable_coverage" in paths:
            coverage_rows, schemas["focusable_coverage"] = _read_coverage(
                paths["focusable_coverage"]
            )
        if "focusable_inventory" in paths:
            inventory_rows, schemas["focusable_inventory"] = _read_inventory(
                paths["focusable_inventory"]
            )
    except ObservationArtifactError as exc:
        diagnostics.append(exc.to_dict())
        corrupt_status = exc.status

    if rows and coverage_rows:
        _attach_coverage(rows, coverage_rows)
    if not rows and coverage_rows:
        rows = _coverage_only_rows(coverage_rows)
    if not rows and inventory_rows:
        rows = [
            {
                "scenario_id": item.get("scenario_id"),
                "step_index": item.get("step_index"),
                "resource_id": item.get("view_id"),
                "class_name": item.get("class_name"),
                "bounds": item.get("bounds"),
                "visible_text": item.get("label"),
                "clickable": item.get("clickable"),
                "focusable": item.get("focusable"),
                "enabled": item.get("enabled"),
                "selected": item.get("selected"),
                "checked": item.get("checked"),
            }
            for item in inventory_rows
        ]
    if not rows and source.reviewed_limitations:
        rows = _limitation_only_rows(source)

    evidence_artifact = next(
        (
            item
            for item in public_artifacts
            if item.get("artifact_type") == "evidence_ledger"
            and item.get("availability") == "AVAILABLE"
        ),
        None,
    )
    xlsx_artifact = next(
        (
            item
            for item in public_artifacts
            if item.get("artifact_type") == "xlsx"
            and item.get("availability") == "AVAILABLE"
        ),
        None,
    )
    coverage_artifact = next(
        (
            item
            for item in public_artifacts
            if item.get("artifact_type") == "focusable_coverage"
            and item.get("availability") == "AVAILABLE"
        ),
        None,
    )
    provenance_by_type = {
        str(item.get("artifact_type")): {
            "source_artifact": item.get("artifact_type"),
            "artifact_digest": item.get("artifact_digest"),
            "logical_reference": item.get("logical_reference"),
            "schema_version": schemas.get(str(item.get("artifact_type")))
            or item.get("schema_version"),
        }
        for item in public_artifacts
        if item.get("availability") == "AVAILABLE"
    }
    observation_provenance = tuple(
        provenance_by_type[name]
        for name in ("evidence_ledger", "xlsx", "focusable_coverage")
        if name in provenance_by_type
    )
    observations = tuple(
        build_observation(
            row,
            locale=str(source.environment.get("locale") or ""),
            provenance=(
                *observation_provenance,
                {
                    "row_record_locator": (
                        f"raw!row={row['xlsx_row_number']}"
                        if row.get("xlsx_row_number")
                        else f"transaction={row.get('transaction_id')}"
                        if row.get("transaction_id")
                        else f"scenario={row.get('scenario_id')}"
                    )
                },
            ),
            viewport=viewport,
            dynamic_device_names=dynamic_device_names,
        )
        for row in rows
    )
    if corrupt_status is not None:
        availability = corrupt_status
        source_quality = "CORRUPT_SOURCE"
    elif evidence_artifact and xlsx_artifact and observations:
        availability = ObservationAvailability.COMPLETE
        source_quality = "EVIDENCE_XLSX_VERIFIED"
    elif observations:
        availability = ObservationAvailability.PARTIAL
        source_quality = (
            "EVIDENCE_ONLY"
            if evidence_artifact
            else "XLSX_ONLY"
            if xlsx_artifact
            else "COVERAGE_OR_LIMITATION_ONLY"
        )
    else:
        availability = ObservationAvailability.UNAVAILABLE
        source_quality = "NO_OBSERVATION_SOURCE"
    identity_source = {
        "observation_set_schema": OBSERVATION_SET_SCHEMA_VERSION,
        "source_id": source.source_id,
        "locale": source.environment.get("locale"),
        "artifact_digests": sorted(
            {
                str(item.get("artifact_digest"))
                for item in public_artifacts
                if item.get("artifact_digest")
                and item.get("availability") == "AVAILABLE"
            }
        ),
        "observations": [item.to_dict() for item in observations],
    }
    digest = canonical_sha256(identity_source) if observations else None
    return ObservationSet(
        observation_set_schema=OBSERVATION_SET_SCHEMA_VERSION,
        source_kind=source.source_kind.value,
        source_id=source.source_id,
        locale=str(source.environment.get("locale") or ""),
        app_package=str(source.environment.get("app_package") or ""),
        app_version_name=(
            str(source.environment.get("app_version_name"))
            if source.environment.get("app_version_name") is not None
            else None
        ),
        app_version_code=(
            int(source.environment["app_version_code"])
            if source.environment.get("app_version_code") is not None
            else None
        ),
        availability=availability,
        source_quality=source_quality,
        observations=observations,
        artifacts=tuple(public_artifacts),
        observation_identity_digest=digest,
        diagnostics=tuple(diagnostics),
    )


__all__ = [
    "ObservationArtifactError",
    "load_observation_set",
]
