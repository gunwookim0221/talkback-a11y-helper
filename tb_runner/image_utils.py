import hashlib
import shutil
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from talkback_lib import A11yAdbClient
from tb_runner.constants import ENABLE_FULL_SCREEN_EVIDENCE, ENABLE_IMAGE_CROP, ENABLE_IMAGE_INSERT_TO_EXCEL
from tb_runner.logging_utils import log
from tb_runner.utils import parse_bounds_str, sanitize_filename

EXCEL_IMAGE_THUMBNAIL_VERSION = "1.2.0"


def create_excel_thumbnail(
    image_path: str | Path,
    *,
    max_width: int = 160,
    max_height: int = 96,
    as_bytes: bool = False,
) -> str | bytes | None:
    try:
        path = Path(image_path)
        if not path.exists():
            return None
        with Image.open(path) as src:
            thumb = src.copy()
            thumb.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            if as_bytes:
                from io import BytesIO

                output = BytesIO()
                thumb.save(output, format="PNG", optimize=True)
                thumb.close()
                return output.getvalue()
            with tempfile.NamedTemporaryFile(suffix=".png", prefix="excel_thumb_", delete=False) as temp_file:
                temp_path = temp_file.name
            thumb.save(temp_path, format="PNG", optimize=True)
            thumb.close()
            return temp_path
    except Exception:
        return None


def capture_full_screenshot(client: A11yAdbClient, dev: str, save_path: str) -> None:
    # talkback_lib 내부 private helper 재사용
    client._take_snapshot(dev, save_path)


def crop_image_by_bounds(
    screenshot_path: str,
    bounds_str: str,
    crop_path: str,
    shrink_px: int = 0,
) -> bool:
    bounds = parse_bounds_str(bounds_str)
    if not bounds:
        return False

    left, top, right, bottom = bounds
    with Image.open(screenshot_path) as img:
        width, height = img.size

        left = max(0, left + shrink_px)
        top = max(0, top + shrink_px)
        right = min(width, right - shrink_px)
        bottom = min(height, bottom - shrink_px)

        if right <= left or bottom <= top:
            return False

        cropped = img.crop((left, top, right, bottom))
        Path(crop_path).parent.mkdir(parents=True, exist_ok=True)
        cropped.save(crop_path)
        cropped.close()
    return True


def maybe_capture_focus_crop(
    client: A11yAdbClient,
    dev: str,
    row: dict,
    output_base_dir: str,
) -> dict:
    row["t_before_crop"] = round(time.monotonic() - float(row.get("_step_mono_start", time.monotonic())), 3) if row.get("_step_mono_start") else 0.0
    row["crop_image_path"] = ""
    row["crop_image_saved"] = False
    row["crop_bounds"] = str(row.get("focus_bounds", "") or "").strip()
    row["crop_source"] = str(row.get("crop_source", "") or "focus_bounds")
    row["crop_focus_confidence_low"] = False
    row["full_screenshot_path"] = ""
    row["full_screenshot_saved"] = False
    row["full_screenshot_width"] = 0
    row["full_screenshot_height"] = 0
    row["full_screenshot_capture_timestamp"] = 0.0
    row["full_screenshot_correlation_id"] = ""

    if not ENABLE_IMAGE_CROP and not ENABLE_FULL_SCREEN_EVIDENCE:
        row["t_after_crop"] = row["t_before_crop"]
        return row

    bounds_str = str(row.get("focus_bounds", "") or "").strip()
    if not bounds_str and not ENABLE_FULL_SCREEN_EVIDENCE:
        row["t_after_crop"] = row["t_before_crop"]
        return row

    tab_name = sanitize_filename(str(row.get("tab_name", "unknown")))
    step_index = row.get("step_index", -1)
    visible_label = sanitize_filename(str(row.get("visible_label", "") or "")[:40])

    scenario_id = sanitize_filename(str(row.get("scenario_id", "unknown")))
    context_type = sanitize_filename(str(row.get("context_type", "main")))
    correlation_seed = f"{scenario_id}:{context_type}:{step_index}:{row.get('parent_step_index', '')}"
    correlation_id = hashlib.sha256(correlation_seed.encode("utf-8")).hexdigest()[:12]
    crop_path = Path(output_base_dir) / "crops" / f"{scenario_id}_{tab_name}_step_{step_index}_{visible_label}_{correlation_id}.png"
    full_path = Path(output_base_dir) / "screens" / f"{scenario_id}_{context_type}_{step_index}_{correlation_id}.full.png"

    capture_started = time.perf_counter()
    screenshot_path = ""
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".png",
            prefix=f"tb_step_{step_index}_",
            delete=False,
        ) as temp_file:
            screenshot_path = temp_file.name
        capture_full_screenshot(client, dev, screenshot_path)
        if ENABLE_FULL_SCREEN_EVIDENCE:
            full_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(screenshot_path, full_path)
            full_image = Image.open(full_path)
            width, height = full_image.size
            close = getattr(full_image, "close", None)
            if callable(close):
                close()
            row["full_screenshot_path"] = str(full_path)
            row["full_screenshot_saved"] = True
            row["full_screenshot_width"] = width
            row["full_screenshot_height"] = height
            row["full_screenshot_capture_timestamp"] = round(time.time(), 3)
            row["full_screenshot_correlation_id"] = correlation_id
        if ENABLE_IMAGE_CROP and bounds_str:
            ok = crop_image_by_bounds(
                screenshot_path=screenshot_path,
                bounds_str=bounds_str,
                crop_path=str(crop_path),
                shrink_px=2,
            )
            if ok:
                row["crop_image_path"] = str(crop_path)
                row["crop_image_saved"] = True
        row["screenshot_capture_elapsed"] = round(time.perf_counter() - capture_started, 3)
    except (AttributeError, OSError, RuntimeError, TypeError) as exc:
        log(f"[IMAGE] crop failed step={step_index}: {exc}")
        if full_path.is_file() and not row["full_screenshot_saved"]:
            full_path.unlink(missing_ok=True)
    finally:
        if screenshot_path:
            try:
                Path(screenshot_path).unlink(missing_ok=True)
            except Exception:
                pass
        row["crop_elapsed_sec"] = round(time.perf_counter() - capture_started, 3)
        if row.get("_step_mono_start"):
            row["t_after_crop"] = round(time.monotonic() - float(row["_step_mono_start"]), 3)
        else:
            row["t_after_crop"] = row.get("t_before_crop", 0.0)
        payload_source = str(row.get("focus_payload_source", "") or "").lower()
        response_success = bool(row.get("get_focus_response_success", False))
        focus_view_id = str(row.get("focus_view_id", "") or "").strip()
        row["crop_focus_confidence_low"] = bool(
            (payload_source == "top_level" and not response_success)
            or (not focus_view_id and bool(row.get("crop_bounds", "")))
        )

    return row


def insert_images_to_excel(
    excel_path: str,
    image_col_name: str = "crop_image",
    sheet_name: str = "raw",
) -> None:
    if not ENABLE_IMAGE_INSERT_TO_EXCEL:
        return

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    headers = [cell.value for cell in ws[1]]
    if image_col_name not in headers or "crop_image_path" not in headers:
        wb.save(excel_path)
        return

    image_col_idx = headers.index(image_col_name) + 1
    path_col_idx = headers.index("crop_image_path") + 1

    col_letter = ws.cell(row=1, column=image_col_idx).column_letter
    temp_thumb_paths: list[str] = []

    try:
        for row_idx in range(2, ws.max_row + 1):
            path_value = ws.cell(row=row_idx, column=path_col_idx).value
            if not path_value:
                continue

            img_path = Path(str(path_value))
            if not img_path.exists():
                continue

            thumb_path = create_excel_thumbnail(img_path)
            if not thumb_path:
                continue
            temp_thumb_paths.append(thumb_path)

            try:
                img = XLImage(thumb_path)
                ws.add_image(img, f"{col_letter}{row_idx}")
                row_height = (float(getattr(img, "height", 0) or 0) * 0.75) + 6.0
                ws.row_dimensions[row_idx].height = max(float(ws.row_dimensions[row_idx].height or 0), row_height)
            except Exception as exc:
                log(f"[EXCEL] image insert failed row={row_idx}: {exc}")
        ws.column_dimensions[col_letter].width = 24
        wb.save(excel_path)
    finally:
        for thumb_path in temp_thumb_paths:
            try:
                Path(thumb_path).unlink(missing_ok=True)
            except Exception:
                pass
